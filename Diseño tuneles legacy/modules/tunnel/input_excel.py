"""Plantilla e importador de entradas para SALVI Tunnel Engine."""
from __future__ import annotations

from io import BytesIO
from collections.abc import Mapping
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


_BRISCO_NOTE = (
    "Escenario Brisco - Arguayo precargado con los valores de contraste "
    "indicados durante la revisión: longitud 1.016 m, Lth 91 cd/m², "
    "umbral 85 m, transición hasta 430 m y Lin 2 cd/m². "
    "Confirmar contra el proyecto constructivo antes de emitir un cálculo."
)

# grupo, clave que entiende la SPA/API, valor de partida, unidad, descripción
INPUT_FIELDS = [
    ("Proyecto", "project_name", "Túnel Brisco - Arguayo", "", "Nombre del proyecto."),
    ("Proyecto", "tube_id", "T1", "", "Identificador del tubo."),
    ("Geometría", "length_m", 1016, "m", "Longitud interior entre bocas."),
    ("Geometría", "speed_kmh", 80, "km/h", "Velocidad de proyecto. Confirmar."),
    ("Geometría", "gradient_pct", 0, "%", "Pendiente; positivo = bajada."),
    ("Geometría", "curvature_radius_m", "", "m", "Vacío si el eje es recto."),
    ("Geometría", "mu_friction", "", "", "Coeficiente de rozamiento; vacío = valor automático."),
    ("Geometría", "t_reaction", 2.5, "s", "Tiempo de reacción para distancia de parada."),
    ("Geometría", "traffic_direction", "one_way", "", "one_way o two_way."),
    ("Geometría", "width_m", 9, "m", "Anchura física pared a pared."),
    ("Geometría", "height_m", 5.5, "m", "Altura libre en clave. Confirmar."),
    ("Geometría", "tunnel_shape", "horseshoe", "", "horseshoe, circular o rectangular."),
    ("Geometría", "H_pared_m", 3, "m", "Altura del tramo recto de pared."),
    ("Calzada", "num_lanes", 2, "", "Número de carriles."),
    ("Calzada", "lane_width_m", 3.5, "m", "Anchura de cada carril."),
    ("Calzada", "shoulder_left_m", 1, "m", "Arcén izquierdo."),
    ("Calzada", "shoulder_right_m", 1, "m", "Arcén derecho."),
    ("Calzada", "sidewalk_left_m", 0, "m", "Acera o paso técnico izquierdo."),
    ("Calzada", "sidewalk_right_m", 0, "m", "Acera o paso técnico derecho."),
    ("Entorno", "portal_orientation", "S", "", "N, NE, E, SE, S, SW, W o NW."),
    ("Entorno", "environment_type", "open_country_flat", "", "Clasificación del entorno CIE 88."),
    ("Entorno", "sky_condition", "clear", "", "clear, intermediate u overcast."),
    ("Entorno", "daylight_penetration", "poor", "", "poor o good."),
    ("Entorno", "wall_reflectance", 0.4, "", "Reflectancia usada por CIE 88."),
    ("Entorno", "road_surface", "dark_asphalt", "", "Pavimento CIE 144: dark_asphalt, medium_asphalt, light_asphalt, concrete o bright_concrete."),
    ("Entorno", "rho_wall", 0.4, "", "Reflectancia de paredes para radiosidad."),
    ("Entorno", "rho_ceiling", 0.25, "", "Reflectancia de techo para radiosidad."),
    ("Entorno", "wall_luminance_height_m", 2, "m", "Franja de pared evaluada."),
    ("Entorno", "exit_visible", False, "", "TRUE/FALSE."),
    ("Entorno", "illuminated_road", False, "", "TRUE/FALSE."),
    ("Entorno", "lat", "", "°", "Latitud WGS84 opcional."),
    ("Entorno", "lng", "", "°", "Longitud WGS84 opcional."),
    ("Geolocalización OSM", "osm_tunnel_id", "", "", "Identificador OSM del tramo."),
    ("Geolocalización OSM", "osm_tunnel_name", "", "", "Nombre/ref del túnel en OSM."),
    ("Geolocalización OSM", "osm_source", "", "", "Fuente del eje: OSM, topografía o proyecto."),
    ("Geolocalización OSM", "osm_tunnel_length_m", 1016, "m", "Longitud del tramo OSM; se conserva separada de length_m."),
    ("Tráfico", "traffic_veh_h", 500, "veh/h", "Tráfico de diseño."),
    ("Tráfico", "imd", "", "veh/día", "IMD opcional; se combina con k_peak."),
    ("Tráfico", "k_peak", 0.1, "", "Coeficiente de hora de diseño."),
    ("Tráfico", "has_pedestrians", False, "", "TRUE/FALSE."),
    ("CIE 88 / proyecto", "interior_luminance_override", 2, "cd/m²", "Lin de proyecto; vacío = cálculo normativo."),
    ("CIE 88 / proyecto", "lth_override", 91, "cd/m²", "Lth A de proyecto; vacío = cálculo normativo."),
    ("CIE 88 / proyecto", "lth_b_override", "", "cd/m²", "Lth B para dos sentidos."),
    ("CIE 88 / proyecto", "threshold_length_override_m", 85, "m", "Longitud de umbral A de proyecto."),
    ("CIE 88 / proyecto", "threshold_length_b_override_m", "", "m", "Longitud de umbral B."),
    ("CIE 88 / proyecto", "transition_end_override_m", 430, "m", "Fin de transición A desde boca."),
    ("CIE 88 / proyecto", "transition_end_b_override_m", "", "m", "Fin de transición B desde boca."),
    ("CIE 88 / proyecto", "exit_length_override_m", "", "m", "Longitud de salida si se fija por proyecto."),
    ("CIE 88 / proyecto", "exit_luminance_ratio_override", 100, "%", "Objetivo de salida relativo a Lin; 100 = igual al interior."),
    ("CIE 88 / proyecto", "l20_override", "", "cd/m²", "L20 A; vacío = cálculo."),
    ("CIE 88 / proyecto", "l20_b_override", "", "cd/m²", "L20 B; vacío = cálculo."),
    ("CIE 88 / proyecto", "tunnel_class", "auto", "", "auto, 1, 2, 3 o 4."),
    ("CIE 88 / proyecto", "l20_method", "model", "", "model o table."),
    ("CIE 88 / proyecto", "lth_method", "k_factor", "", "k_factor o lseq."),
    ("CIE 88 / proyecto", "lth_standard", "oc36_2015", "", "Norma/método aplicable."),
    ("CIE 88 / proyecto", "stopping_distance_override_m", "", "m", "Distancia de parada A; vacío = cálculo."),
    ("CIE 88 / proyecto", "stopping_distance_b_override_m", "", "m", "Distancia de parada B; vacío = cálculo."),
    ("CIE 88 / proyecto", "k_lth_override", "", "", "Factor k de Lth A; vacío = cálculo."),
    ("CIE 88 / proyecto", "k_lth_b_override", "", "", "Factor k de Lth B; vacío = cálculo."),
    ("CIE 88 / proyecto", "lseq_override", "", "cd/m²", "Lseq A de proyecto; vacío = cálculo."),
    ("CIE 88 / proyecto", "lseq_b_override", "", "cd/m²", "Lseq B de proyecto; vacío = cálculo."),
    ("CIE 88 / proyecto", "qc_override", 0.1, "", "Factor qc de proyecto."),
    ("CIE 88 / proyecto", "contrast_observation", 0.04, "", "Contraste de observación para Lth."),
    ("CIE 88 / proyecto", "profile_stepped", False, "", "TRUE/FALSE para perfil de transición escalonado."),
    ("CIE 88 / proyecto", "n_steps", 4, "", "Número de escalones."),
    ("CIE 88 / proyecto", "n_transition_groups", 2, "", "Grupos de transición."),
    ("CIE 140", "wall_ratio_override", "", "", "Ratio pared/calzada de proyecto; vacío = clase CIE."),
    ("CIE 140", "calc_mode", "direct", "", "direct o radiosity."),
    ("Control", "control_protocol", "DALI", "", "Protocolo de control."),
    ("Control", "control_architecture", "permanent_base_plus_portal_reinforcement", "", "Arquitectura de capas instalada."),
    ("Control", "annual_operation_hours", 8760, "h/año", "Horas anuales para energía."),
    ("Control", "energy_tariff_eur_kwh", 0.15, "€/kWh", "Tarifa energética."),
    ("Luminarias", "luminaire.I_max_mA", 500, "mA", "Corriente máxima."),
    ("Luminarias", "luminaire.cct", "4000K", "", "Temperatura de color."),
    ("Luminarias", "luminaire.optic", "auto", "", "Óptica APHEX o auto."),
    ("Luminarias", "luminaire.mounting_height_m", 4.5, "m", "Altura de montaje."),
    ("Luminarias", "luminaire.wall_offset_m", 0.3, "m", "Coordenada desde pared izquierda."),
    ("Luminarias", "luminaire.axis_offset_m", 0.3, "m", "Desplazamiento de eje, si aplica."),
    ("Luminarias", "luminaire.arrangement", "central_single", "", "Disposición de filas."),
    ("Luminarias", "luminaire.maintenance_factor", 0.7, "", "Factor de mantenimiento."),
    ("Luminarias", "luminaire.road_surface", "dark_asphalt", "", "Pavimento aplicado al cálculo fotométrico."),
    ("Luminarias", "luminaire.U0_obj", 0.4, "", "Objetivo U0."),
    ("Luminarias", "luminaire.Ul_obj", 0.6, "", "Objetivo Ul."),
    ("Luminarias", "luminaire.I_min_pct", 30, "%", "Mínimo de regulación."),
    ("Luminarias", "luminaire.tilt_max", 20, "°", "Inclinación máxima."),
    ("Luminarias", "luminaire.d_fixed", "", "m", "Interdistancia fija; vacío = automática."),
    ("Luminarias", "luminaire.d_min", 1, "m", "Interdistancia mínima."),
    ("Luminarias", "luminaire.optimization_goal", "min_luminaires", "", "min_luminaires o min_power."),
    ("Luminarias", "luminaire.max_luminaire_increase_pct", 15, "%", "Límite de aumento de luminarias."),
    ("Luminarias", "luminaire.max_base_spacing_reduction_pct", 20, "%", "Límite de reducción de interdistancia base."),
    ("Luminarias", "luminaire.spacing_quantum_m", 0.5, "m", "Cuantización de la interdistancia."),
    ("Luminarias", "luminaire.constructive_min_separation_m", 0.5, "m", "Separación constructiva mínima."),
    ("Luminarias", "luminaire.transition_spacing_step_m", 2, "m", "Paso de interdistancia en transición."),
    ("Luminarias", "luminaire.luminance_margin_pct", 4, "%", "Margen de diseño de luminancia."),
    ("Luminarias", "luminaire.daylight_contribution_enabled", False, "", "TRUE/FALSE."),
    ("Luminarias", "luminaire.daylight_portal_a", True, "", "TRUE/FALSE."),
    ("Luminarias", "luminaire.daylight_portal_b", True, "", "TRUE/FALSE."),
    ("Luminarias", "luminaire.daylight_penetration_length_m", 60, "m", "Longitud modelada de aporte natural."),
    ("Luminarias", "luminaire.daylight_mouth_contribution_pct", 10, "%", "Aporte natural en boca."),
    ("Luminarias", "luminaire.daylight_decay_exponent", 1, "", "Exponente de decaimiento natural."),
]

_DROPDOWN_OPTIONS = {
    "traffic_direction": [("one_way", "Un sentido"), ("two_way", "Bidireccional")],
    "tunnel_shape": [("horseshoe", "Herradura / bóveda"), ("circular", "Circular"), ("rectangular", "Rectangular")],
    "portal_orientation": [(code, code) for code in ("N", "NE", "E", "SE", "S", "SW", "W", "NW")],
    "sky_condition": [("clear", "Despejado"), ("intermediate", "Parcial"), ("overcast", "Cubierto")],
    "daylight_penetration": [("poor", "Pobre"), ("good", "Buena")],
    "road_surface": [(code, label) for code, label in (
        ("dark_asphalt", "Asfalto oscuro R3"), ("medium_asphalt", "Asfalto medio R2"),
        ("light_asphalt", "Asfalto claro R1"), ("concrete", "Hormigón C1"),
        ("bright_concrete", "Hormigón claro C2"),
    )],
    "l20_method": [("model", "Modelo CIE 88"), ("table", "Tabla CIE 88")],
    "lth_method": [("k_factor", "Factor k"), ("lseq", "Lseq")],
    "lth_standard": [("oc36_2015", "OC 36/2015"), ("cie88", "CIE 88")],
    "tunnel_class": [("auto", "Automática"), ("1", "Clase 1"), ("2", "Clase 2"), ("3", "Clase 3"), ("4", "Clase 4")],
    "calc_mode": [("direct", "Directo"), ("radiosity", "Radiosidad")],
    "control_protocol": [("DALI", "DALI"), ("DALI-continuous", "DALI continuo"), ("SmartEC", "SmartEC continuo")],
    "control_architecture": [("permanent_base_plus_portal_reinforcement", "Base + refuerzo portal"), ("legacy_zonal", "Zonal legacy")],
    "cct": [("3000K", "3000 K"), ("4000K", "4000 K")],
    "optic": [("auto", "Automática"), ("F151", "F151"), ("F2M2", "F2M2"), ("F2MD", "F2MD")],
    "arrangement": [(code, label) for code, label in (
        ("central_single", "Fila central única"), ("central_offset", "Fila central desplazada"),
        ("central_double", "Doble fila central"), ("lateral_left", "Lateral izquierda"),
        ("lateral_right", "Lateral derecha"), ("bilateral_sym", "Bilateral simétrico"),
        ("bilateral_stag", "Bilateral tresbolillo"), ("bilateral", "Bilateral"),
        ("unilateral", "Unilateral"),
    )],
    "optimization_goal": [("min_luminaires", "Mínimo de luminarias"), ("min_power", "Mínima potencia")],
}

_BOOL_KEYS = {
    "exit_visible", "illuminated_road",
    "has_pedestrians", "profile_stepped", "daylight_contribution_enabled",
    "daylight_portal_a", "daylight_portal_b",
}


def _coerce_value(key: str, value: Any) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    base_key = key.split(".")[-1]
    if base_key in _BOOL_KEYS:
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"1", "true", "verdadero", "si", "sí", "yes", "y", "x", "on"}:
            return True
        if normalized in {"0", "false", "falso", "no", "n", "off"}:
            return False
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    return str(value).strip()


def create_tunnel_input_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos de entrada"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = "SALVI Tunnel Engine - Plantilla de entradas"
    ws["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1A3A6B")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:E2")
    ws["A2"] = _BRISCO_NOTE
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="5B4A18")
    ws["A2"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34
    headers = ["Grupo", "Campo", "Valor", "Unidad", "Descripción"]
    for col, label in enumerate(headers, 1):
        cell = ws.cell(4, col, label)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A56B0")
        cell.alignment = Alignment(horizontal="center")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for row, (group, key, value, unit, description) in enumerate(INPUT_FIELDS, 5):
        # CÃ³digos booleanos universales, visibles tambiÃ©n fuera de Excel en
        # espaÃ±ol y compatibles con la lista desplegable TRUE/FALSE.
        cell_value = (
            "TRUE" if value is True else "FALSE" if value is False else value
        )
        values = [group, key, cell_value, unit, description]
        for col, item in enumerate(values, 1):
            cell = ws.cell(row, col, item)
            cell.font = Font(name="Arial", size=10, color="0000FF" if col == 3 else "1F1F1F")
            cell.fill = PatternFill("solid", fgColor="FFF2CC" if col == 3 else "FFFFFF")
            cell.alignment = Alignment(vertical="top", wrap_text=col == 5)
            cell.border = border
        ws.cell(row, 3).comment = Comment(
            "Celda editable. Dejar vacía para no modificar ese dato al importar.",
            "SALVI",
        )
    for column, width in {"A": 20, "B": 38, "C": 18, "D": 12, "E": 62}.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{4 + len(INPUT_FIELDS)}"

    # Listas visibles para que el usuario pueda consultar códigos y etiquetas.
    lists = wb.create_sheet("Listas de opciones")
    lists.sheet_view.showGridLines = False
    lists.append(["Campo", "Valor que se importa", "Descripción"])
    for cell in lists[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A56B0")
    option_ranges = {}
    for key, options in _DROPDOWN_OPTIONS.items():
        start = lists.max_row + 1
        for value, label in options:
            lists.append([key, value, label])
        option_ranges[key] = (start, lists.max_row)
        range_name = f"salvi_opt_{key.replace('.', '_')}"
        defined = DefinedName(
            range_name,
            attr_text=f"'Listas de opciones'!$B${start}:$B${lists.max_row}",
        )
        if hasattr(wb.defined_names, "add"):
            wb.defined_names.add(defined)
        else:
            wb.defined_names.append(defined)
    lists.column_dimensions["A"].width = 28
    lists.column_dimensions["B"].width = 38
    lists.column_dimensions["C"].width = 52
    lists.freeze_panes = "A2"
    lists.auto_filter.ref = f"A1:C{lists.max_row}"

    # Validación de datos en la columna Valor. Se referencia la hoja visible
    # para que el código elegido siempre coincida con lo que espera la API.
    row_by_key = {key: row for row, (_group, key, *_rest) in enumerate(INPUT_FIELDS, 5)}
    for key, options in _DROPDOWN_OPTIONS.items():
        row_key = key if key in row_by_key else f"luminaire.{key}"
        if row_key not in row_by_key:
            continue
        start, end = option_ranges[key]
        validation = DataValidation(
            type="list",
            formula1=f"=salvi_opt_{key.replace('.', '_')}",
            allow_blank=True,
        )
        validation.error = "Elige un valor de la lista de opciones."
        validation.errorTitle = "Valor no válido"
        validation.prompt = "Selecciona el código que usa SALVI Tunnel Engine."
        validation.promptTitle = "Entrada SALVI"
        ws.add_data_validation(validation)
        validation.add(ws.cell(row_by_key[row_key], 3))
    for key in _BOOL_KEYS:
        row_key = key if key in row_by_key else f"luminaire.{key}"
        if row_key not in row_by_key:
            continue
        validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        validation.error = "Usa TRUE o FALSE."
        validation.errorTitle = "Booleano no válido"
        ws.add_data_validation(validation)
        validation.add(ws.cell(row_by_key[row_key], 3))

    # Eje longitudinal opcional procedente de OSM.
    route = wb.create_sheet("Tramo OSM")
    route.sheet_view.showGridLines = False
    route.merge_cells("A1:D1")
    route["A1"] = "Tramo OSM / eje georreferenciado"
    route["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    route["A1"].fill = PatternFill("solid", fgColor="1A3A6B")
    route["A2"] = "Introduzca una fila por punto del eje. La aplicación importará orden, latitud y longitud."
    route["A2"].font = Font(name="Arial", italic=True, color="5B4A18")
    route["A2"].fill = PatternFill("solid", fgColor="FFF2CC")
    for col, title in enumerate(("Orden", "Latitud WGS84", "Longitud WGS84", "Cota (m)"), 1):
        cell = route.cell(4, col, title)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A56B0")
    route.cell(5, 1, 1)
    route.cell(5, 2, "")
    route.cell(5, 3, "")
    route.cell(5, 4, "")
    for col, width in {"A": 12, "B": 20, "C": 20, "D": 14}.items():
        route.column_dimensions[col].width = width
    route.freeze_panes = "A5"

    guide = wb.create_sheet("Instrucciones")
    guide.sheet_view.showGridLines = False
    guide["A1"] = "Cómo usar esta plantilla"
    guide["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="1A3A6B")
    guide.merge_cells("A1:B1")
    instructions = [
        ("1", "Edite únicamente la columna Valor (amarilla)."),
        ("2", "Una celda Valor vacía no sobrescribe el valor actual al importar."),
        ("3", "No modifique la columna Campo: es la clave de importación."),
        ("4", "Los campos luminaire.* se cargan en la configuración de luminarias."),
        ("5", _BRISCO_NOTE),
        ("6", "Las aceras no entran en L/U0/Ul; los arcenes sólo lo hacen si se activa su campo específico."),
        ("7", "Los overrides por luminaria (posiciones, flujo, tilt y tándem) se mantienen en su tabla editable; no se reemplazan al importar esta planilla."),
        ("8", "La hoja Tramo OSM permite cargar el eje por puntos (latitud/longitud). Si se deja vacía, se conserva la geolocalización puntual."),
    ]
    for row, (step, text) in enumerate(instructions, 3):
        guide.cell(row, 1, step).font = Font(name="Arial", bold=True, color="1A3A6B")
        guide.cell(row, 2, text).font = Font(name="Arial", size=10)
        guide.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row].height = 28
    guide.column_dimensions["A"].width = 8
    guide.column_dimensions["B"].width = 110

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def create_tunnel_input_workbook(form: Mapping[str, Any] | None = None) -> bytes:
    """Genera un XLSX con la configuraciÃ³n vigente de un tubo.

    ``create_tunnel_input_template`` sigue siendo la plantilla base con el
    caso Brisco--Arguayo. Esta funciÃ³n la rellena con el formulario actual y
    conserva ademÃ¡s el eje OSM completo en la hoja ``Tramo OSM``.
    """
    if not isinstance(form, Mapping):
        return create_tunnel_input_template()

    workbook = load_workbook(BytesIO(create_tunnel_input_template()))
    worksheet = workbook["Datos de entrada"]
    worksheet["A1"] = "SALVI Tunnel Engine - ConfiguraciÃ³n actual"
    worksheet["A2"] = (
        "ExportaciÃ³n del estado vigente del tubo. Edite la columna Valor y vuelva a importar si necesita ajustar entradas."
    )
    luminaire = form.get("lum_config")
    if not isinstance(luminaire, Mapping):
        luminaire = form.get("luminaire")
    if not isinstance(luminaire, Mapping):
        luminaire = {}

    for row, (_group, key, default, _unit, _description) in enumerate(
        INPUT_FIELDS, 5
    ):
        if key.startswith("luminaire."):
            field_key = key.split(".", 1)[1]
            value = luminaire[field_key] if field_key in luminaire else default
        else:
            # La exportaciÃ³n representa el estado real, no vuelve a rellenar
            # opcionales ausentes con los valores de la plantilla Brisco.
            value = form[key] if key in form else ""
        if value is None:
            value = ""
        elif isinstance(value, bool):
            value = "TRUE" if value else "FALSE"
        worksheet.cell(row, 3).value = value

    route = workbook["Tramo OSM"]
    # La plantilla contiene una fila vacÃ­a de ejemplo; se elimina antes de
    # escribir el eje actual para que una exportaciÃ³n no arrastre puntos.
    if route.max_row >= 5:
        route.delete_rows(5, route.max_row - 4)
    geometry = form.get("osm_tunnel_geometry")
    if not isinstance(geometry, (list, tuple)):
        geometry = []
    for index, point in enumerate(geometry, 1):
        if not isinstance(point, Mapping):
            continue
        lat = point.get("lat")
        lng = point.get("lng", point.get("lon"))
        if lat in (None, "") or lng in (None, ""):
            continue
        row = route.max_row + 1
        route.cell(row, 1).value = index
        route.cell(row, 2).value = lat
        route.cell(row, 3).value = lng
        elevation = point.get("elevation_m", point.get("elevation"))
        if elevation not in (None, ""):
            route.cell(row, 4).value = elevation

    workbook.properties.title = "SALVI Tunnel Engine - ConfiguraciÃ³n actual"
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def parse_tunnel_input_workbook(stream) -> dict:
    wb = load_workbook(stream, read_only=True, data_only=False)
    if "Datos de entrada" not in wb.sheetnames:
        raise ValueError("Falta la hoja 'Datos de entrada' de la plantilla SALVI.")
    ws = wb["Datos de entrada"]
    header = {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(next(ws.iter_rows(min_row=4, max_row=4)), 1)
        if cell.value is not None
    }
    key_col = header.get("campo")
    value_col = header.get("valor")
    if not key_col or not value_col:
        raise ValueError("La hoja debe contener las columnas 'Campo' y 'Valor'.")

    allowed = {item[1] for item in INPUT_FIELDS}
    form: dict[str, Any] = {}
    luminaire: dict[str, Any] = {}
    ignored: list[str] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if len(row) < max(key_col, value_col):
            continue
        key = row[key_col - 1]
        if key is None:
            continue
        key = str(key).strip()
        if key not in allowed:
            ignored.append(key)
            continue
        value = _coerce_value(key, row[value_col - 1])
        if value is None:
            continue
        if key.startswith("luminaire."):
            luminaire[key.split(".", 1)[1]] = value
        else:
            form[key] = value
    if "Tramo OSM" in wb.sheetnames:
        route = wb["Tramo OSM"]
        route_points = []
        for row in route.iter_rows(min_row=5, values_only=True):
            if len(row) < 3 or row[1] in (None, "") or row[2] in (None, ""):
                continue
            try:
                point = {"lat": float(row[1]), "lng": float(row[2])}
                if len(row) > 3 and row[3] not in (None, ""):
                    point["elevation_m"] = float(row[3])
                route_points.append((float(row[0] or len(route_points) + 1), point))
            except (TypeError, ValueError):
                continue
        if route_points:
            form["osm_tunnel_geometry"] = [
                point for _order, point in sorted(route_points, key=lambda item: item[0])
            ]
    if luminaire:
        form["lum_config"] = luminaire
    if not form:
        raise ValueError("No hay valores importables en la columna 'Valor'.")
    return {"form": form, "ignored_fields": ignored}
