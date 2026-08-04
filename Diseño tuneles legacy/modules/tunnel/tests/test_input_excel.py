from io import BytesIO

from openpyxl import load_workbook

from modules.tunnel.input_excel import (
    create_tunnel_input_workbook,
    parse_tunnel_input_workbook,
)


def test_current_configuration_roundtrip_includes_osm_and_booleans():
    form = {
        "tube_id": "T2",
        "project_name": "Túnel de prueba",
        "length_m": 1234,
        "lat": 42.12345,
        "lng": -0.23456,
        "exit_visible": True,
        "illuminated_road": False,
        "lum_config": {
            "arrangement": "bilateral_sym",
            "daylight_portal_a": False,
        },
        "osm_tunnel_id": "way/99",
        "osm_tunnel_length_m": 1234,
        "osm_tunnel_geometry": [
            {"lat": 42.1, "lng": -0.2, "elevation_m": 100},
            {"lat": 42.2, "lng": -0.3},
        ],
    }

    raw = create_tunnel_input_workbook(form)
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=False)
    worksheet = workbook["Datos de entrada"]
    values = {
        worksheet.cell(row, 2).value: worksheet.cell(row, 3).value
        for row in range(5, worksheet.max_row + 1)
    }
    assert values["exit_visible"] == "TRUE"
    assert values["illuminated_road"] == "FALSE"
    assert values["luminaire.daylight_portal_a"] == "FALSE"

    route = workbook["Tramo OSM"]
    assert route.max_row == 6
    assert route.cell(5, 2).value == 42.1
    assert route.cell(5, 3).value == -0.2

    parsed = parse_tunnel_input_workbook(BytesIO(raw))["form"]
    assert parsed["lat"] == 42.12345
    assert parsed["lng"] == -0.23456
    assert parsed["exit_visible"] is True
    assert parsed["illuminated_road"] is False
    assert parsed["lum_config"]["daylight_portal_a"] is False
    assert parsed["osm_tunnel_geometry"] == [
        {"lat": 42.1, "lng": -0.2, "elevation_m": 100.0},
        {"lat": 42.2, "lng": -0.3},
    ]


def test_boolean_import_accepts_english_and_spanish_values():
    form = {"exit_visible": "Verdadero", "illuminated_road": "Falso"}
    parsed = parse_tunnel_input_workbook(
        BytesIO(create_tunnel_input_workbook(form))
    )["form"]
    assert parsed["exit_visible"] is True
    assert parsed["illuminated_road"] is False
