#!/usr/bin/env python3
"""SALVI GIS - API Backend. Pure Python stdlib. Port 8733."""
import sqlite3, json, os, re, io, time, math, uuid, hmac, hashlib, base64 as _b64, secrets as _sec
from contextlib import contextmanager
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

# ── Load .env file if present ──────────────────────────────────────────────────
_ENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_ENV_PATH):
    with open(_ENV_PATH, encoding="utf-8") as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _, _v = _line.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip().strip('"').strip("'"))

# ── Constants ──────────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "db", "salvi_gis.db")

DEFAULT_COLORS = [
    '#4caf82','#e67e22','#3498db','#9b59b6','#e74c3c',
    '#1abc9c','#f39c12','#2980b9','#8e44ad','#c0392b',
]

_AI_KEY   = os.environ.get("ANTHROPIC_API_KEY", "")
_AI_MODEL = os.environ.get("SALVI_AI_MODEL", "claude-haiku-4-5-20251001")

# Local legacy deployments can run without the login screen. Set this to 0 in
# .env to restore the original user/password authentication.
_AUTH_DISABLED = os.environ.get("SALVI_AUTH_DISABLED", "1").strip().lower() not in ("0", "false", "no", "off")

# ── Auth secret ────────────────────────────────────────────────────────────────
_AUTH_SECRET = os.environ.get("AUTH_SECRET", "")
if not _AUTH_SECRET:
    _AUTH_SECRET = _sec.token_hex(32)
    try:
        with open(_ENV_PATH, 'a', encoding='utf-8') as _ef:
            _ef.write(f"\nAUTH_SECRET={_AUTH_SECRET}\n")
        print("[SALVI] AUTH_SECRET generado y guardado en .env")
    except Exception: pass
_TOKEN_TTL    = 86400 * 7   # 7 días (login normal)
_RESET_TOKENS = {}          # token -> {uid, username, exp}  (in-memory, 1h TTL)

# ── Pending import cache (expires after 30 min) ────────────────────────────────
_pending_imports = {}  # { temp_id: {"rows": [...], "headers": [...], "ts": float} }

# ── DB context manager ─────────────────────────────────────────────────────────
@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

# ── Schema (only for fresh installs — never modifies existing data) ─────────────
_SCHEMA_SQL = [
    """CREATE TABLE IF NOT EXISTS projects (
        id TEXT PRIMARY KEY, name TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now')))""",
    """CREATE TABLE IF NOT EXISTS zones (
        id TEXT PRIMARY KEY, name TEXT NOT NULL,
        type TEXT DEFAULT '', color TEXT DEFAULT '#4caf82', priority INTEGER DEFAULT 2,
        center_lat REAL, center_lon REAL, zoom INTEGER DEFAULT 12,
        bbox TEXT DEFAULT '', description TEXT DEFAULT '',
        est_primary REAL DEFAULT 0, est_secondary REAL DEFAULT 0,
        est_tertiary REAL DEFAULT 0, est_residential REAL DEFAULT 0,
        est_unclassified REAL DEFAULT 0,
        corridors TEXT DEFAULT '[]', bounds_polygon TEXT DEFAULT '[]',
        osm_relation INTEGER DEFAULT NULL,
        source TEXT DEFAULT 'manual',
        created_at TEXT DEFAULT (datetime('now')), project_id TEXT DEFAULT NULL)""",
    """CREATE TABLE IF NOT EXISTS zone_config (
        zone_id TEXT PRIMARY KEY, spacing INTEGER DEFAULT 30,
        watt_hps REAL DEFAULT 150, watt_led REAL DEFAULT 60,
        efficacy REAL DEFAULT 90, hours_night REAL DEFAULT 11.5,
        updated_at TEXT DEFAULT (datetime('now')))""",
    """CREATE TABLE IF NOT EXISTS zone_osm_data (
        zone_id TEXT PRIMARY KEY, km_by_type TEXT DEFAULT '{}',
        ways TEXT DEFAULT '[]', source TEXT DEFAULT 'estimated',
        loaded_at TEXT DEFAULT (datetime('now')))""",
    """CREATE TABLE IF NOT EXISTS luminaires (
        id INTEGER PRIMARY KEY AUTOINCREMENT, project_id TEXT,
        zone_id TEXT NOT NULL, road_type TEXT, lighting_class TEXT,
        street_name TEXT, lat REAL NOT NULL, lon REAL NOT NULL,
        watts REAL, spacing REAL, placed_at TEXT DEFAULT (datetime('now')))""",
    """CREATE TABLE IF NOT EXISTS inventory_luminaires (
        id INTEGER PRIMARY KEY AUTOINCREMENT, zone_id TEXT NOT NULL,
        point_id TEXT, lat REAL NOT NULL, lon REAL NOT NULL,
        power_w REAL, height_m REAL, brand TEXT, model TEXT,
        lamp_type TEXT, support_type TEXT, circuit_id TEXT, line_id TEXT,
        extra TEXT DEFAULT '{}', way_key TEXT, road_type TEXT,
        imported_at TEXT DEFAULT (datetime('now')))""",
    """CREATE TABLE IF NOT EXISTS project_ui_config (
        project_id TEXT NOT NULL,
        config_key TEXT NOT NULL,
        config_value TEXT DEFAULT '{}',
        updated_at TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (project_id, config_key))""",
        """CREATE TABLE IF NOT EXISTS photometric_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT, zone_id TEXT NOT NULL,
        segment_name TEXT, match_key TEXT,
        road_width REAL, spacing REAL, lighting_class TEXT,
        power_w REAL, lm_em REAL, uo REAL, ui REAL, ti REAL, sr REAL,
        model TEXT, lente TEXT, tilt REAL, phi_lm REAL,
        cumple TEXT, notes TEXT, imported_at TEXT DEFAULT (datetime('now')),
        UNIQUE(zone_id, match_key))""",
    """CREATE TABLE IF NOT EXISTS zone_trees (
        zone_id TEXT PRIMARY KEY, trees TEXT DEFAULT '[]',
        loaded_at TEXT DEFAULT (datetime('now')))""",
    """CREATE TABLE IF NOT EXISTS users (
        id TEXT PRIMARY KEY,
        username TEXT NOT NULL UNIQUE,
        email TEXT DEFAULT '',
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'user',
        active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        last_login TEXT DEFAULT NULL)""",
]

_MIGRATIONS = [
    # Add osm_relation column to existing zones tables that lack it
    "ALTER TABLE zones ADD COLUMN osm_relation INTEGER DEFAULT NULL",
    "ALTER TABLE users ADD COLUMN active INTEGER DEFAULT 1",
    "ALTER TABLE users ADD COLUMN last_login TEXT DEFAULT NULL",
    # Fase 3 — per-luminaire overrides (tilt/height/arm/distribution)
    "ALTER TABLE luminaires ADD COLUMN tilt REAL DEFAULT NULL",
    "ALTER TABLE luminaires ADD COLUMN height_m REAL DEFAULT NULL",
    "ALTER TABLE luminaires ADD COLUMN arm_len REAL DEFAULT NULL",
    "ALTER TABLE luminaires ADD COLUMN distribution TEXT DEFAULT NULL",
]

# ── Auth helpers ──────────────────────────────────────────────────────────────
def _hash_pw(pw: str) -> str:
    salt = os.urandom(32)
    key  = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt, 600_000)
    return _b64.b64encode(salt + key).decode()

def _verify_pw(pw: str, stored: str) -> bool:
    try:
        raw  = _b64.b64decode(stored.encode())
        salt, key = raw[:32], raw[32:]
        test = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt, 600_000)
        return hmac.compare_digest(key, test)
    except Exception: return False

def _jwt_make(uid: str, username: str, role: str, ttl: int = None) -> str:
    if ttl is None: ttl = _TOKEN_TTL
    h64  = _b64.urlsafe_b64encode(b'{"alg":"HS256","typ":"JWT"}').rstrip(b"=")
    now  = int(time.time())
    body = json.dumps({"sub": uid, "usr": username, "role": role,
                       "iat": now, "exp": now + ttl}).encode()
    p64  = _b64.urlsafe_b64encode(body).rstrip(b"=")
    msg  = h64 + b"." + p64
    sig  = _b64.urlsafe_b64encode(
               hmac.new(_AUTH_SECRET.encode(), msg, hashlib.sha256).digest()
           ).rstrip(b"=")
    return (msg + b"." + sig).decode()

def _jwt_verify(token: str) -> dict:
    parts = token.split(".")
    if len(parts) != 3: raise ValueError("token malformado")
    msg = (parts[0] + "." + parts[1]).encode()
    sig = _b64.urlsafe_b64encode(
              hmac.new(_AUTH_SECRET.encode(), msg, hashlib.sha256).digest()
          ).rstrip(b"=").decode()
    if not hmac.compare_digest(sig, parts[2]):
        raise ValueError("firma inválida")
    pad = (4 - len(parts[1]) % 4) % 4
    payload = json.loads(_b64.urlsafe_b64decode(parts[1] + "=" * pad))
    if payload.get("exp", 0) < time.time(): raise ValueError("token expirado")
    return payload

def init_db():
    """Create tables if they don't exist. Apply additive migrations. Non-fatal."""
    try:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        with db() as conn:
            for sql in _SCHEMA_SQL:
                conn.execute(sql)
            # Apply additive migrations — ignore errors (column already exists)
            for sql in _MIGRATIONS:
                try: conn.execute(sql)
                except Exception: pass
            conn.commit()
        try:
            if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
                print("[SALVI] ⚠  Sin usuarios — abre la app para crear el primer administrador.")
        except Exception: pass
    except Exception as e:
        print(f"[SALVI] init_db note (DB already configured): {e}")

# ── Helpers ────────────────────────────────────────────────────────────────────
def _fval(v):
    try: return float(v) if v is not None else None
    except: return None

def _sval(v):
    return str(v).strip() if v is not None else None

def _row2dict(row):
    return dict(row) if row else None

def _rows2list(rows):
    return [dict(r) for r in rows]

def _parse_json_field(v, default=None):
    """Safely parse a JSON field that may be None, a string, or already parsed."""
    if v is None: return default if default is not None else []
    if isinstance(v, str):
        try: return json.loads(v)
        except: return default if default is not None else []
    return v

# ── Column detection for inventory import ─────────────────────────────────────
_FIELD_KEYS = {
    'lat':          [' lat', 'latitud', 'latitude', 'coord_y', 'y_coord', 'coordenada y'],
    'lon':          [' lon', ' lng', 'longitud', 'longitude', 'coord_x', 'x_coord', 'coordenada x'],
    'power_w':      [' w ', 'potencia', 'power', 'watt', 'kw'],
    'height_m':     ['altura', 'height', 'alto', 'h_m', 'h (m)'],
    'brand':        ['marca', 'brand', 'fabricante'],
    'model':        ['modelo', 'model', 'tipo'],
    'lamp_type':    ['lampara', 'lamp', 'tecnologia', 'tech'],
    'point_id':     ['id punto', 'punto_id', 'point_id', 'cod', 'codigo', 'num'],
    'circuit_id':   ['cuadro', 'panel', 'circuit', 'armario'],
    'line_id':      ['linea', 'ramal', 'line', 'circuito'],
    'support_type': ['soporte', 'support', 'bac', 'columna', 'poste'],
}
_FIELD_ORDER = list(_FIELD_KEYS)

def _detect_inv_columns(headers):
    m = {}
    for i, h in enumerate(headers):
        hl = ' ' + h.lower().strip()
        for field in _FIELD_ORDER:
            if field not in m and any(k in hl for k in _FIELD_KEYS[field]):
                m[field] = i; break
    return m

# ── Geometry helpers ───────────────────────────────────────────────────────────
def latlon_to_utm(lat, lon):
    lat_r = math.radians(lat); lon_r = math.radians(lon)
    a = 6378137.0; f = 1/298.257223563; b = a*(1-f); e2 = 1-(b/a)**2
    zone = int((lon+180)/6)+1
    lon0_r = math.radians((zone-1)*6-180+3)
    N0 = 0; E0 = 500000; k0 = 0.9996
    N = a/math.sqrt(1-e2*math.sin(lat_r)**2)
    T = math.tan(lat_r)**2; C = e2/(1-e2)*math.cos(lat_r)**2
    A_ = math.cos(lat_r)*(lon_r-lon0_r)
    e4 = e2*e2; e6 = e4*e2
    M = a*((1-e2/4-3*e4/64-5*e6/256)*lat_r
           -(3*e2/8+3*e4/32+45*e6/1024)*math.sin(2*lat_r)
           +(15*e4/256+45*e6/1024)*math.sin(4*lat_r)
           -(35*e6/3072)*math.sin(6*lat_r))
    x = k0*N*(A_+(1-T+C)*A_**3/6+(5-18*T+T**2)*A_**5/120)+E0
    y = k0*(M+N*math.tan(lat_r)*(A_**2/2+(5-T+9*C+4*C**2)*A_**4/24))+N0
    return x, y

def offset_polyline(pts_utm, dist_m):
    if len(pts_utm) < 2: return pts_utm
    out = []
    for i in range(len(pts_utm)):
        segs = []
        if i > 0:
            dx = pts_utm[i][0]-pts_utm[i-1][0]; dy = pts_utm[i][1]-pts_utm[i-1][1]
            L = math.hypot(dx, dy)
            if L > 0: segs.append((dy/L, -dx/L))
        if i < len(pts_utm)-1:
            dx = pts_utm[i+1][0]-pts_utm[i][0]; dy = pts_utm[i+1][1]-pts_utm[i][1]
            L = math.hypot(dx, dy)
            if L > 0: segs.append((dy/L, -dx/L))
        if not segs: out.append(pts_utm[i]); continue
        nx = sum(s[0] for s in segs)/len(segs)
        ny = sum(s[1] for s in segs)/len(segs)
        L2 = math.hypot(nx, ny)
        if L2 > 0: nx /= L2; ny /= L2
        out.append((pts_utm[i][0]+nx*dist_m, pts_utm[i][1]+ny*dist_m))
    return out

def road_polygon(pts_utm, half_w):
    left = offset_polyline(pts_utm, half_w)
    right = offset_polyline(pts_utm, -half_w)
    return left + right[::-1] + [left[0]]

# ── Export: DXF ───────────────────────────────────────────────────────────────
_DXF_CLR = {'motorway':1,'motorway_link':1,'trunk':14,'trunk_link':14,
             'primary':30,'primary_link':30,'secondary':2,'secondary_link':2,
             'tertiary':3,'tertiary_link':3,'residential':4,'unclassified':9,
             'living_street':8,'pedestrian':6,'service':8,'tunnel':5,'trees':82}
_DXF_LW  = {'motorway':50,'trunk':40,'primary':35,'secondary':30,
             'tertiary':25,'residential':18,'tunnel':30}

def _dxf_ldef(name, color, lw=18, lt='CONTINUOUS'):
    return ["0","LAYER","2",name,"70","0","62",str(color),"6",lt,"370",str(lw)]

def _perp_off(lon1, lat1, lon2, lat2, half_m):
    mid = math.radians((lat1 + lat2) / 2)
    cos_mid = math.cos(mid) or 0.001
    dlat_m = (lat2 - lat1) * 111320
    dlon_m = (lon2 - lon1) * 111320 * cos_mid
    dist   = math.sqrt(dlat_m**2 + dlon_m**2)
    if dist < 0.01: return 0.0, 0.0
    return (-dlon_m/dist * half_m / 111320,
             dlat_m/dist * half_m / (111320 * cos_mid))

def build_dxf(zone_id, opts=None):
    if opts is None: opts = {}
    def _on(k): return opts.get(k, '1') != '0'
    inc_cl    = _on('centerlines')
    inc_wd    = _on('widths')
    inc_lb    = _on('labels')
    inc_lum   = _on('designed_lum')
    inc_inv   = _on('inventory')
    inc_bnd   = _on('boundary')
    inc_trees = _on('trees')

    with db() as conn:
        lum_rows = _rows2list(conn.execute(
            "SELECT lat,lon,road_type,watts FROM luminaires WHERE zone_id=?", (zone_id,)))
        inv_rows = _rows2list(conn.execute(
            "SELECT lat,lon FROM inventory_luminaires WHERE zone_id=?", (zone_id,))) if inc_inv else []
        osm  = conn.execute("SELECT ways FROM zone_osm_data WHERE zone_id=?", (zone_id,)).fetchone()
        zrow = conn.execute("SELECT bounds_polygon FROM zones WHERE id=?", (zone_id,)).fetchone()
        trow = conn.execute("SELECT trees FROM zone_trees WHERE zone_id=?", (zone_id,)).fetchone() if inc_trees else None
    tree_rows = _parse_json_field(trow['trees'] if trow else None, [])

    ways     = _parse_json_field(osm["ways"] if osm else None, [])
    boundary = _parse_json_field(zrow["bounds_polygon"] if zrow else None, [])
    rtypes   = sorted({w.get('type','road') for w in ways})

    # Layer table
    layers = [("0", 7, 18, "CONTINUOUS")]
    if inc_cl:
        for rt in rtypes:
            layers.append((f"STREETS_{rt.upper()}", _DXF_CLR.get(rt,7), _DXF_LW.get(rt,18), "CONTINUOUS"))
    if inc_wd:
        for rt in rtypes:
            layers.append((f"WIDTH_{rt.upper()}", _DXF_CLR.get(rt,7), 9, "DASHED"))
    if inc_lb:
        layers.append(("STREET_LABELS", 7, 13, "CONTINUOUS"))
    if inc_lum:
        for rt in sorted({r.get('road_type') or 'GEN' for r in lum_rows}):
            layers.append((f"LUM_{rt.upper()}", 50, 18, "CONTINUOUS"))
    if inc_inv and inv_rows:
        layers.append(("INVENTORY", 140, 18, "CONTINUOUS"))
    if inc_bnd and boundary:
        layers.append(("ZONE_BOUNDARY", 7, 25, "CONTINUOUS"))
    if inc_trees and tree_rows:
        layers.append(("TREES", 82, 18, "CONTINUOUS"))

    L = []
    L += ["0","SECTION","2","HEADER","0","ENDSEC"]
    # TABLES
    L += ["0","SECTION","2","TABLES"]
    L += ["0","TABLE","2","LTYPE","70","2"]
    L += ["0","LTYPE","2","CONTINUOUS","70","0","3","Solid","72","65","73","0","40","0.0"]
    L += ["0","LTYPE","2","DASHED","70","0","3","__ __","72","65","73","2","40","0.75",
          "49","0.5","74","0","49","-0.25","74","0"]
    L += ["0","ENDTAB"]
    L += ["0","TABLE","2","LAYER","70",str(len(layers))]
    for nm,clr,lw,lt in layers:
        L += _dxf_ldef(nm, clr, lw, lt)
    L += ["0","ENDTAB","0","ENDSEC"]
    # ENTITIES
    L += ["0","SECTION","2","ENTITIES"]

    if inc_cl:
        for w in ways:
            geom = w.get("geom",[])
            if len(geom)<2: continue
            rt=w.get('type','road'); lnm=f"STREETS_{rt.upper()}"
            clr=_DXF_CLR.get(rt,7); lw=_DXF_LW.get(rt,18)
            for i in range(len(geom)-1):
                p0,p1=geom[i],geom[i+1]
                L+=["0","LINE","8",lnm,"62",str(clr),"370",str(lw),
                    "10",f"{p0['lon']:.6f}","20",f"{p0['lat']:.6f}","30","0.0",
                    "11",f"{p1['lon']:.6f}","21",f"{p1['lat']:.6f}","31","0.0"]

    if inc_wd:
        for w in ways:
            geom=w.get("geom",[])
            if len(geom)<2: continue
            rt=w.get('type','road'); lnm=f"WIDTH_{rt.upper()}"
            clr=_DXF_CLR.get(rt,7); half=(w.get('estWidth') or 6.0)/2.0
            for i in range(len(geom)-1):
                p0,p1=geom[i],geom[i+1]
                dlat,dlon=_perp_off(p0['lon'],p0['lat'],p1['lon'],p1['lat'],half)
                if not dlat and not dlon: continue
                for s in (1,-1):
                    L+=["0","LINE","8",lnm,"62",str(clr),"370","9",
                        "10",f"{p0['lon']+s*dlon:.6f}","20",f"{p0['lat']+s*dlat:.6f}","30","0.0",
                        "11",f"{p1['lon']+s*dlon:.6f}","21",f"{p1['lat']+s*dlat:.6f}","31","0.0"]

    if inc_lb:
        seen=set()
        for w in ways:
            nm=w.get('name')
            if not nm or nm in seen: continue
            seen.add(nm)
            geom=w.get("geom",[])
            if not geom: continue
            mid=geom[len(geom)//2]
            L+=["0","TEXT","8","STREET_LABELS","62","7",
                "10",f"{mid['lon']:.6f}","20",f"{mid['lat']:.6f}","30","0.0",
                "40","0.000045","1",nm[:63]]

    if inc_lum:
        for r in lum_rows:
            rt=r.get('road_type') or 'GEN'
            L+=["0","POINT","8",f"LUM_{rt.upper()}","62","50",
                "10",f"{r['lon']:.6f}","20",f"{r['lat']:.6f}","30","0.0"]

    if inc_inv:
        for r in inv_rows:
            L+=["0","POINT","8","INVENTORY","62","140",
                "10",f"{r['lon']:.6f}","20",f"{r['lat']:.6f}","30","0.0"]

    if inc_bnd and boundary:
        n=len(boundary)
        for i in range(n):
            p0=boundary[i]; p1=boundary[(i+1)%n]
            lat0,lon0=(p0[0],p0[1]) if isinstance(p0,list) else (p0['lat'],p0['lon'])
            lat1,lon1=(p1[0],p1[1]) if isinstance(p1,list) else (p1['lat'],p1['lon'])
            L+=["0","LINE","8","ZONE_BOUNDARY","62","7","370","25",
                "10",f"{lon0:.6f}","20",f"{lat0:.6f}","30","0.0",
                "11",f"{lon1:.6f}","21",f"{lat1:.6f}","31","0.0"]

    if inc_trees:
        for t in tree_rows:
            L+=["0","POINT","8","TREES","62","82",
                "10",f"{t['lon']:.6f}","20",f"{t['lat']:.6f}","30","0.0"]

    L+=["0","ENDSEC","0","EOF"]
    return "\n".join(L).encode("utf-8")

# ── Export: Luminaires XLSX ───────────────────────────────────────────────────
def build_luminaires_xlsx(zone_id):
    import openpyxl
    with db() as conn:
        rows = _rows2list(conn.execute(
            "SELECT id,zone_id,road_type,lighting_class,street_name,lat,lon,watts,spacing,placed_at "
            "FROM luminaires WHERE zone_id=? ORDER BY road_type,id", (zone_id,)))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Luminarias"
    ws.append(["ID","Zona","Tipo via","Clase ilum.","Calle","Lat","Lon","W","Espaciado","Colocada"])
    for r in rows:
        ws.append([r["id"],r["zone_id"],r["road_type"],r["lighting_class"],
                   r["street_name"],r["lat"],r["lon"],r["watts"],r["spacing"],r["placed_at"]])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Export: Plantilla luminotecnica XLSX ──────────────────────────────────────
def build_plantilla_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Plantilla luminotecnica"
    HEADERS = [
        "Zona","Tipo via","Longitud (km)","N luminarias","Disposicion",
        "Clase ilum.","Ancho calzada (m)","Espaciado (m)",
        "Fabricante","Gama","Potencia (W)","Difusor","Lente","Tipo LED",
        "CCT (K)","CRI","Brazo (m)","Inclinacion","Pavimento","MF",
    ]
    gold = "C8A96E"; dark = "1C1C1A"
    hfill = PatternFill("solid", fgColor=gold)
    hfont = Font(bold=True, color=dark, size=10)
    thin  = Side(style="thin", color="2A3A4A")
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = bord
    for r in rows:
        ws.append([r.get("zone_id",""), r.get("hw",""), r.get("km",""), r.get("n_lum",""),
                   r.get("arrangement",""), r.get("lighting_class",""), r.get("road_width",""),
                   r.get("spacing",""), r.get("manufacturer","Salvi"), r.get("gama",""),
                   r.get("power",""), r.get("difusor",""), r.get("lente",""), r.get("led_type",""),
                   r.get("cct",""), r.get("cri",""), r.get("arm_length",""),
                   r.get("tilt",""), r.get("pavement",""), r.get("mf","")])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Import: Photometric results ────────────────────────────────────────────────
def _import_photometric(zone_id, raw_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws = wb.active; rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw: return 0
    headers = [str(h).lower().strip() if h else "" for h in rows_raw[0]]
    def _col(names):
        for n in names:
            if n in headers: return headers.index(n)
        return None
    ci = {
        "segment_name":  _col(["segmento","segment","name","tramo"]),
        "match_key":     _col(["match_key","key","clave"]),
        "road_width":    _col(["ancho","road_width","w_m","width"]),
        "spacing":       _col(["espaciado","spacing","sp"]),
        "lighting_class":_col(["clase","class","lc","lighting_class"]),
        "power_w":       _col(["w","potencia","power","watts"]),
        "lm_em":         _col(["lm/m2","lm_em","em","e_m"]),
        "uo":            _col(["uo","u0"]),
        "ui":            _col(["ui"]),
        "ti":            _col(["ti"]),
        "sr":            _col(["sr"]),
        "model":         _col(["modelo","model"]),
        "lente":         _col(["lente","lens"]),
        "tilt":          _col(["tilt","inclinacion"]),
        "phi_lm":        _col(["phi","flujo","phi_lm","lm"]),
        "cumple":        _col(["cumple","ok","pass","resultado"]),
        "notes":         _col(["notas","notes","observaciones"]),
    }
    imported = 0
    with db() as conn:
        for row in rows_raw[1:]:
            if not any(row): continue
            def g(k): idx = ci.get(k); return row[idx] if idx is not None and idx < len(row) else None
            road_w  = _fval(g("road_width")); spacing = _fval(g("spacing"))
            lc      = _sval(g("lighting_class"))
            mk      = _sval(g("match_key")) or (f"{road_w}|{spacing}|{lc}" if road_w and spacing and lc else None)
            if not mk: continue
            conn.execute("""
                INSERT INTO photometric_results
                    (zone_id,segment_name,match_key,road_width,spacing,lighting_class,
                     power_w,lm_em,uo,ui,ti,sr,model,lente,tilt,phi_lm,cumple,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(zone_id,match_key) DO UPDATE SET
                    road_width=excluded.road_width, spacing=excluded.spacing,
                    lighting_class=excluded.lighting_class, power_w=excluded.power_w,
                    lm_em=excluded.lm_em, uo=excluded.uo, ui=excluded.ui,
                    ti=excluded.ti, sr=excluded.sr, model=excluded.model,
                    lente=excluded.lente, tilt=excluded.tilt, phi_lm=excluded.phi_lm,
                    cumple=excluded.cumple, notes=excluded.notes,
                    imported_at=datetime('now')
            """, (zone_id, _sval(g("segment_name")), mk, road_w, spacing, lc,
                  _fval(g("power_w")), _fval(g("lm_em")), _fval(g("uo")), _fval(g("ui")),
                  _fval(g("ti")), _fval(g("sr")), _sval(g("model")), _sval(g("lente")),
                  _fval(g("tilt")), _fval(g("phi_lm")), _sval(g("cumple")), _sval(g("notes"))))
            imported += 1
        conn.commit()
    return imported

# ── AI: Build project context ──────────────────────────────────────────────────
_DB_SCHEMA_SUMMARY = """
ESQUEMA DE LA BASE DE DATOS (SQLite):
- projects(id, name, created_at)
- zones(id, name, type, color, priority, center_lat, center_lon, zoom, bbox, description,
        est_primary, est_secondary, est_tertiary, est_residential, est_unclassified,
        corridors[JSON], bounds_polygon[JSON], osm_relation, source, created_at, project_id)
- zone_config(zone_id, spacing[m], watt_hps, watt_led, efficacy, hours_night)
- zone_osm_data(zone_id, km_by_type[JSON:{primary,secondary,tertiary,residential,unclassified,...}],
                ways[JSON], source, loaded_at)
- luminaires(id, project_id, zone_id, road_type, lighting_class, street_name, lat, lon,
             watts, spacing, placed_at)  -- luminarias diseñadas/colocadas
- inventory_luminaires(id, zone_id, point_id, lat, lon, power_w, height_m, brand, model,
                       lamp_type, support_type, circuit_id, line_id, extra[JSON], way_key,
                       road_type, imported_at)  -- inventario de campo
- photometric_results(id, zone_id, segment_name, match_key, road_width, spacing,
                      lighting_class, power_w, lm_em, uo, ui, ti, sr, model, lente, tilt,
                      phi_lm, cumple, notes, imported_at)
- project_ui_config(project_id, config_key, config_value[JSON])
"""

def _build_project_context(project_id):
    with db() as conn:
        proj = _row2dict(conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone())
        if not proj:
            proj = _row2dict(conn.execute("SELECT * FROM projects LIMIT 1").fetchone())
        if not proj: return "No hay proyectos en la base de datos."
        pid = proj["id"]
        zones = _rows2list(conn.execute("SELECT * FROM zones WHERE project_id=?", (pid,)).fetchall())
        lines = [f"PROYECTO: {proj['name']} (id={pid})", f"Zonas: {len(zones)}", ""]
        total_proj_km = 0
        total_proj_lum = 0
        for z in zones:
            zid = z["id"]
            cfg = _row2dict(conn.execute("SELECT * FROM zone_config WHERE zone_id=?", (zid,)).fetchone()) or {}
            osm = _row2dict(conn.execute("SELECT km_by_type FROM zone_osm_data WHERE zone_id=?", (zid,)).fetchone())
            km_by_type = _parse_json_field(osm["km_by_type"] if osm else None, {})
            lum_count   = conn.execute("SELECT COUNT(*) FROM luminaires WHERE zone_id=?", (zid,)).fetchone()[0]
            inv_count   = conn.execute("SELECT COUNT(*) FROM inventory_luminaires WHERE zone_id=?", (zid,)).fetchone()[0]
            photo_count = conn.execute("SELECT COUNT(*) FROM photometric_results WHERE zone_id=?", (zid,)).fetchone()[0]
            total_km = sum(float(v) for v in km_by_type.values() if v)
            spacing  = cfg.get("spacing", 30) or 30
            watt_led = cfg.get("watt_led", 60) or 60
            hours    = cfg.get("hours_night", 11.5) or 11.5
            n_est    = int(total_km * 1000 / spacing) if spacing > 0 else 0
            kw_total = n_est * watt_led / 1000
            mwh_yr   = kw_total * hours * 365 / 1000
            total_proj_km  += total_km
            total_proj_lum += n_est
            lines.append(f"ZONA: {z['name']} (id={zid})")
            if km_by_type:
                for t, km in sorted(km_by_type.items(), key=lambda x: -float(x[1] or 0)):
                    if km: lines.append(f"  - {t}: {float(km):.2f} km")
                lines.append(f"  Total red viaria: {total_km:.2f} km")
            lines.append(f"  Config: espaciado={spacing}m, LED={watt_led}W, horas/noche={hours}h")
            lines.append(f"  Luminarias estimadas: {n_est} | Potencia: {kw_total:.1f} kW | Energia: {mwh_yr:.1f} MWh/año")
            if lum_count: lines.append(f"  Luminarias diseñadas en mapa: {lum_count}")
            if inv_count:
                inv_types = _rows2list(conn.execute(
                    "SELECT lamp_type, ROUND(AVG(power_w),0) avg_w, COUNT(*) n FROM inventory_luminaires WHERE zone_id=? GROUP BY lamp_type ORDER BY n DESC LIMIT 5",
                    (zid,)).fetchall())
                inv_summary = "; ".join(f"{r['lamp_type'] or 'desconocido'} {int(r['avg_w'] or 0)}W ×{r['n']}" for r in inv_types)
                lines.append(f"  Inventario real: {inv_count} puntos ({inv_summary})")
            if photo_count:
                photo_ok = conn.execute("SELECT COUNT(*) FROM photometric_results WHERE zone_id=? AND cumple='SI'", (zid,)).fetchone()[0]
                lines.append(f"  Fotometría: {photo_count} tramos ({photo_ok} cumplen normativa)")
            # Corridors
            corridors = _parse_json_field(z.get("corridors"), [])
            if corridors:
                names = [c.get("name","?") for c in corridors if c.get("name")][:6]
                lines.append(f"  Corredores ({len(corridors)}): {', '.join(names)}")
            lines.append("")
        lines.insert(2, f"Totales: {total_proj_km:.1f} km red viaria · {total_proj_lum} luminarias estimadas")
    return "\n".join(lines)


def _safe_db_query(sql, params=None):
    """Execute a safe SELECT-only query. Returns (columns, rows) or raises."""
    sql_stripped = sql.strip()
    upper = sql_stripped.upper()
    if not (upper.startswith("SELECT") or upper.startswith("WITH")):
        raise ValueError("Solo se permiten consultas SELECT o WITH")
    if sql_stripped.count(";") > 1 or (sql_stripped.endswith(";") and sql_stripped[:-1].count(";") > 0):
        raise ValueError("No se permiten múltiples sentencias SQL")
    with db() as conn:
        cur = conn.execute(sql_stripped, params or [])
        cols = [d[0] for d in cur.description] if cur.description else []
        rows = [list(r) for r in cur.fetchmany(500)]
    return cols, rows

# ── Nominatim proxy ────────────────────────────────────────────────────────────
_NOM_HEADERS = {"User-Agent": "SalviGIS/1.0 (contact@salvi.es)", "Accept-Language": "es,en"}

def _nom_ssl_ctx():
    """Return a relaxed SSL context for nominatim (handles Windows cert store issues)."""
    import ssl
    try:
        ctx = ssl.create_default_context()
        return ctx
    except Exception:
        return ssl._create_unverified_context()

def _nominatim_search(q, featuretype=None):
    """Proxy Nominatim search server-side to avoid browser CORS/rate-limit issues."""
    import urllib.request as _ur
    import urllib.parse as _up
    base = "https://nominatim.openstreetmap.org/search"
    params = f"q={_up.quote(q)}&format=json&addressdetails=1&polygon_geojson=1&limit=10"
    if featuretype:
        params += f"&featuretype={featuretype}"
    req = _ur.Request(f"{base}?{params}", headers=_NOM_HEADERS)
    with _ur.urlopen(req, timeout=12, context=_nom_ssl_ctx()) as resp:
        return json.loads(resp.read().decode("utf-8"))

def _nominatim_reverse(lat, lon, zoom=14):
    """Proxy Nominatim reverse geocoding server-side."""
    import urllib.request as _ur
    params = f"lat={lat}&lon={lon}&format=json&polygon_geojson=1&zoom={zoom}"
    req = _ur.Request(f"https://nominatim.openstreetmap.org/reverse?{params}", headers=_NOM_HEADERS)
    with _ur.urlopen(req, timeout=12, context=_nom_ssl_ctx()) as resp:
        return json.loads(resp.read().decode("utf-8"))

# ── AI: Call Anthropic API ─────────────────────────────────────────────────────
def _call_claude_api(messages, system_prompt, max_tokens=2000):
    import urllib.request as _ur, urllib.error as _ue
    body = json.dumps({"model": _AI_MODEL, "max_tokens": max_tokens,
                       "system": system_prompt, "messages": messages}).encode()
    req = _ur.Request("https://api.anthropic.com/v1/messages", data=body,
                      headers={"Content-Type": "application/json",
                               "x-api-key": _AI_KEY,
                               "anthropic-version": "2023-06-01"})
    try:
        with _ur.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())
            return data["content"][0]["text"], data.get("usage", {})
    except _ue.HTTPError as e:
        raise RuntimeError(f"Anthropic API {e.code}: {e.read().decode('utf-8','replace')}")

# ── Route table builder ────────────────────────────────────────────────────────
def _rt(pairs):
    return [(re.compile(p), meth) for p, meth in pairs]

_GET = _rt([
    (r'^/api/auth/me$',               'h_auth_me'),
    (r'^/api/users$',                 'h_users_list'),
    (r'^/api/nominatim/search$',          'h_nominatim_search'),
    (r'^/api/nominatim/reverse$',         'h_nominatim_reverse'),
    (r'^/api/projects$',                  'h_projects'),
    (r'^/api/zones$',                     'h_zones'),
    (r'^/api/zones/osm/all$',             'h_zones_osm_all'),
    (r'^/api/zones/([^/]+)/osm$',         'h_zone_osm'),
    (r'^/api/zones/([^/]+)/inventory$',   'h_zone_inventory'),
    (r'^/api/zones/([^/]+)/trees$',        'h_zone_trees_get'),
    (r'^/api/zones/([^/]+)/photometric$', 'h_zone_photometric'),
    (r'^/api/luminaires$',                'h_luminaires'),
    (r'^/api/projects/([^/]+)/ui-config$',  'h_project_ui_config_get'),
    (r'^/api/luminaires/export$',         'h_luminaires_export'),
    (r'^/api/export/dxf$',                'h_export_dxf'),
])
_POST = _rt([
    (r'^/api/auth/login$',          'h_auth_login'),
    (r'^/api/auth/setup$',          'h_auth_setup'),
    (r'^/api/auth/reset-request$',  'h_auth_reset_request'),
    (r'^/api/auth/reset-apply$',    'h_auth_reset_apply'),
    (r'^/api/users$',        'h_users_create'),
    (r'^/api/projects$',                         'h_project_create'),
    (r'^/api/zones$',                            'h_zone_create'),
    (r'^/api/luminaires/bulk$',                  'h_luminaires_bulk'),
    (r'^/api/export/plantilla_luminotecnica$',   'h_export_plantilla'),
    (r'^/api/parse/inventory_excel$',            'h_parse_inventory'),
    (r'^/api/import/inventory$',                 'h_import_inventory'),
    (r'^/api/import/photometric$',               'h_import_photometric'),
    (r'^/api/ai/ask$',                           'h_ai_ask'),
    (r'^/api/db/query$',                         'h_db_query'),
])
_PUT = _rt([
    (r'^/api/users/([^/]+)$',         'h_users_update'),
    (r'^/api/projects/([^/]+)/ui-config$',  'h_project_ui_config_put'),
    (r'^/api/zones/([^/]+)/trees$',   'h_zone_trees_put'),
    (r'^/api/zones/([^/]+)/osm$',    'h_zone_osm_save'),
    (r'^/api/zones/([^/]+)/config$', 'h_zone_config'),
    (r'^/api/zones/([^/]+)$',        'h_zone_update'),
])
_DELETE = _rt([
    (r'^/api/users/([^/]+)$',         'h_users_delete'),
    (r'^/api/projects/([^/]+)$',              'h_project_delete'),
    (r'^/api/zones/([^/]+)$',                 'h_zone_delete'),
    (r'^/api/luminaires/([^/]+)/([^/]+)$',    'h_luminaires_delete'),
])

# ── Request handler ────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def _send(self, code, obj, content_type="application/json"):
        body = json.dumps(obj).encode() if content_type == "application/json" else obj
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", len(body))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _send_binary(self, code, body, content_type, filename):
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", len(body))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _body(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b""
        try: return json.loads(raw)
        except: return {}

    def _raw_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length) if length else b""

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type,Content-Length,Authorization")
        self.end_headers()

    def _dispatch(self, table, qs):
        _PUBLIC = {"/api/auth/login", "/api/auth/setup"}
        p = urlparse(self.path).path.rstrip("/")
        self._current_user = None
        for pat, method in table:
            mo = pat.match(p)
            if mo:
                if _AUTH_DISABLED:
                    # Preserve admin-only features for this single-user local mode.
                    self._current_user = {"sub": "local", "usr": "local", "role": "admin"}
                elif p not in _PUBLIC:
                    auth = self.headers.get("Authorization", "")
                    if not auth.startswith("Bearer "):
                        with db() as _c:
                            needs = _c.execute(
                                "SELECT COUNT(*) FROM users WHERE active=1"
                            ).fetchone()[0] == 0
                        self._send(401, {"error": "unauthenticated", "needs_setup": bool(needs)}); return
                    try:
                        self._current_user = _jwt_verify(auth[7:])
                    except ValueError as e:
                        self._send(401, {"error": str(e)}); return
                getattr(self, method)(qs, mo); return
        self._send(404, {"error": "not found"})

    def do_GET(self):    self._dispatch(_GET,    parse_qs(urlparse(self.path).query))
    def do_POST(self):   self._dispatch(_POST,   parse_qs(urlparse(self.path).query))
    def do_PUT(self):    self._dispatch(_PUT,    parse_qs(urlparse(self.path).query))
    def do_DELETE(self): self._dispatch(_DELETE, parse_qs(urlparse(self.path).query))

    # ── Projects ────────────────────────────────────────────────────────────────
    def h_nominatim_search(self, qs, m):
        import urllib.error as _ue
        q = (qs.get('q') or [''])[0].strip()
        if not q:
            self._send(400, {"error": "Missing q parameter"}); return
        try:
            # Try with featuretype=settlement; if that call itself fails, skip gracefully
            data = []
            try:
                data = _nominatim_search(q, featuretype='settlement')
            except Exception:
                pass
            if not data:
                data = _nominatim_search(q)
            self._send(200, data)
        except _ue.URLError as e:
            self._send(502, {"error": f"Nominatim unreachable: {e.reason}"})
        except Exception as e:
            self._send(502, {"error": str(e)})

    def h_nominatim_reverse(self, qs, m):
        import urllib.error as _ue
        lat = (qs.get('lat') or [''])[0].strip()
        lon = (qs.get('lon') or [''])[0].strip()
        zoom = (qs.get('zoom') or ['14'])[0].strip()
        if not lat or not lon:
            self._send(400, {"error": "Missing lat/lon"}); return
        try:
            data = _nominatim_reverse(lat, lon, zoom)
            self._send(200, data)
        except _ue.URLError as e:
            self._send(502, {"error": f"Nominatim unreachable: {e.reason}"})
        except Exception as e:
            self._send(502, {"error": str(e)})

    def h_projects(self, qs, m):
        with db() as conn:
            rows = _rows2list(conn.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall())
        self._send(200, rows)

    def h_project_create(self, qs, m):
        b    = self._body()
        name = _sval(b.get("name")) or "Nuevo proyecto"
        pid  = str(uuid.uuid4())[:8]
        with db() as conn:
            conn.execute("INSERT INTO projects (id,name) VALUES (?,?)", (pid, name))
            conn.commit()
            row = _row2dict(conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone())
        self._send(201, row)

    def h_project_delete(self, qs, m):
        pid = m.group(1)
        with db() as conn:
            conn.execute("UPDATE zones SET project_id=NULL WHERE project_id=?", (pid,))
            conn.execute("DELETE FROM projects WHERE id=?", (pid,))
            conn.commit()
        self._send(200, {"ok": True})

    # ── Zones ───────────────────────────────────────────────────────────────────
    def _format_zones(self, rows):
        """Shared formatting for zone rows: parse JSON fields, build est dict."""
        for r in rows:
            r["corridors"]      = _parse_json_field(r.get("corridors"), [])
            r["bounds_polygon"] = _parse_json_field(r.get("bounds_polygon"), [])
            r["est"] = {
                "primary":      _fval(r.pop("est_primary",    0)) or 0,
                "secondary":    _fval(r.pop("est_secondary",  0)) or 0,
                "tertiary":     _fval(r.pop("est_tertiary",   0)) or 0,
                "residential":  _fval(r.pop("est_residential",0)) or 0,
                "unclassified": _fval(r.pop("est_unclassified",0)) or 0,
            }
            # Derive center_lat/center_lon from bounds_polygon centroid when null
            if r.get("center_lat") is None and r["bounds_polygon"]:
                pts = r["bounds_polygon"]
                if pts:
                    r["center_lat"] = sum(p[0] for p in pts) / len(pts)
                    r["center_lon"] = sum(p[1] for p in pts) / len(pts)
        return rows

    def h_zones(self, qs, m):
        project_id = (qs.get("project_id") or [None])[0]
        with db() as conn:
            if project_id:
                rows = _rows2list(conn.execute(
                    "SELECT z.*, COALESCE(c.spacing,30) AS spacing "
                    "FROM zones z LEFT JOIN zone_config c ON z.id=c.zone_id "
                    "WHERE z.project_id=? ORDER BY z.created_at DESC", (project_id,)).fetchall())
            else:
                rows = _rows2list(conn.execute(
                    "SELECT z.*, COALESCE(c.spacing,30) AS spacing "
                    "FROM zones z LEFT JOIN zone_config c ON z.id=c.zone_id "
                    "ORDER BY z.created_at DESC").fetchall())
        self._send(200, self._format_zones(rows))

    def h_zone_create(self, qs, m):
        b   = self._body()
        zid = b.get("id") or str(uuid.uuid4())[:12]
        est = b.get("est") or {}
        # Accept both camelCase (osmRelation) and snake_case (osm_relation)
        osm_rel_raw = b.get("osm_relation") or b.get("osmRelation")
        osm_rel = int(osm_rel_raw) if osm_rel_raw is not None else None
        # Accept center_lat/center_lon flat fields OR center:[lat,lon] array
        _center = b.get("center") or []
        _clat = _fval(b.get("center_lat")) or (_fval(_center[0]) if _center else None)
        _clon = _fval(b.get("center_lon")) or (_fval(_center[1]) if _center else None)
        with db() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO zones "
                "(id,name,type,color,priority,center_lat,center_lon,zoom,"
                "bbox,description,est_primary,est_secondary,est_tertiary,"
                "est_residential,est_unclassified,corridors,bounds_polygon,"
                "osm_relation,source,project_id) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (zid, _sval(b.get("name")) or "Zona",
                 _sval(b.get("type")) or "",
                 _sval(b.get("color")) or DEFAULT_COLORS[0],
                 int(b.get("priority") or 2),
                 _clat, _clon,
                 int(b.get("zoom") or 12),
                 _sval(b.get("bbox")) or "",
                 _sval(b.get("description")) or "",
                 _fval(est.get("primary")),   _fval(est.get("secondary")),
                 _fval(est.get("tertiary")),  _fval(est.get("residential")),
                 _fval(est.get("unclassified")),
                 json.dumps(b.get("corridors") or []),
                 json.dumps(b.get("bounds_polygon") or []),
                 osm_rel,
                 _sval(b.get("source")) or "manual",
                 _sval(b.get("project_id"))))
            conn.execute("INSERT OR IGNORE INTO zone_config (zone_id) VALUES (?)", (zid,))
            conn.commit()
            row = _row2dict(conn.execute(
                "SELECT z.*, COALESCE(c.spacing,30) AS spacing "
                "FROM zones z LEFT JOIN zone_config c ON z.id=c.zone_id WHERE z.id=?", (zid,)).fetchone())
        self._send(201, self._format_zones([row])[0])

    def h_zone_update(self, qs, m):
        zid = m.group(1); b = self._body()
        allowed = {"name","type","color","priority","center_lat","center_lon","zoom",
                   "bbox","description","source","project_id","corridors","bounds_polygon",
                   "est_primary","est_secondary","est_tertiary","est_residential","est_unclassified"}
        sets = []; vals = []
        for k, v in b.items():
            if k not in allowed: continue
            if k in ("corridors","bounds_polygon") and isinstance(v, (list, dict)):
                v = json.dumps(v)
            sets.append(f"{k}=?"); vals.append(v)
        if sets:
            vals.append(zid)
            with db() as conn:
                conn.execute(f"UPDATE zones SET {','.join(sets)} WHERE id=?", vals)
                conn.commit()
        self._send(200, {"ok": True})

    def h_zone_delete(self, qs, m):
        zid = m.group(1)
        with db() as conn:
            for tbl in ("luminaires","inventory_luminaires","photometric_results",
                        "zone_osm_data","zone_config","zones"):
                col = "id" if tbl == "zones" else "zone_id"
                conn.execute(f"DELETE FROM {tbl} WHERE {col}=?", (zid,))
            conn.commit()
        self._send(200, {"ok": True})

    # ── Zone OSM ────────────────────────────────────────────────────────────────
    def h_zone_osm(self, qs, m):
        zid = m.group(1)
        with db() as conn:
            row = _row2dict(conn.execute("SELECT * FROM zone_osm_data WHERE zone_id=?", (zid,)).fetchone())
        if not row: self._send(200, {}); return
        row["km_by_type"] = _parse_json_field(row["km_by_type"], {})
        row["ways"]       = _parse_json_field(row["ways"], [])
        self._send(200, row)

    def h_zone_osm_save(self, qs, m):
        zid = m.group(1); b = self._body()
        with db() as conn:
            conn.execute(
                "INSERT INTO zone_osm_data (zone_id,km_by_type,ways,source) VALUES (?,?,?,?) "
                "ON CONFLICT(zone_id) DO UPDATE SET km_by_type=excluded.km_by_type, "
                "ways=excluded.ways, source=excluded.source, loaded_at=datetime('now')",
                (zid, json.dumps(b.get("kmByType") or {}),
                 json.dumps(b.get("ways") or []),
                 _sval(b.get("source")) or "osm"))
            conn.commit()
        self._send(200, {"ok": True})

    def h_zones_osm_all(self, qs, m):
        project_id = (qs.get("project_id") or [None])[0]
        with db() as conn:
            if project_id:
                rows = _rows2list(conn.execute(
                    "SELECT o.* FROM zone_osm_data o "
                    "JOIN zones z ON z.id=o.zone_id WHERE z.project_id=?", (project_id,)).fetchall())
            else:
                rows = _rows2list(conn.execute("SELECT * FROM zone_osm_data").fetchall())
        result = {}
        for r in rows:
            r["km_by_type"] = _parse_json_field(r["km_by_type"], {})
            r["ways"]       = _parse_json_field(r["ways"], [])
            result[r["zone_id"]] = r
        self._send(200, result)

    def h_zone_config(self, qs, m):
        zid = m.group(1); b = self._body()
        allowed = {"spacing","watt_hps","watt_led","efficacy","hours_night"}
        sets = []; vals = []
        for k, v in b.items():
            if k in allowed: sets.append(f"{k}=?"); vals.append(v)
        sets.append("updated_at=datetime('now')"); vals.append(zid)
        with db() as conn:
            conn.execute("INSERT OR IGNORE INTO zone_config (zone_id) VALUES (?)", (zid,))
            conn.execute(f"UPDATE zone_config SET {','.join(sets)} WHERE zone_id=?", vals)
            conn.commit()
        self._send(200, {"ok": True})

    # ── Luminaires ──────────────────────────────────────────────────────────────
    def h_luminaires(self, qs, m):
        zid = (qs.get("zone_id") or [None])[0]
        with db() as conn:
            if zid:
                rows = _rows2list(conn.execute(
                    "SELECT * FROM luminaires WHERE zone_id=? ORDER BY road_type,id", (zid,)).fetchall())
            else:
                rows = _rows2list(conn.execute(
                    "SELECT * FROM luminaires ORDER BY zone_id,road_type,id").fetchall())
        self._send(200, rows)

    def h_luminaires_bulk(self, qs, m):
        b = self._body()
        zid = b.get("zone_id"); road_type = b.get("road_type"); items = b.get("items") or []
        if not zid or not items: self._send(400, {"error": "zone_id and items required"}); return
        with db() as conn:
            if road_type:
                conn.execute("DELETE FROM luminaires WHERE zone_id=? AND road_type=?", (zid, road_type))
            else:
                conn.execute("DELETE FROM luminaires WHERE zone_id=?", (zid,))
            conn.executemany(
                "INSERT INTO luminaires (project_id,zone_id,road_type,lighting_class,street_name,lat,lon,"
                "watts,spacing,tilt,height_m,arm_len,distribution) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [(it.get("project_id"), zid, road_type or it.get("road_type"), it.get("lighting_class"),
                  it.get("street_name"), _fval(it.get("lat")), _fval(it.get("lon")),
                  _fval(it.get("watts")), _fval(it.get("spacing")), _fval(it.get("tilt")),
                  _fval(it.get("height_m")), _fval(it.get("arm_len")), it.get("distribution"))
                 for it in items])
            conn.commit()
        self._send(200, {"inserted": len(items)})

    def h_luminaires_export(self, qs, m):
        zid = (qs.get("zone_id") or [None])[0]
        if not zid: self._send(400, {"error": "zone_id required"}); return
        try:
            self._send_binary(200, build_luminaires_xlsx(zid),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"luminarias_{zid}.xlsx")
        except Exception as e:
            self._send(500, {"error": str(e)})

    def h_luminaires_delete(self, qs, m):
        with db() as conn:
            conn.execute("DELETE FROM luminaires WHERE zone_id=? AND road_type=?",
                         (m.group(1), m.group(2)))
            conn.commit()
        self._send(200, {"ok": True})

    # ── Exports ─────────────────────────────────────────────────────────────────
    def h_export_plantilla(self, qs, m):
        b = self._body(); rows = b.get("rows") or []
        if not rows: self._send(400, {"error": "rows required"}); return
        try:
            self._send_binary(200, build_plantilla_xlsx(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"plantilla_luminotecnica_{b.get('zone_id','zona')}.xlsx")
        except Exception as e:
            self._send(500, {"error": str(e)})

    def h_export_dxf(self, qs, m):
        zid = (qs.get("zone_id") or [None])[0]
        if not zid: self._send(400, {"error": "zone_id required"}); return
        try:
            opts = {k: v[0] for k, v in qs.items() if k != "zone_id"}
            self._send_binary(200, build_dxf(zid, opts), "application/dxf", f"salvi_{zid}.dxf")
        except Exception as e:
            self._send(500, {"error": str(e)})

    # ── Inventory ───────────────────────────────────────────────────────────────
    def h_parse_inventory(self, qs, m):
        raw = self._raw_body()
        if not raw: self._send(400, {"error": "empty body"}); return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active; all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows: self._send(400, {"error": "empty sheet"}); return
            headers   = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(all_rows[0])]
            data_rows = all_rows[1:]
            row_count = sum(1 for r in data_rows if any(c is not None for c in r))
            sample    = [{headers[i]: (str(r[i]) if r[i] is not None else "")
                         for i in range(len(headers))}
                        for r in data_rows[:5] if any(c is not None for c in r)]
            detected  = _detect_inv_columns(headers)
            temp_id   = str(uuid.uuid4())
            _pending_imports[temp_id] = {
                "rows": [[str(c) if c is not None else "" for c in r] for r in data_rows],
                "headers": headers, "ts": time.time()
            }
            cutoff = time.time() - 1800
            for k in [k for k, v in list(_pending_imports.items()) if v["ts"] < cutoff]:
                del _pending_imports[k]
            self._send(200, {"temp_id": temp_id, "headers": headers,
                             "sample": sample, "row_count": row_count, "detected": detected})
        except Exception as e:
            self._send(500, {"error": str(e)})

    def h_import_inventory(self, qs, m):
        b = self._body()
        temp_id    = b.get("temp_id"); mapping = b.get("mapping") or {}
        zone_name  = _sval(b.get("zone_name")) or "Inventario"
        project_id = _sval(b.get("project_id"))
        color      = _sval(b.get("color")) or DEFAULT_COLORS[0]
        if not temp_id or temp_id not in _pending_imports:
            self._send(400, {"error": "temp_id not found or expired"}); return
        cache = _pending_imports.pop(temp_id); rows = cache["rows"]
        lats, lons, lum_rows = [], [], []
        for row in rows:
            def gm(f):
                idx = mapping.get(f)
                if idx is None: return None
                idx = int(idx); return row[idx] if idx < len(row) else None
            lat = _fval(gm("lat")); lon = _fval(gm("lon"))
            if lat is None or lon is None: continue
            lats.append(lat); lons.append(lon)
            lum_rows.append({
                "lat": lat, "lon": lon, "power_w": _fval(gm("power_w")),
                "height_m": _fval(gm("height_m")), "brand": _sval(gm("brand")),
                "model": _sval(gm("model")), "lamp_type": _sval(gm("lamp_type")),
                "point_id": _sval(gm("point_id")), "circuit_id": _sval(gm("circuit_id")),
                "line_id": _sval(gm("line_id")), "support_type": _sval(gm("support_type")),
            })
        if not lum_rows: self._send(400, {"error": "No valid lat/lon rows"}); return
        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)
        center_lat = (min_lat+max_lat)/2; center_lon = (min_lon+max_lon)/2
        bounds_poly = [
            {"lat":min_lat,"lon":min_lon},{"lat":max_lat,"lon":min_lon},
            {"lat":max_lat,"lon":max_lon},{"lat":min_lat,"lon":max_lon},
            {"lat":min_lat,"lon":min_lon},
        ]
        zid = str(uuid.uuid4())[:12]
        with db() as conn:
            conn.execute(
                "INSERT INTO zones (id,name,color,center_lat,center_lon,zoom,"
                "bounds_polygon,source,project_id) VALUES (?,?,?,?,?,?,?,?,?)",
                (zid, zone_name, color, center_lat, center_lon, 14,
                 json.dumps(bounds_poly), "inventory", project_id))
            conn.execute("INSERT OR IGNORE INTO zone_config (zone_id) VALUES (?)", (zid,))
            conn.executemany(
                "INSERT INTO inventory_luminaires "
                "(zone_id,lat,lon,power_w,height_m,brand,model,lamp_type,"
                "point_id,circuit_id,line_id,support_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                [(zid, r["lat"], r["lon"], r["power_w"], r["height_m"], r["brand"],
                  r["model"], r["lamp_type"], r["point_id"], r["circuit_id"],
                  r["line_id"], r["support_type"]) for r in lum_rows])
            conn.commit()
            zone_row = _row2dict(conn.execute(
                "SELECT z.*, COALESCE(c.spacing,30) AS spacing "
                "FROM zones z LEFT JOIN zone_config c ON z.id=c.zone_id WHERE z.id=?", (zid,)).fetchone())
        self._send(201, {"zone": self._format_zones([zone_row])[0], "count": len(lum_rows)})

    def h_zone_inventory(self, qs, m):
        zid = m.group(1)
        with db() as conn:
            rows = _rows2list(conn.execute(
                "SELECT * FROM inventory_luminaires WHERE zone_id=? ORDER BY id", (zid,)).fetchall())
        self._send(200, rows)

    # ── Photometric ─────────────────────────────────────────────────────────────
    def h_import_photometric(self, qs, m):
        zid = (qs.get("zone_id") or [None])[0]
        if not zid: self._send(400, {"error": "zone_id required"}); return
        raw = self._raw_body()
        if not raw: self._send(400, {"error": "empty body"}); return
        try:
            n = _import_photometric(zid, raw)
            with db() as conn:
                rows = _rows2list(conn.execute(
                    "SELECT * FROM photometric_results WHERE zone_id=? ORDER BY id", (zid,)).fetchall())
            self._send(200, {"imported": n, "rows": rows})
        except Exception as e:
            self._send(500, {"error": str(e)})

    def h_zone_photometric(self, qs, m):
        zid = m.group(1)
        with db() as conn:
            rows = _rows2list(conn.execute(
                "SELECT * FROM photometric_results WHERE zone_id=? ORDER BY id", (zid,)).fetchall())
        self._send(200, rows)

    # ── Project UI config ────────────────────────────────────────────────────────
    def h_project_ui_config_get(self, qs, m):
        pid = m.group(1)
        with db() as conn:
            rows = conn.execute(
                "SELECT config_key, config_value FROM project_ui_config WHERE project_id=?", (pid,)
            ).fetchall()
        result = {r[0]: json.loads(r[1]) if r[1] and r[1].startswith('[') or r[1] and r[1].startswith('{') else r[1]
                  for r in rows}
        self._send(200, result)

    def h_project_ui_config_put(self, qs, m):
        pid = m.group(1)
        b = self._body()
        key = _sval(b.get("key"))
        val = b.get("value")
        if not key: self._send(400, {"error": "key required"}); return
        val_str = json.dumps(val) if not isinstance(val, str) else val
        with db() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO project_ui_config (project_id, config_key, config_value, updated_at) "
                "VALUES (?, ?, ?, datetime('now'))", (pid, key, val_str)
            )
            conn.commit()
        self._send(200, {"ok": True})

        # ── AI ──────────────────────────────────────────────────────────────────────
    def h_ai_ask(self, qs, m):
        if not _AI_KEY:
            self._send(503, {"error": "ANTHROPIC_API_KEY no configurada."}); return
        b = self._body()
        messages   = b.get("messages") or []
        project_id = b.get("project_id") or "angola"
        if not messages: self._send(400, {"error": "messages required"}); return
        try:
            context = _build_project_context(project_id)
            system  = (
                "Eres SALVI AI, el asistente experto del software SALVI GIS para analisis de iluminacion publica en Angola.\n"
                "Respondes siempre en espanol. Eres tecnico, preciso y analitico.\n"
                "Usa los datos exactos del proyecto. No inventes cifras.\n\n"
                "CAPACIDADES:\n"
                "- Tienes acceso completo a los datos del proyecto (zonas, luminarias, inventario, fotometria).\n"
                "- Puedes ejecutar consultas SQL sobre la base de datos usando el bloque especial:\n"
                "  <sql>SELECT ...</sql>\n"
                "  El sistema ejecutara la consulta y te devolvera los resultados para que puedas responder con datos precisos.\n"
                "- Puedes generar scripts SQL, Python o JavaScript para el usuario.\n"
                "- Cuando el usuario pida algo que requiera datos especificos, genera la consulta SQL adecuada.\n\n"
                + _DB_SCHEMA_SUMMARY + "\n"
                "DATOS DEL PROYECTO ACTIVO:\n" + context
            )
            # First pass
            answer, usage = _call_claude_api(messages, system, max_tokens=2000)

            # Two-pass: detect <sql>...</sql> blocks, execute them, feed results back
            sql_queries = re.findall(r'<sql>(.*?)</sql>', answer, re.DOTALL | re.IGNORECASE)
            if sql_queries:
                results_text = []
                for sql in sql_queries:
                    sql = sql.strip()
                    try:
                        cols, rows = _safe_db_query(sql)
                        if rows:
                            header = " | ".join(cols)
                            sep    = "-" * len(header)
                            row_lines = [" | ".join(str(v) for v in r) for r in rows[:50]]
                            results_text.append(f"Resultado de: {sql[:80]}...\n{header}\n{sep}\n" + "\n".join(row_lines))
                        else:
                            results_text.append(f"Resultado de: {sql[:80]}...\n(Sin resultados)")
                    except Exception as qe:
                        results_text.append(f"Error en consulta: {qe}")
                # Second pass with results
                followup_messages = list(messages) + [
                    {"role": "assistant", "content": answer},
                    {"role": "user",      "content": "Resultados de las consultas SQL:\n\n" + "\n\n".join(results_text) + "\n\nAhora responde la pregunta original con estos datos."}
                ]
                answer, usage2 = _call_claude_api(followup_messages, system, max_tokens=2000)
                usage["input_tokens"]  = usage.get("input_tokens",0)  + usage2.get("input_tokens",0)
                usage["output_tokens"] = usage.get("output_tokens",0) + usage2.get("output_tokens",0)

            self._send(200, {"answer": answer, "usage": usage})
        except RuntimeError as ex:
            self._send(502, {"error": str(ex)})
        except Exception as ex:
            import traceback
            self._send(500, {"error": str(ex), "detail": traceback.format_exc()})

    # ── Auth ──────────────────────────────────────────────────────────────────────
    def h_auth_login(self, qs, m):
        b = self._body()
        username    = (b.get("username") or "").strip()
        password    = b.get("password") or ""
        remember_me = bool(b.get("remember_me"))          # 30-day token when true
        if not username or not password:
            self._send(400, {"error": "usuario y contraseña requeridos"}); return
        with db() as conn:
            # Case-insensitive username lookup
            row = conn.execute(
                "SELECT * FROM users WHERE lower(username)=lower(?) AND active=1", (username,)
            ).fetchone()
        if not row or not _verify_pw(password, row["password_hash"]):
            self._send(401, {"error": "Credenciales incorrectas"}); return
        with db() as conn:
            conn.execute("UPDATE users SET last_login=datetime('now') WHERE id=?", (row["id"],))
            conn.commit()
        ttl   = 86400 * 30 if remember_me else _TOKEN_TTL   # 30 days or 7 days
        token = _jwt_make(row["id"], row["username"], row["role"], ttl=ttl)
        self._send(200, {"token": token, "user": {
            "id": row["id"], "username": row["username"],
            "role": row["role"], "email": row["email"] or ""}})

    def h_auth_setup(self, qs, m):
        with db() as conn:
            if conn.execute("SELECT COUNT(*) FROM users WHERE active=1").fetchone()[0] > 0:
                self._send(403, {"error": "Setup ya completado"}); return
            b = self._body()
            username = (b.get("username") or "").strip()
            password = b.get("password") or ""
            email    = (b.get("email") or "").strip()
            if not username or not password:
                self._send(400, {"error": "usuario y contraseña requeridos"}); return
            if len(password) < 8:
                self._send(400, {"error": "La contraseña debe tener al menos 8 caracteres"}); return
            uid = str(uuid.uuid4())[:12]
            try:
                conn.execute(
                    "INSERT INTO users (id,username,email,password_hash,role) VALUES (?,?,?,?,?)",
                    (uid, username, email, _hash_pw(password), "admin"))
                conn.commit()
            except Exception as e:
                self._send(400, {"error": str(e)}); return
        token = _jwt_make(uid, username, "admin")
        self._send(201, {"token": token, "user": {
            "id": uid, "username": username, "role": "admin", "email": email}})

    def h_auth_me(self, qs, m):
        u = self._current_user
        if not u: self._send(401, {"error": "unauthenticated"}); return
        if _AUTH_DISABLED:
            self._send(200, {"id": "local", "username": "local", "email": "", "role": "admin", "local_access": True})
            return
        with db() as conn:
            row = conn.execute(
                "SELECT id,username,email,role,last_login FROM users WHERE id=? AND active=1", (u["sub"],)
            ).fetchone()
        if not row: self._send(401, {"error": "usuario no encontrado"}); return
        self._send(200, dict(row))

    def h_auth_reset_request(self, qs, m):
        """Generate a 1-hour password-reset token for a given username/email.
        Admin-only OR unauthenticated (for forgotten password flow).
        Returns the token in the response so an admin can copy+share it, or
        sends email if SMTP is configured in .env."""
        b = self._body()
        identifier = (b.get("username") or b.get("email") or "").strip()
        if not identifier:
            self._send(400, {"error": "Proporciona usuario o email"}); return
        with db() as conn:
            row = conn.execute(
                "SELECT id,username,email FROM users "
                "WHERE (lower(username)=lower(?) OR lower(email)=lower(?)) AND active=1",
                (identifier, identifier)
            ).fetchone()
        if not row:
            # Don't reveal whether user exists — but still return 200
            self._send(200, {"ok": True, "info": "Si el usuario existe recibirá instrucciones."}); return
        tok = _sec.token_urlsafe(32)
        _RESET_TOKENS[tok] = {
            "uid": row["id"], "username": row["username"],
            "exp": int(time.time()) + 3600
        }
        # Try SMTP if configured
        smtp_sent = False
        smtp_host = os.environ.get("SMTP_HOST", "")
        smtp_user = os.environ.get("SMTP_USER", "")
        smtp_pass = os.environ.get("SMTP_PASS", "")
        smtp_port = int(os.environ.get("SMTP_PORT", "587"))
        if smtp_host and smtp_user and smtp_pass and row["email"]:
            try:
                import smtplib, email.mime.text as _mt
                reset_url = f"http://localhost:8733/reset?token={tok}"
                body = (f"Hola {row['username']},\n\n"
                        f"Para restablecer tu contraseña en SALVI GIS, copia este enlace:\n\n"
                        f"  {reset_url}\n\n"
                        f"El enlace expira en 1 hora.\n\n"
                        f"Si no solicitaste este cambio, ignora este mensaje.")
                msg = _mt.MIMEText(body, "plain", "utf-8")
                msg["Subject"] = "SALVI GIS — Restablecer contraseña"
                msg["From"]    = smtp_user
                msg["To"]      = row["email"]
                with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as s:
                    s.starttls()
                    s.login(smtp_user, smtp_pass)
                    s.sendmail(smtp_user, [row["email"]], msg.as_string())
                smtp_sent = True
            except Exception as e:
                print(f"[SALVI] SMTP error: {e}")
        self._send(200, {
            "ok": True,
            "token": tok,              # always return token so admin can copy if no SMTP
            "smtp_sent": smtp_sent,
            "expires_in": 3600,
            "username": row["username"],
        })

    def h_auth_reset_apply(self, qs, m):
        """Apply a reset: given token + new password, update the user's password."""
        b = self._body()
        tok      = (b.get("token") or "").strip()
        password = b.get("password") or ""
        if not tok or not password:
            self._send(400, {"error": "token y contraseña requeridos"}); return
        entry = _RESET_TOKENS.get(tok)
        if not entry or entry["exp"] < int(time.time()):
            _RESET_TOKENS.pop(tok, None)
            self._send(401, {"error": "Token inválido o expirado"}); return
        if len(password) < 8:
            self._send(400, {"error": "La contraseña debe tener al menos 8 caracteres"}); return
        uid = entry["uid"]; username = entry["username"]
        with db() as conn:
            conn.execute("UPDATE users SET password_hash=? WHERE id=?", (_hash_pw(password), uid))
            conn.commit()
        _RESET_TOKENS.pop(tok, None)
        token = _jwt_make(uid, username, "user")
        # Re-fetch role
        with db() as conn:
            row = conn.execute("SELECT role FROM users WHERE id=?", (uid,)).fetchone()
        role = row["role"] if row else "user"
        token = _jwt_make(uid, username, role)
        self._send(200, {"ok": True, "token": token,
                         "user": {"id": uid, "username": username, "role": role}})

    def h_users_list(self, qs, m):
        u = self._current_user
        if not u or u.get("role") != "admin":
            self._send(403, {"error": "Solo administradores"}); return
        with db() as conn:
            rows = _rows2list(conn.execute(
                "SELECT id,username,email,role,active,created_at,last_login FROM users ORDER BY created_at"
            ).fetchall())
        self._send(200, rows)

    def h_users_create(self, qs, m):
        u = self._current_user
        if not u or u.get("role") != "admin":
            self._send(403, {"error": "Solo administradores"}); return
        b = self._body()
        username = (b.get("username") or "").strip()
        password = b.get("password") or ""
        email    = (b.get("email") or "").strip()
        role     = b.get("role", "user")
        if not username or not password:
            self._send(400, {"error": "usuario y contraseña requeridos"}); return
        if len(password) < 8:
            self._send(400, {"error": "La contraseña debe tener al menos 8 caracteres"}); return
        if role not in ("admin", "user"): role = "user"
        uid = str(uuid.uuid4())[:12]
        try:
            with db() as conn:
                conn.execute(
                    "INSERT INTO users (id,username,email,password_hash,role) VALUES (?,?,?,?,?)",
                    (uid, username, email, _hash_pw(password), role))
                conn.commit()
                row = _row2dict(conn.execute(
                    "SELECT id,username,email,role,active,created_at FROM users WHERE id=?", (uid,)
                ).fetchone())
            self._send(201, row)
        except sqlite3.IntegrityError:
            self._send(409, {"error": f'El usuario "{username}" ya existe'})

    def h_users_update(self, qs, m):
        u = self._current_user
        if not u or u.get("role") != "admin":
            self._send(403, {"error": "Solo administradores"}); return
        uid = m.group(1); b = self._body()
        updates = []; params = []
        if "username" in b and str(b["username"]).strip():
            updates.append("username=?"); params.append(str(b["username"]).strip())
        if "email" in b:
            updates.append("email=?"); params.append(str(b["email"]).strip())
        if "role" in b and b["role"] in ("admin", "user"):
            updates.append("role=?"); params.append(b["role"])
        if "active" in b:
            updates.append("active=?"); params.append(1 if b["active"] else 0)
        if "password" in b and b["password"]:
            if len(b["password"]) < 8:
                self._send(400, {"error": "La contraseña debe tener al menos 8 caracteres"}); return
            updates.append("password_hash=?"); params.append(_hash_pw(b["password"]))
        if not updates:
            self._send(400, {"error": "Nada que actualizar"}); return
        params.append(uid)
        with db() as conn:
            conn.execute(f"UPDATE users SET {','.join(updates)} WHERE id=?", params)
            conn.commit()
            row = _row2dict(conn.execute(
                "SELECT id,username,email,role,active,created_at,last_login FROM users WHERE id=?", (uid,)
            ).fetchone())
        if not row: self._send(404, {"error": "Usuario no encontrado"}); return
        self._send(200, row)

    def h_users_delete(self, qs, m):
        u = self._current_user
        if not u or u.get("role") != "admin":
            self._send(403, {"error": "Solo administradores"}); return
        uid = m.group(1)
        if uid == u.get("sub"):
            self._send(400, {"error": "No puedes eliminar tu propia cuenta"}); return
        with db() as conn:
            row = conn.execute("SELECT role FROM users WHERE id=?", (uid,)).fetchone()
            if not row: self._send(404, {"error": "Usuario no encontrado"}); return
            if row["role"] == "admin":
                cnt = conn.execute(
                    "SELECT COUNT(*) FROM users WHERE role='admin' AND active=1"
                ).fetchone()[0]
                if cnt <= 1:
                    self._send(400, {"error": "No puedes eliminar el único administrador"}); return
            conn.execute("DELETE FROM users WHERE id=?", (uid,))
            conn.commit()
        self._send(200, {"ok": True})

    def h_zone_trees_get(self, qs, m):
        zid = m.group(1)
        with db() as conn:
            row = conn.execute(
                "SELECT trees, loaded_at FROM zone_trees WHERE zone_id=?", (zid,)
            ).fetchone()
        if not row:
            self._send(404, {"error": "no trees"}); return
        self._send(200, {
            "trees": _parse_json_field(row["trees"], []),
            "loaded_at": row["loaded_at"]
        })

    def h_zone_trees_put(self, qs, m):
        zid = m.group(1)
        body = self._body()
        trees = body.get("trees", [])
        with db() as conn:
            conn.execute(
                "INSERT INTO zone_trees (zone_id, trees) VALUES (?, ?) "
                "ON CONFLICT(zone_id) DO UPDATE SET trees=excluded.trees, "
                "loaded_at=datetime('now')",
                (zid, json.dumps(trees))
            )
            conn.commit()
        self._send(200, {"ok": True, "count": len(trees)})

    def h_db_query(self, qs, m):
        '''Execute an arbitrary SELECT query (used by the AI assistant).'''
        u = self._current_user
        if not u:
            self._send(401, {"error": "unauthenticated"}); return
        b = self._body()
        sql = (b.get("sql") or "").strip()
        if not sql:
            self._send(400, {"error": "sql requerido"}); return
        # Restrict to SELECT only
        if not re.match(r'^\s*SELECT\b', sql, re.IGNORECASE):
            self._send(400, {"error": "Solo se permiten consultas SELECT"}); return
        try:
            cols, rows = _safe_db_query(sql, b.get("params") or [])
            self._send(200, {"columns": cols, "rows": rows, "count": len(rows)})
        except Exception as ex:
            self._send(400, {"error": str(ex)})


def run():
    init_db()
    port = int(os.environ.get("PORT", 8733))
    from socketserver import ThreadingMixIn
    class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
        daemon_threads = True
    server = ThreadedHTTPServer(("", port), Handler)
    print(f"[SALVI GIS] Servidor escuchando en http://localhost:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[SALVI GIS] Servidor detenido.")


if __name__ == "__main__":
    run()
