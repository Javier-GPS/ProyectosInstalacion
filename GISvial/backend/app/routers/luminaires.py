"""Luminaires — CRUD, bulk, export, delete + inventory."""
import io
import time
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from sqlalchemy.orm import Session

from ..core.database import get_db
from ..models import (
    GisLuminaire, GisInventoryLuminaire, GisZone, GisZoneConfig,
    User, ensure_gis_tables,
)
from ..schemas.luminaires import GisImportInventoryBody
from .deps import current_user

router = APIRouter()

# ── Pending import cache ──────────────────────────────────────────────────
_PENDING_IMPORTS: dict[str, dict] = {}


def _row2dict(r) -> dict | None:
    if r is None:
        return None
    return {c.name: getattr(r, c.name) for c in r.__table__.columns}


def _rows2list(rows) -> list[dict]:
    return [_row2dict(r) for r in rows]


def _fval(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _sval(v):
    return str(v).strip() if v is not None else None


# ── Luminaires CRUD ──────────────────────────────────────────────────────
@router.get("/api/luminaires")
async def gis_luminaires_list(
    zone_id: Optional[str] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    ensure_gis_tables()
    query = db.query(GisLuminaire)
    if zone_id:
        query = query.filter(GisLuminaire.zone_id == zone_id)
    return _rows2list(query.order_by(GisLuminaire.id).all())


@router.post("/api/luminaires/bulk")
async def gis_luminaires_bulk(body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    zone_id = body.get("zone_id") or (body.get("luminaires", [{}])[0].get("zone_id") if body.get("luminaires") else None)
    if not zone_id:
        raise HTTPException(status_code=400, detail="zone_id required")
    luminaires = body.get("luminaires", [])
    if not luminaires:
        return {"ok": True, "count": 0}
    db.query(GisLuminaire).filter(GisLuminaire.zone_id == zone_id).delete()
    for item in luminaires:
        db.add(GisLuminaire(
            project_id=item.get("project_id"), zone_id=zone_id,
            road_type=item.get("road_type"), lighting_class=item.get("lighting_class"),
            street_name=item.get("street_name"), lat=item["lat"], lon=item["lon"],
            watts=_fval(item.get("watts")), spacing=_fval(item.get("spacing")),
            tilt=_fval(item.get("tilt")),
            height_m=_fval(item.get("height_m") or item.get("height")),
            arm_len=_fval(item.get("arm_len") or item.get("arm_length")),
            distribution=item.get("distribution"),
        ))
    db.commit()
    return {"ok": True, "count": len(luminaires)}


@router.delete("/api/luminaires/{zone_id}/{lum_id}")
async def gis_luminaires_delete(zone_id: str, lum_id: int, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    lum = db.get(GisLuminaire, lum_id)
    if not lum or lum.zone_id != zone_id:
        raise HTTPException(status_code=404, detail="Luminaire not found")
    db.delete(lum)
    db.commit()
    return {"ok": True}


# ── Luminaires export ────────────────────────────────────────────────────
@router.get("/api/luminaires/export")
async def gis_luminaires_export(
    zone_id: str = Query(...), user: User = Depends(current_user), db: Session = Depends(get_db),
):
    import openpyxl
    ensure_gis_tables()
    luminaires = db.query(GisLuminaire).filter(GisLuminaire.zone_id == zone_id).order_by(GisLuminaire.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Luminarias"
    ws.append(["ID", "Zona", "Tipo via", "Clase ilum.", "Calle", "Lat", "Lon", "W", "Espaciado", "Colocada"])
    for r in luminaires:
        ws.append([r.id, r.zone_id, r.road_type, r.lighting_class, r.street_name,
                   r.lat, r.lon, r.watts, r.spacing,
                   r.placed_at.isoformat() if r.placed_at else ""])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="luminaires_{zone_id}.xlsx"'},
    )


# ── Inventory ────────────────────────────────────────────────────────────
@router.get("/api/zones/{zone_id}/inventory")
async def gis_zone_inventory(zone_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    rows = db.query(GisInventoryLuminaire).filter(
        GisInventoryLuminaire.zone_id == zone_id
    ).order_by(GisInventoryLuminaire.id).all()
    return _rows2list(rows)


@router.post("/api/parse/inventory_excel")
async def gis_parse_inventory(request: Request):
    import openpyxl
    raw = await request.body()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise HTTPException(status_code=400, detail="Excel vacio")
    headers = [str(h).lower().strip() if h else "" for h in rows_iter[0]]
    FIELD_KEYS = {
        'lat': [' lat', 'latitud', 'latitude', 'coord_y', 'y_coord', 'coordenada y'],
        'lon': [' lon', ' lng', 'longitud', 'longitude', 'coord_x', 'x_coord', 'coordenada x'],
        'power_w': [' w ', 'potencia', 'power', 'watt', 'kw'],
        'height_m': ['altura', 'height', 'alto', 'h_m', 'h (m)'],
        'brand': ['marca', 'brand', 'fabricante'],
        'model': ['modelo', 'model', 'tipo'],
        'lamp_type': ['lampara', 'lamp', 'tecnologia', 'tech'],
        'point_id': ['id punto', 'punto_id', 'point_id', 'cod', 'codigo', 'num'],
        'circuit_id': ['cuadro', 'panel', 'circuit', 'armario'],
        'line_id': ['linea', 'ramal', 'line', 'circuito'],
        'support_type': ['soporte', 'support', 'bac', 'columna', 'poste'],
    }
    col_map = {}
    for i, h in enumerate(headers):
        hl = f" {h.lower().strip()}"
        for field, aliases in FIELD_KEYS.items():
            if field not in col_map and any(k in hl for k in aliases):
                col_map[field] = i
                break
    temp_id = uuid.uuid4().hex[:12]
    _PENDING_IMPORTS[temp_id] = {"rows": [list(r) for r in rows_iter[1:]], "headers": headers, "ts": time.time()}
    sample_rows = []
    for row in rows_iter[1:11]:
        if any(row):
            sample_rows.append([str(v) if v is not None else "" for v in row])
    return {
        "temp_id": temp_id, "filename": "upload.xlsx", "total_rows": len(rows_iter) - 1,
        "detected_columns": col_map, "headers": headers,
        "sample_rows": sample_rows, "field_order": list(FIELD_KEYS.keys()),
    }


@router.post("/api/import/inventory")
async def gis_import_inventory(body: GisImportInventoryBody, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    pending = _PENDING_IMPORTS.pop(body.temp_id, None)
    if not pending:
        raise HTTPException(status_code=400, detail="temp_id no valido o expirado")
    if time.time() - pending["ts"] > 1800:
        raise HTTPException(status_code=400, detail="Datos expirados, sube el archivo otra vez")
    rows_raw = pending["rows"]
    mapping = body.mapping

    def _val(row, field):
        idx = mapping.get(field)
        if idx is not None and idx < len(row):
            v = row[idx]
            return str(v).strip() if v is not None else None
        return None

    def _num(row, field):
        v = _val(row, field)
        if v is None or v == "":
            return None
        try:
            return float(v.replace(",", ".").replace(" ", ""))
        except ValueError:
            return None

    zid = uuid.uuid4().hex[:12]
    zone = GisZone(
        id=zid, name=body.zone_name.strip(), color=body.color or "#4caf82",
        source="inventory", project_id=body.project_id,
    )
    db.add(zone)
    db.add(GisZoneConfig(zone_id=zid))
    db.flush()

    count = 0
    for row in rows_raw:
        if not any(row):
            continue
        lat = _num(row, "lat")
        lon = _num(row, "lon")
        if lat is None or lon is None:
            continue
        db.add(GisInventoryLuminaire(
            zone_id=zid, lat=lat, lon=lon,
            point_id=_val(row, "point_id"),
            power_w=_num(row, "power_w"), height_m=_num(row, "height_m"),
            brand=_val(row, "brand"), model=_val(row, "model"),
            lamp_type=_val(row, "lamp_type"),
            support_type=_val(row, "support_type"),
            circuit_id=_val(row, "circuit_id"), line_id=_val(row, "line_id"),
        ))
        count += 1
    db.commit()
    db.refresh(zone)
    return {"zone": {
        "id": zone.id, "name": zone.name, "color": zone.color,
        "source": zone.source, "project_id": zone.project_id,
        "corridors": [], "bounds_polygon": [],
    }, "count": count}
