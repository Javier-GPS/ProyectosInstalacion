"""Plantilla export service — Excel for LuxStudio."""
import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def build_plantilla(zone_id: str, rows: list[dict]) -> bytes:
    """Build the 'plantilla luminotecnica' Excel workbook. Returns raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla luminotecnica"

    HEADERS = [
        "Zona", "Tipo via", "Longitud (km)", "N luminarias", "Disposicion",
        "Clase ilum.", "Ancho calzada (m)", "Espaciado (m)",
        "Fabricante", "Gama", "Potencia (W)", "Difusor", "Lente", "Tipo LED",
        "CCT (K)", "CRI", "Brazo (m)", "Inclinacion", "Pavimento", "MF",
    ]
    gold = "C8A96E"
    dark = "1C1C1A"
    hfill = PatternFill("solid", fgColor=gold)
    hfont = Font(bold=True, color=dark, size=10)
    thin = Side(style="thin", color="2A3A4A")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bord

    for r in rows:
        ws.append([
            zone_id, r.get("name", ""), r.get("description", ""), "",
            r.get("arrangement", ""), r.get("lighting_class", ""),
            r.get("road_width", ""), r.get("spacing", ""),
            r.get("manufacturer", "Salvi"), r.get("gama", ""),
            r.get("power", "") or "",
            r.get("difusor", ""), r.get("lente", ""), r.get("led_type", ""),
            r.get("cct", ""), r.get("cri", ""), r.get("arm_length", ""),
            r.get("tilt", ""), r.get("pavement", ""), r.get("mf", ""),
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
