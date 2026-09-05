from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class ExcelHandler:
    """Manejador de archivos Excel para lectura y escritura"""

    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_study_from_form(self, form_data, output_path):
        """Crea un Excel con los datos del formulario manual"""
        wb = Workbook()

        # Hoja 1: Datos del Proyecto
        ws_proyecto = wb.active
        ws_proyecto.title = "Proyecto"

        self._write_header(ws_proyecto, "DATOS DEL PROYECTO")

        row = 3
        fields = [
            ('Nombre del proyecto', form_data.get('nombre_proyecto')),
            ('Cliente final', form_data.get('cliente_final')),
            ('Localización', form_data.get('localizacion')),
            ('Proyectista', form_data.get('proyectista')),
            ('Fecha del estudio', form_data.get('fecha_estudio')),
            ('Nº de referencia', form_data.get('nro_referencia')),
            ('Norma aplicable', form_data.get('norma_aplicable')),
            ('Notas', form_data.get('notas')),
        ]

        for label, value in fields:
            ws_proyecto[f'A{row}'] = label
            ws_proyecto[f'B{row}'] = value
            self._style_cell(ws_proyecto[f'A{row}'], bold=True)
            row += 1

        # Hoja 2: Geometría
        ws_geom = wb.create_sheet("Geometría")
        self._write_header(ws_geom, "GEOMETRÍA DE LA VÍA")

        row = 3
        geom_fields = [
            ('Identificador modelo', form_data.get('identificador_modelo')),
            ('Disposición', form_data.get('disposicion')),
            ('Altura montaje (m)', form_data.get('altura_montaje')),
            ('Interdistancia (m)', form_data.get('interdistancia')),
            ('Saliente brazo (m)', form_data.get('saliente_brazo')),
            ('Inclinación brazo (°)', form_data.get('inclinacion_brazo')),
        ]

        for label, value in geom_fields:
            ws_geom[f'A{row}'] = label
            ws_geom[f'B{row}'] = value
            self._style_cell(ws_geom[f'A{row}'], bold=True)
            row += 1

        # Subsección: Calzada 1
        row += 1
        ws_geom[f'A{row}'] = "CALZADA 1"
        self._style_cell(ws_geom[f'A{row}'], bold=True, bg_color='E8E8E8')
        row += 1

        calzada1_fields = [
            ('Ancho (m)', form_data.get('calzada1_ancho')),
            ('Nº carriles', form_data.get('calzada1_carriles')),
            ('Tipo pavimento', form_data.get('calzada1_pavimento')),
            ('Q0', form_data.get('calzada1_q0')),
            ('Clase', form_data.get('calzada1_clase')),
        ]

        for label, value in calzada1_fields:
            ws_geom[f'A{row}'] = label
            ws_geom[f'B{row}'] = value
            self._style_cell(ws_geom[f'A{row}'], bold=True)
            row += 1

        # Hoja 3: Luminarias
        ws_lum = wb.create_sheet("Luminarias")
        self._write_header(ws_lum, "LUMINARIAS")

        row = 3
        headers = ['Modelo', 'Óptica LDT', 'Potencia (W)', 'Objetivo']
        for col, header in enumerate(headers, 1):
            cell = ws_lum.cell(row=row, column=col, value=header)
            self._style_cell(cell, bold=True, bg_color='D3D3D3')

        row += 1

        luminarias = form_data.get('luminarias', [])
        for lum in luminarias:
            ws_lum.cell(row=row, column=1, value=lum.get('modelo'))
            ws_lum.cell(row=row, column=2, value=lum.get('optica'))
            ws_lum.cell(row=row, column=3, value=lum.get('potencia'))
            ws_lum.cell(row=row, column=4, value=lum.get('objetivo'))
            row += 1

        # Hoja 4: Energía & Ambiental
        ws_energia = wb.create_sheet("Energía")
        self._write_header(ws_energia, "PARÁMETROS ENERGÍA Y AMBIENTAL")

        row = 3
        energia_fields = [
            ('Horas funcionamiento/año', form_data.get('horas_funcionamiento')),
            ('Tarifa eléctrica (€/kWh)', form_data.get('tarifa_electrica')),
            ('Factor CO2 (kg/kWh)', form_data.get('factor_co2')),
            ('Zona ambiental CIE 150', form_data.get('zona_ambiental')),
            ('ULOR máximo (%)', form_data.get('ulor_maximo')),
            ('Sensibilidad fauna', form_data.get('sensibilidad_fauna')),
        ]

        for label, value in energia_fields:
            ws_energia[f'A{row}'] = label
            ws_energia[f'B{row}'] = value
            self._style_cell(ws_energia[f'A{row}'], bold=True)
            row += 1

        # Ajustar ancho de columnas
        for ws in [ws_proyecto, ws_geom, ws_lum, ws_energia]:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40

        wb.save(output_path)

    def process_imported_excel(self, excel_path):
        """Procesa Excel importado y extrae datos de estudios"""
        wb = load_workbook(excel_path)

        if 'Plantilla' not in wb.sheetnames:
            raise ValueError("Hoja 'Plantilla' no encontrada en el Excel")

        ws = wb['Plantilla']
        results = []

        # Leer encabezados de la primera fila
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)

        if not headers:
            raise ValueError("No se encontraron encabezados en la hoja 'Plantilla'")

        # Mapeo de columnas esperadas a índices
        column_map = {header: idx for idx, header in enumerate(headers)}

        # Leer datos de la plantilla (asume formato específico)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:  # Fin de datos
                break

            try:
                study = {
                    'identificador': row[column_map.get('Identificador modelo', 0)],
                    'disposicion': row[column_map.get('Disposición de las luminarias', 1)],
                    'altura_montaje': float(row[column_map.get('Altura de montaje (h)', 2)] or 0),
                    'interdistancia': float(row[column_map.get('Interdistancia (d)', 3)] or 0),
                    'ancho_calzada1': float(row[column_map.get('Ancho de calzada 1 (W1)', 4)] or 0),
                    'clase_calzada': row[column_map.get('Clase calzada', 5)],
                    'modelo_luminaria': row[column_map.get('Modelo luminaria', 6)],
                    'optica_ldt': row[column_map.get('Óptica / código LDT', 7)],
                    'potencia': float(row[column_map.get('Potencia nominal', 8)] or 0),
                    'q0': float(row[column_map.get('Q0 (lux.m2)', 9)] or 0),
                    'pavimento': row[column_map.get('Tipo pavimento', 10)] or 'Asfalto',
                }

                # Validar que tenga identificador y modelo de luminaria
                if not study['identificador'] or not study['modelo_luminaria']:
                    continue

                results.append(study)
            except (ValueError, TypeError) as e:
                # Saltar filas con errores de conversión
                continue

        return results

    def create_results_excel(self, results, output_path):
        """Crea Excel de resultados después de procesar importación"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        self._write_header(ws, "RESULTADOS DEL CÁLCULO FOTOMÉTRICO")

        row = 3
        headers = ['Identificador', 'Modelo Luminaria', 'Altura (m)', 'Clase', 'Disposición',
                   'Potencia (W)', 'Ancho Calzada (m)', 'Pavimento']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._style_cell(cell, bold=True, bg_color='D3D3D3')

        row += 1

        for result in results:
            ws.cell(row=row, column=1, value=result.get('identificador'))
            ws.cell(row=row, column=2, value=result.get('modelo_luminaria'))
            ws.cell(row=row, column=3, value=result.get('altura_montaje'))
            ws.cell(row=row, column=4, value=result.get('clase_calzada'))
            ws.cell(row=row, column=5, value=result.get('disposicion'))
            ws.cell(row=row, column=6, value=result.get('potencia'))
            ws.cell(row=row, column=7, value=result.get('ancho_calzada1'))
            ws.cell(row=row, column=8, value=result.get('pavimento'))
            row += 1

        # Ajustar ancho de columnas
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 18

        wb.save(output_path)

    def _write_header(self, ws, title):
        """Escribe encabezado en hoja"""
        ws['A1'] = title
        self._style_cell(ws['A1'], bold=True, font_size=14, bg_color='4472C4', font_color='FFFFFF')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 25

    def _style_cell(self, cell, bold=False, bg_color=None, font_color='000000', font_size=11):
        """Aplica estilos a una celda"""
        cell.font = Font(bold=bold, color=font_color, size=font_size)

        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = self.thin_border
