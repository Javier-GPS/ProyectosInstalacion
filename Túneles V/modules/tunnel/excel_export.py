"""
SALVI Tunnel Engine — Exportador Excel (CIE 88:2004)
Genera un .xlsx multi-hoja a partir del resultado completo del motor.
Requiere: openpyxl >= 3.0
"""

import io
import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

C_BRAND      = "1A3A6B"
C_BRAND_MED  = "1A56B0"
C_BRAND_LT   = "D5E8F5"
C_WHITE      = "FFFFFF"
C_GRAY_H     = "F5F5F5"
C_BLACK      = "000000"

C_ZONE = {
    "threshold":  "FFF3CD",
    "threshold_b":"FFF3CD",
    "transition": "CFF4FC",
    "transition_b":"CFF4FC",
    "interior":   "D1E7DD",
    "interior_base":"D1E7DD",
    "exit":       "E2D9F3",
    "access":     "F8D7DA",
    "parting":    "F8D7DA",
}

C_SCENE = {
    "sunny":    "FFFDE7",
    "normal":   "F5F5F5",
    "overcast": "ECEFF1",
    "dusk":     "FCE4EC",
    "night":    "E8EAF6",
}


def _font(bold=False, size=10, color=C_BLACK, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    s = Side(style="medium", color=C_BRAND_MED)
    return Border(bottom=s)

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _merge_header(ws, row, col_start, col_end, text,
                  bg=C_BRAND, fg=C_WHITE, size=11, bold=True):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font      = _font(bold=bold, size=size, color=fg)
    cell.fill      = _fill(bg)
    cell.alignment = _align("center")
    return row + 1

def _header_row(ws, row, headers, bg=C_BRAND_MED, fg=C_WHITE, start_col=1, size=9):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font      = _font(bold=True, size=size, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _thin_border()
    return row + 1

def _data_row(ws, row, values, bg=C_WHITE, bold_cols=None, start_col=1, size=9, color=C_BLACK):
    bold_cols = bold_cols or []
    for i, v in enumerate(values):
        col = start_col + i
        c   = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=(col in bold_cols), size=size, color=color)
        c.fill      = _fill(bg)
        c.alignment = _align("center" if isinstance(v, (int, float)) else "left")
        c.border    = _thin_border()
    return row + 1

def _kv_row(ws, row, key, value, bg=C_GRAY_H):
    c_k = ws.cell(row=row, column=1, value=key)
    c_k.font      = _font(bold=True, size=9, color=C_BRAND)
    c_k.fill      = _fill(bg)
    c_k.alignment = _align("left")
    c_k.border    = _thin_border()
    c_v = ws.cell(row=row, column=2, value=value)
    c_v.font      = _font(size=9)
    c_v.fill      = _fill(bg)
    c_v.alignment = _align("left")
    c_v.border    = _thin_border()
    return row + 1

def _section_title(ws, row, text, col_span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, size=10, color=C_BRAND)
    c.fill      = _fill(C_BRAND_LT)
    c.alignment = _align("left")
    c.border    = Border(bottom=Side(style="medium", color=C_BRAND_MED))
    return row + 1


# -- Hoja 1: Portada

def _build_portada(wb, result, params):
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False
    summary        = result.get("summary", {})
    classification = result.get("classification", {})
    speed          = result.get("speed", {})
    _set_col_widths(ws, {"A": 28, "B": 30})

    row = _merge_header(ws, 1, 1, 2, "SALVI TUNNEL ENGINE", size=18)
    ws.row_dimensions[1].height = 36
    row = _merge_header(ws, row, 1, 2, "Informe Tecnico de Iluminacion de Tuneles",
                        bg=C_BRAND_MED, size=12)
    ws.row_dimensions[row-1].height = 22
    row = _merge_header(ws, row, 1, 2,
                        "CIE 88:2004 — Guide for the Lighting of Road Tunnels and Underpasses",
                        bg=C_BRAND_LT, fg=C_BRAND, size=9, bold=False)
    row += 1

    row = _merge_header(ws, row, 1, 2, "DATOS DEL PROYECTO", bg=C_BRAND_MED, size=10)
    fields = [
        ("Proyecto",              summary.get("project", params.get("project_name", "—"))),
        ("Tubo",                  summary.get("tube_id", "—")),
        ("Longitud total",        f"{summary.get('length_m', '—')} m"),
        ("Velocidad de diseno",   f"{summary.get('speed_kmh', '—')} km/h"),
        ("Distancia de parada",   f"{speed.get('SD_m', summary.get('SD_m', '—'))} m"),
        ("Trafico de diseno",     f"{params.get('traffic_veh_h', '—')} veh/h"),
        ("Clasificacion optica",  classification.get("optical", "—").replace("_", " ")),
        ("Clasificacion geometrica", classification.get("geometric", "—").replace("_", " ")),
        ("Iluminacion diurna",    classification.get("daylighting", "—")),
        ("Fecha del informe",     datetime.date.today().strftime("%d/%m/%Y")),
    ]
    for k, v in fields:
        bg = C_GRAY_H if row % 2 == 0 else C_WHITE
        row = _kv_row(ws, row, k, v, bg=bg)

    row += 1
    row = _merge_header(ws, row, 1, 2, "LUMINANCIAS DE DISENO", bg=C_BRAND_MED, size=10)
    lum_fields = [
        ("L20 — Campo 20",     f"{summary.get('L20', '—')} cd/m2"),
        ("Lth — Umbral",       f"{summary.get('Lth', '—')} cd/m2"),
        ("Lin — Interior",     f"{summary.get('Lin', '—')} cd/m2"),
        (
            "L noche normal",
            f"{summary.get('L_night_normal', summary.get('Lin', '—'))} cd/m2",
        ),
        (
            "L noche reducida",
            f"{summary.get('L_night_reduced', summary.get('L_night', '—'))} cd/m2",
        ),
        ("Factor k",           summary.get("k_factor", "—")),
        ("Coeficiente qc",     summary.get("qc", "—")),
    ]
    for k, v in lum_fields:
        bg = C_GRAY_H if row % 2 == 0 else C_WHITE
        row = _kv_row(ws, row, k, v, bg=bg)

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1,
                value="Generado con SALVI Tunnel Engine (STE) — CIE 88:2004.")
    c.font      = _font(size=8, italic=True, color="888888")
    c.alignment = _align("center")


# -- Hoja 2: Zonas CIE 88

def _build_zonas(wb, result):
    ws = wb.create_sheet("Zonas CIE 88")
    ws.sheet_view.showGridLines = False
    zones_raw  = result.get("zones", {})
    zones_list = list(zones_raw.values()) if isinstance(zones_raw, dict) else zones_raw
    _set_col_widths(ws, {"A": 20, "B": 13, "C": 13, "D": 14, "E": 17, "F": 17, "G": 20})

    row = _merge_header(ws, 1, 1, 7, "ZONAS NORMATIVAS CIE 88:2004", size=13)
    row += 1
    row = _header_row(ws, row, [
        "Zona", "s inicio (m)", "s fin (m)", "Longitud (m)",
        "L inicio (cd/m2)", "L fin (cd/m2)", "L min. req. (cd/m2)",
    ])
    zone_labels = {
        "threshold":  "Umbral (CTH)",
        "transition": "Transicion (CTR)",
        "interior":   "Interior (CIN)",
        "exit":       "Salida (CEX)",
        "access":     "Acceso",
        "parting":    "Particion",
    }
    data_start = row
    for z in zones_list:
        z_type  = str(z.get("zone_type") or z.get("type") or "interior").lower()
        bg      = C_ZONE.get(z_type, C_WHITE)
        label   = zone_labels.get(z_type, z_type.replace("_", " ").title())
        L_start = float(z.get("L_start", z.get("L_min_required", 0)))
        L_end   = float(z.get("L_end",   z.get("L_min_required", 0)))
        length  = float(z.get("length", float(z.get("s_end", 0)) - float(z.get("s_start", 0))))
        _data_row(ws, row, [
            label,
            round(float(z.get("s_start", 0)), 1),
            round(float(z.get("s_end", 0)),   1),
            round(length, 1),
            round(L_start, 1), round(L_end, 1),
            round(float(z.get("L_min_required", 0)), 1),
        ], bg=bg, bold_cols=[1, 7])
        row += 1
    if zones_list:
        _data_row(ws, row,
                  ["TOTAL", "", "", f"=SUM(D{data_start}:D{row-1})", "", "", ""],
                  bg=C_BRAND_LT, bold_cols=[1, 4])
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Leyenda:").font = _font(bold=True, size=9, color=C_BRAND)
    row += 1
    for z_type, bg in C_ZONE.items():
        c = ws.cell(row=row, column=1, value=zone_labels.get(z_type, z_type))
        c.fill = _fill(bg); c.font = _font(size=9); c.border = _thin_border()
        c.alignment = _align("center")
        row += 1


# -- Hoja 3: Resultados

def _build_resultados(wb, result, params):
    ws = wb.create_sheet("Resultados")
    ws.sheet_view.showGridLines = False
    summary  = result.get("summary",  {})
    speed    = result.get("speed",    {})
    l20_data = result.get("l20",      {})
    lth_data = result.get("lth",      {})
    _set_col_widths(ws, {"A": 30, "B": 22, "C": 4, "D": 30, "E": 22})

    row = _merge_header(ws, 1, 1, 5, "RESULTADOS DEL CALCULO CIE 88:2004", size=13)
    row += 1

    row = _section_title(ws, row, "1. Velocidad de Diseno y Distancia de Parada", 5)
    for k, v in [
        ("Velocidad de diseno",    f"{summary.get('speed_kmh', '—')} km/h"),
        ("Distancia de parada SD", f"{speed.get('SD_m', summary.get('SD_m', '—'))} m"),
        ("Distancia de reaccion",  f"{speed.get('d_reaction_m', '—')} m"),
        ("Distancia de frenado",   f"{speed.get('d_braking_m', '—')} m"),
    ]:
        row = _kv_row(ws, row, k, v, bg=C_GRAY_H if row % 2 == 0 else C_WHITE)

    row = _section_title(ws, row + 1, "2. Luminancia Campo 20 (L20)", 5)
    for k, v in [
        ("L20",          f"{l20_data.get('L20', summary.get('L20', '—'))} cd/m2"),
        ("Metodo",       l20_data.get("method", "—")),
        ("Confianza",    l20_data.get("confidence", "—")),
        ("Lseq estimada",f"{summary.get('Lseq', '—')} cd/m2"),
    ]:
        row = _kv_row(ws, row, k, v, bg=C_GRAY_H if row % 2 == 0 else C_WHITE)

    row = _section_title(ws, row + 1, "3. Luminancia de Umbral (Lth)", 5)
    for k, v in [
        ("Lth",            f"{summary.get('Lth', '—')} cd/m2"),
        ("Factor k",       lth_data.get("k_factor", summary.get("k_factor", "—"))),
        ("Coef. qc",       lth_data.get("qc", summary.get("qc", "—"))),
        ("Lin (interior)", f"{summary.get('Lin', '—')} cd/m2"),
        (
            "L noche normal",
            f"{summary.get('L_night_normal', summary.get('Lin', '—'))} cd/m2",
        ),
        (
            "L noche reducida",
            f"{summary.get('L_night_reduced', summary.get('L_night', '—'))} cd/m2",
        ),
    ]:
        row = _kv_row(ws, row, k, v, bg=C_GRAY_H if row % 2 == 0 else C_WHITE)

    warnings = result.get("warnings", [])
    if warnings:
        row = _section_title(ws, row + 1, "Advertencias", 5)
        for w in warnings:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value=f"- {w}")
            c.font = _font(size=9, color="C0392B"); c.alignment = _align("left")
            row += 1


# -- Hoja 4: Plan de Control

def _build_control(wb, result):
    control = result.get("control")
    if not control:
        return
    ws = wb.create_sheet("Plan de Control")
    ws.sheet_view.showGridLines = False
    groups   = control.get("groups", [])
    scenes   = control.get("scenes", [])
    protocol = control.get("protocol", "DALI")
    if not groups or not scenes:
        return

    n_cols = 2 + len(scenes)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    for i in range(len(scenes)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    row = _merge_header(ws, 1, 1, n_cols, "PLAN DE CONTROL CIE 88:2004 §87-111", size=13)
    for k, v in [
        ("Protocolo", protocol),
        ("N grupos", control.get("n_groups", len(groups))),
        ("N escenas", control.get("n_scenes", len(scenes))),
    ]:
        row = _kv_row(ws, row, k, v)

    row += 1
    row = _merge_header(ws, row, 1, n_cols,
                        "Tabla de Regulacion (grupos x escenas)", bg=C_BRAND_MED, size=10)
    scene_headers = ["Grupo", "L diseno (cd/m2)"] + [
        f"{s['name']} / L20={s['L20']:.0f}" for s in scenes
    ]
    row = _header_row(ws, row, scene_headers)

    for g in groups:
        z_type   = g.get("zone_type", "interior").lower()
        bg       = C_ZONE.get(z_type, C_WHITE)
        row_vals = [g["name"], g.get("L_design", "—")]
        for scene in scenes:
            sid  = str(scene["scene_id"])
            pct  = g.get("dimming_levels", {}).get(sid, "—")
            dali = g.get("dali_levels", {}).get(sid, "")
            row_vals.append(f"{pct}%  /  DALI {dali}" if protocol == "DALI" and dali else f"{pct}%")
        _data_row(ws, row, row_vals, bg=bg, bold_cols=[1])
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1,
                value="CTH=Umbral - CTRn=Transicion - CIN=Interior - CEX=Salida. "
                      "Niveles DALI segun IEC 62386 (rango 0-254).")
    c.font = _font(size=8, italic=True, color="666666"); c.alignment = _align("left")

    reg_curves = control.get("regulation_curves", [])
    if reg_curves:
        cth_curve = next((c for c in reg_curves if "CTH" in c.get("group_name", "")),
                         reg_curves[0])
        if cth_curve and cth_curve.get("points"):
            row += 3
            row = _merge_header(ws, row, 1, 4, "Curva de Regulacion — Grupo CTH",
                                bg=C_BRAND_MED, size=10)
            row = _header_row(ws, row, ["L20 (cd/m2)", "Dimming (%)", "DALI level"])
            data_start = row
            for pt in cth_curve["points"]:
                ws.cell(row=row, column=1, value=pt.get("L20", ""))
                ws.cell(row=row, column=2, value=pt.get("dimming_pct", ""))
                ws.cell(row=row, column=3, value=pt.get("dali_level", ""))
                row += 1
            data_end = row - 1
            if data_end > data_start:
                chart = LineChart()
                chart.title = "Curva de Regulacion CTH"
                chart.style = 10
                chart.y_axis.title = "Dimming (%)"
                chart.x_axis.title = "L20 (cd/m2)"
                chart.height = 12; chart.width = 18
                chart.add_data(Reference(ws, min_col=2, min_row=data_start-1, max_row=data_end),
                               titles_from_data=True)
                chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
                ws.add_chart(chart, "F" + str(data_start))


# -- Hoja 5: Luminarias

def _build_luminarias(wb, lum_result):
    if not lum_result:
        return
    ws = wb.create_sheet("Luminarias")
    ws.sheet_view.showGridLines = False
    lum    = lum_result.get("luminaire", {})
    zones  = lum_result.get("zones", [])
    totals = lum_result.get("totals", {})
    road_s = lum_result.get("road_surface", {})
    _set_col_widths(ws, {"A": 16, "B": 14, "C": 16, "D": 18, "E": 18, "F": 16, "G": 14, "H": 18, "I": 18})

    row = _merge_header(ws, 1, 1, 9, "CALCULO DE LUMINARIAS POR ZONA", size=13)
    row += 1
    row = _section_title(ws, row, "Luminaria de Diseno", 4)
    for k, v in [
        ("Referencia/Modelo",       lum.get("name", "—")),
        ("Flujo luminoso",          f"{lum.get('flux_lm', '—')} lm"),
        ("Potencia",                f"{lum.get('power_w', '—')} W"),
        ("Eficacia luminosa",       f"{lum.get('efficacy_lm_w', '—')} lm/W"),
        ("Factor de utilizacion",   lum.get("efficiency", "—")),
        ("Disposicion",             lum.get("arrangement_label", "—")),
        ("Factor de mantenimiento", lum.get("maintenance_factor", "—")),
        ("Superficie calzada",      road_s.get("label", "—")),
        ("Coef. rho_eff",           f"{road_s.get('rho', '—')} cd/m2/lux"),
    ]:
        row = _kv_row(ws, row, k, v, bg=C_GRAY_H if row % 2 == 0 else C_WHITE)

    row += 1
    row = _section_title(ws, row, "Diseno por Zona", 9)
    row = _header_row(ws, row, [
        "Zona", "Longitud (m)", "L req. (cd/m2)", "E req. (lux)",
        "d max. (m)", "d usado (m)", "N luminarias",
        "Potencia zona (W)", "Densidad (W/m2)",
    ])
    data_start = row
    for z in zones:
        z_type = z.get("zone_type", "interior").lower()
        bg     = C_ZONE.get(z_type, C_WHITE)
        _data_row(ws, row, [
            z.get("zone_name", ""), z.get("zone_length", 0),
            z.get("L_required", 0), z.get("E_required", 0),
            z.get("d_max", 0), z.get("d_used", 0),
            z.get("n_luminaires", 0), z.get("power_zone_w", 0),
            z.get("power_density_wm2", 0),
        ], bg=bg, bold_cols=[1, 7])
        row += 1
    data_end = row - 1
    if zones:
        _data_row(ws, row, [
            "TOTAL", f"=SUM(B{data_start}:B{data_end})", "", "", "", "",
            f"=SUM(G{data_start}:G{data_end})",
            f"=SUM(H{data_start}:H{data_end})", "",
        ], bg=C_BRAND_LT, bold_cols=[1, 7, 8])
        row += 1

    row += 1
    row = _section_title(ws, row, "Resumen Instalacion", 4)
    for k, v in [
        ("Total luminarias",     totals.get("n_luminaires", 0)),
        ("Potencia total",       f"{totals.get('power_kw', 0)} kW"),
        ("Densidad media",       f"{totals.get('power_density_wm2', 0)} W/m2"),
        ("Flujo instalado",      f"{totals.get('flux_lm', 0):,.0f} lm"),
    ]:
        row = _kv_row(ws, row, k, v, bg=C_GRAY_H if row % 2 == 0 else C_WHITE)

    if zones:
        row += 2
        ws.cell(row=row, column=1, value="Zona").font = _font(bold=True, size=8)
        ws.cell(row=row, column=2, value="Luminarias").font = _font(bold=True, size=8)
        row += 1
        chart_start = row
        for z in zones:
            ws.cell(row=row, column=1, value=z.get("zone_name", ""))
            ws.cell(row=row, column=2, value=z.get("n_luminaires", 0))
            row += 1
        chart_end = row - 1
        chart = BarChart()
        chart.type = "col"; chart.title = "Luminarias por zona"
        chart.style = 10; chart.height = 10; chart.width = 18
        chart.y_axis.title = "N luminarias"
        chart.add_data(Reference(ws, min_col=2, min_row=chart_start-1, max_row=chart_end),
                       titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end))
        ws.add_chart(chart, "D" + str(chart_start - 1))


# -- Hoja 6: Lista de luminarias físicas

def _build_lista_luminarias(wb, lum_result, result=None):
    """Inventario para obra: una fila por luminaria física, no por estación."""
    if not lum_result:
        return
    lum = lum_result.get("luminaire", {}) or {}
    zones = lum_result.get("zones", []) or []
    if not zones:
        return
    result = result or {}
    control = result.get("control", {}) or {}
    scenes = control.get("scenes", []) or []
    groups = control.get("groups", []) or []
    tube_id = lum_result.get("tube_id", result.get("summary", {}).get("tube_id", "T1"))
    arrangement = str(lum_result.get("arrangement", lum.get("arrangement", "central_single")) or "central_single")
    road_width = float(lum_result.get("road_width_m", 0) or 0)
    wall_offset = float(lum.get("wall_offset_m", 0.30) or 0.30)
    mounting_height = float(lum.get("mounting_height_m", 0) or 0)

    def physical_rows(station_index):
        """Coordenadas desde la pared izquierda, idénticas al motor CIE 140."""
        offset = min(max(0.05, wall_offset), max(0.05, road_width / 2 - 0.05))
        if arrangement in {"central_double", "bilateral_sym", "bilateral"}:
            return [(offset, "Izquierda"), (road_width - offset, "Derecha")]
        if arrangement in {"bilateral_stag", "staggered"}:
            return [(offset if station_index % 2 else road_width - offset,
                     "Izquierda" if station_index % 2 else "Derecha")]
        if arrangement in {"central_offset", "lateral_left"}:
            return [(offset, "Izquierda")]
        if arrangement in {"lateral_right", "unilateral"}:
            return [(road_width - offset, "Derecha")]
        return [(road_width / 2, "Eje")]

    def dali_group_for(zone):
        layer = str(zone.get("control_layer", "") or "").lower()
        portal = zone.get("portal")
        candidates = [g for g in groups if str(g.get("layer", "")).lower() == layer]
        if portal is not None:
            portal_candidates = [g for g in candidates if g.get("portal") == portal]
            if portal_candidates:
                candidates = portal_candidates
        return candidates[0] if candidates else {}

    headers = [
        "ID luminaria", "Tubo", "Grupo DALI", "Nombre grupo DALI", "Capa control",
        "Zona", "Tipo zona", "Portal", "Estación", "Fila", "PK longitudinal (m)",
        "Y desde pared izquierda (m)", "Altura montaje (m)", "Disposición",
        "Modelo", "Óptica", "PCB", "Corriente día (mA)", "Potencia día (W)",
        "Flujo día (lm)", "Tilt (°)", "Interdistancia (m)", "L requerida (cd/m²)",
        "L total requerida (cd/m²)", "L estimada (cd/m²)", "U0", "Ul",
        "Corriente noche (mA)", "Potencia noche (W)", "Flujo noche (lm)",
        "L noche estimada (cd/m²)", "Driver floor noche",
    ]
    for scene in scenes:
        scene_name = str(scene.get("name", scene.get("scene_type", "Escena")))
        headers.extend([f"{scene_name} estado", f"{scene_name} (%)", f"{scene_name} DALI"])

    ws = wb.create_sheet("Lista Luminarias")
    ws.sheet_view.showGridLines = False
    row = _merge_header(ws, 1, 1, len(headers), "LISTA DE TODAS LAS LUMINARIAS — MONTAJE Y PUESTA EN MARCHA", size=13)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    note = ws.cell(row=row, column=1, value=(
        "Una fila corresponde a una luminaria física. La coordenada Y se mide desde la pared izquierda; "
        "los niveles DALI proceden del plan de control calculado."
    ))
    note.font = _font(size=8, italic=True, color="666666")
    note.alignment = _align("left", wrap=True)
    row += 2
    header_row = row
    row = _header_row(ws, row, headers, size=8)

    reference = 1
    for zone in zones:
        zone_type = str(zone.get("zone_type", "interior") or "interior").lower()
        bg = C_ZONE.get(zone_type, C_WHITE)
        dali_group = dali_group_for(zone)
        points = zone.get("setpoints", []) or []
        for point_position, point in enumerate(points, start=1):
            station = int(point.get("idx", point_position) or point_position)
            operating = point.get("scenario_operating_points", {}) or {}
            for transverse_y, row_name in physical_rows(station):
                values = [
                    f"{tube_id}-{reference:04d}", tube_id,
                    dali_group.get("group_id", "—"), dali_group.get("name", "—"),
                    zone.get("control_layer", "—"), zone.get("zone_name", "—"), zone_type,
                    zone.get("portal", point.get("portal", "—")), station, row_name,
                    point.get("s", "—"), round(transverse_y, 3), mounting_height, arrangement,
                    point.get("model", zone.get("model", "—")), point.get("optic", zone.get("optic", "—")),
                    point.get("pcb", zone.get("pcb", "—")), point.get("current_mA", "—"),
                    point.get("power_w", "—"), point.get("flux_lm", "—"), point.get("tilt_deg", zone.get("tilt_deg", "—")),
                    point.get("spacing_m", zone.get("d_used", "—")), point.get("L_req", zone.get("L_required", "—")),
                    point.get("L_total_req", zone.get("L_total_required", "—")), point.get("L_est", zone.get("L_estimated", "—")),
                    point.get("U0", "—"), point.get("Ul", "—"), point.get("night_current_mA", "—"),
                    point.get("night_power_w", "—"), point.get("night_flux_lm", "—"),
                    point.get("night_L_est", "—"), "Sí" if point.get("night_driver_floor") else "No",
                ]
                for scene in scenes:
                    scene_id = str(scene.get("scene_id", ""))
                    scene_type = str(scene.get("scene_type", ""))
                    scene_point = operating.get(scene_type, {}) or {}
                    values.extend([
                        scene_point.get("state", "—"),
                        dali_group.get("dimming_levels", {}).get(scene_id, "—"),
                        dali_group.get("dali_levels", {}).get(scene_id, "—"),
                    ])
                _data_row(ws, row, values, bg=bg, bold_cols=[1], size=8)
                row += 1
                reference += 1

    last_row = row - 1
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    widths = {
        1: 16, 2: 9, 3: 11, 4: 27, 5: 15, 6: 22, 7: 16, 8: 10, 9: 10, 10: 12,
        11: 18, 12: 24, 13: 18, 14: 18, 15: 20, 16: 12, 17: 10,
    }
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 15)
    ws.row_dimensions[2].height = 28


# -- Hoja 7: Perfil Longitudinal

def _build_perfil(wb, result):
    chart_data = result.get("chart", {})
    chart_data = chart_data.get("data", []) if isinstance(chart_data, dict) else chart_data
    if not chart_data:
        return
    ws = wb.create_sheet("Perfil Longitudinal")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 14, "B": 20})

    row = _merge_header(ws, 1, 1, 2, "PERFIL LONGITUDINAL DE LUMINANCIA", size=13)
    row += 1
    row = _header_row(ws, row, ["Posicion (m)", "L requerida (cd/m2)"])
    data_start = row

    pts = chart_data
    if len(pts) > 500:
        step = max(1, len(pts) // 500)
        pts = pts[::step]

    for pt in pts:
        ws.cell(row=row, column=1, value=round(float(pt.get("s", pt.get("x", pt.get("position", 0)))), 1))
        ws.cell(row=row, column=2, value=round(float(pt.get("L", pt.get("y", pt.get("L_required", 0)))), 2))
        row += 1
    data_end = row - 1

    chart = LineChart()
    chart.title = "Perfil de Luminancia — CIE 88:2004"
    chart.style = 10
    chart.y_axis.title = "L requerida (cd/m2)"
    chart.x_axis.title = "Posicion (m)"
    chart.height = 15; chart.width = 25
    chart.add_data(Reference(ws, min_col=2, min_row=data_start-1, max_row=data_end),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = C_BRAND_MED
    s.graphicalProperties.line.width     = 20000
    ws.add_chart(chart, "D2")


# ═══════════════════════════════════════════════════════════════════
# FUNCIONES PRINCIPALES
# ═══════════════════════════════════════════════════════════════════

def generate_excel_combined(tubes_data: list) -> bytes:
    """
    Genera un Excel combinado con multiples tubos, cada uno con sus hojas
    renombradas con el ID de tubo (ej. "Portada T1", "Zonas CIE T1", etc.).
    """
    if not tubes_data:
        raise ValueError("Se necesita al menos un tubo")

    if len(tubes_data) == 1:
        t = tubes_data[0]
        return generate_excel(t["result"], t.get("params"), t.get("lum_result"))

    wb = Workbook()

    for t in tubes_data:
        result     = t["result"]
        params     = t.get("params") or {}
        lum_result = t.get("lum_result")
        tube_id    = result.get("summary", {}).get("tube_id", "T?")

        # Generar workbook temporal para este tubo
        wb_tmp = Workbook()
        _build_portada(wb_tmp, result, params)
        _build_zonas(wb_tmp, result)
        _build_resultados(wb_tmp, result, params)
        _build_control(wb_tmp, result)
        _build_luminarias(wb_tmp, lum_result)
        _build_lista_luminarias(wb_tmp, lum_result, result)
        _build_perfil(wb_tmp, result)

        # Copiar hojas al workbook combinado con sufijo de tubo
        for ws_src in wb_tmp.worksheets:
            ws_dst = wb.create_sheet(title=f"{ws_src.title} {tube_id}"[:31])
            for row in ws_src.iter_rows():
                for cell in row:
                    dst_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        dst_cell.font      = cell.font.copy()
                        dst_cell.fill      = cell.fill.copy()
                        dst_cell.border    = cell.border.copy()
                        dst_cell.alignment = cell.alignment.copy()
            # Copiar merged cells
            for mc in ws_src.merged_cells.ranges:
                ws_dst.merge_cells(str(mc))
            # Copiar anchos de columna
            for col, cd in ws_src.column_dimensions.items():
                ws_dst.column_dimensions[col].width = cd.width
            for row, rd in ws_src.row_dimensions.items():
                ws_dst.row_dimensions[row].height = rd.height

    # Eliminar hoja vacia inicial
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if wb.worksheets:
        wb.active = wb.worksheets[0]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_excel(result: dict, params: dict = None, lum_result: dict = None) -> bytes:
    """
    Genera el libro Excel completo.
    """
    if params is None:
        params = {}
    wb = Workbook()
    _build_portada(wb, result, params)
    _build_zonas(wb, result)
    _build_resultados(wb, result, params)
    _build_control(wb, result)
    _build_luminarias(wb, lum_result)
    _build_lista_luminarias(wb, lum_result, result)
    _build_perfil(wb, result)
    wb.active = wb["Portada"]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
