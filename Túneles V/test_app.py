#!/usr/bin/env python
"""
Test simplificado para verificar que Flask funciona
Ejecuta: python test_app.py
"""

from flask import Flask, jsonify, request, render_template, send_file
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib import colors
import json
from io import BytesIO
from photometry import calcular_fotometria
from visualization import generar_graficos_isocurvas
from ldt_reader import obtener_lista_luminarias_json

app = Flask(__name__)

# Carpetas
DOWNLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'downloads')
ASSETS_FOLDER = os.path.join(os.path.dirname(__file__), 'assets')
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
os.makedirs(ASSETS_FOLDER, exist_ok=True)

print(f"\n✓ DOWNLOAD_FOLDER: {DOWNLOAD_FOLDER}")
print(f"✓ ASSETS_FOLDER: {ASSETS_FOLDER}\n")

# Pre-cargar lista de LDTs al inicio
LDTS_CACHE = None

def cargar_ldts_cache():
    """Carga los LDTs en caché al inicio del servidor"""
    global LDTS_CACHE
    ldts_zip = os.path.join(ASSETS_FOLDER, 'LDTs_luminarias.zip')

    if os.path.exists(ldts_zip):
        print(f"✓ Leyendo LDTs desde: {ldts_zip}")
        LDTS_CACHE = obtener_lista_luminarias_json(ldts_zip)
        print(f"✓ {LDTS_CACHE.get('total', 0)} luminarias cargadas")
    else:
        print(f"⚠️ No se encontró ZIP de LDTs: {ldts_zip}")
        LDTS_CACHE = {
            'success': False,
            'luminarias': [],
            'error': 'No se encontró el archivo LDTs_luminarias.zip'
        }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/test-simple')
def test_simple():
    return render_template('test_simple.html')

@app.route('/debug')
def debug():
    return render_template('debug.html')

@app.route('/api/luminarias', methods=['GET'])
def obtener_luminarias():
    """Obtiene lista de luminarias disponibles desde LDTs"""
    global LDTS_CACHE

    if LDTS_CACHE is None:
        print("DEBUG: LDTS_CACHE es None, cargando...")
        cargar_ldts_cache()

    print(f"DEBUG: Devolviendo {LDTS_CACHE.get('total', 0)} luminarias")
    return jsonify(LDTS_CACHE), 200

@app.route('/api/debug/ldts', methods=['GET'])
def debug_ldts():
    """Debug endpoint para LDTs"""
    global LDTS_CACHE

    ldts_zip = os.path.join(ASSETS_FOLDER, 'LDTs_luminarias.zip')

    return jsonify({
        'status': 'ok',
        'ldts_zip_path': ldts_zip,
        'ldts_zip_exists': os.path.exists(ldts_zip),
        'ldts_cache_loaded': LDTS_CACHE is not None,
        'ldts_cache': LDTS_CACHE,
        'assets_folder': ASSETS_FOLDER,
        'assets_contents': os.listdir(ASSETS_FOLDER) if os.path.exists(ASSETS_FOLDER) else []
    }), 200

@app.route('/api/test', methods=['GET'])
def test():
    """Endpoint de prueba"""
    return jsonify({
        'status': 'ok',
        'message': 'Servidor respondiendo correctamente',
        'timestamp': datetime.now().isoformat()
    }), 200

@app.route('/api/download-template', methods=['GET'])
def download_template():
    """Descarga plantilla Excel"""
    try:
        filename = 'plantilla_app_salvilux.xlsx'
        filepath = os.path.join(ASSETS_FOLDER, filename)

        print(f"DEBUG: Buscando plantilla en: {filepath}")
        print(f"DEBUG: Existe: {os.path.exists(filepath)}")

        if not os.path.exists(filepath):
            # Si no existe, buscar en rutas alternativas
            alt_paths = [
                os.path.join(os.path.dirname(__file__), 'assets', filename),
                os.path.join(os.getcwd(), 'assets', filename),
            ]
            for alt_path in alt_paths:
                if os.path.exists(alt_path):
                    filepath = alt_path
                    print(f"DEBUG: Encontrada en: {filepath}")
                    break
            else:
                print(f"ERROR: Plantilla no encontrada")
                print(f"DEBUG: Archivos en assets: {os.listdir(ASSETS_FOLDER) if os.path.exists(ASSETS_FOLDER) else 'NO EXISTE'}")
                return jsonify({'error': 'Plantilla no encontrada'}), 404

        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"ERROR en download_template: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-ldts', methods=['GET'])
def download_ldts():
    """Descarga LDTs"""
    try:
        filename = 'LDTs_luminarias.zip'
        filepath = os.path.join(ASSETS_FOLDER, filename)

        print(f"DEBUG: Buscando LDTs en: {filepath}")
        print(f"DEBUG: Existe: {os.path.exists(filepath)}")

        if not os.path.exists(filepath):
            # Si no existe, buscar en rutas alternativas
            alt_paths = [
                os.path.join(os.path.dirname(__file__), 'assets', filename),
                os.path.join(os.getcwd(), 'assets', filename),
            ]
            for alt_path in alt_paths:
                if os.path.exists(alt_path):
                    filepath = alt_path
                    print(f"DEBUG: Encontrada en: {filepath}")
                    break
            else:
                print(f"ERROR: LDTs no encontrada")
                print(f"DEBUG: Archivos en assets: {os.listdir(ASSETS_FOLDER) if os.path.exists(ASSETS_FOLDER) else 'NO EXISTE'}")
                return jsonify({'error': 'LDTs no encontrada'}), 404

        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/zip'
        )
    except Exception as e:
        print(f"ERROR en download_ldts: {str(e)}")
        return jsonify({'error': str(e)}), 500

def crear_excel_completo(data):
    """Crea un Excel con todos los datos del formulario"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudio Fotométrico"

    # Estilos
    titulo_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # TÍTULO
    ws[f'A{row}'] = "CÁLCULO FOTOMÉTRICO SALVI"
    ws[f'A{row}'].font = titulo_font
    ws[f'A{row}'].fill = titulo_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 2

    # SECCIÓN 1: DATOS DEL PROYECTO
    ws[f'A{row}'] = "DATOS DEL PROYECTO"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    datos_proyecto = [
        ("Nombre del proyecto", data.get('nombre_proyecto', '')),
        ("Cliente final", data.get('cliente_final', '')),
        ("Localización", data.get('localizacion', '')),
        ("Proyectista", data.get('proyectista', '')),
        ("Nº Referencia", data.get('nro_referencia', '')),
        ("Fecha del estudio", data.get('fecha_estudio', '')),
        ("Norma aplicable", data.get('norma_aplicable', '')),
        ("Notas", data.get('notas', '')),
    ]

    for label, value in datos_proyecto:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # SECCIÓN 2: GEOMETRÍA
    ws[f'A{row}'] = "GEOMETRÍA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    datos_geom = [
        ("Identificador modelo", data.get('identificador_modelo', '')),
        ("Disposición", data.get('disposicion', '')),
        ("Altura montaje [m]", data.get('altura_montaje', '')),
        ("Interdistancia [m]", data.get('interdistancia', '')),
        ("Saliente brazo [m]", data.get('saliente_brazo', '')),
        ("Inclinación brazo [°]", data.get('inclinacion_brazo', '')),
    ]

    for label, value in datos_geom:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # CALZADA
    ws[f'A{row}'] = "Calzada Principal"
    ws[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    datos_calzada = [
        ("Ancho calzada [m]", data.get('calzada1_ancho', '')),
        ("Nº carriles", data.get('calzada1_carriles', '')),
        ("Tipo pavimento", data.get('calzada1_pavimento', '')),
        ("q₀ pavimento", data.get('calzada1_q0', '')),
        ("Clase de iluminación", data.get('calzada1_clase', '')),
    ]

    for label, value in datos_calzada:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # ACERA
    ws[f'A{row}'] = "Acera"
    ws[f'A{row}'].font = Font(bold=True, underline="single")
    row += 1

    datos_acera = [
        ("Ancho acera [m]", data.get('acera1_ancho', '')),
        ("Tipo pavimento", data.get('acera1_pavimento', '')),
        ("Clase de iluminación", data.get('acera1_clase', '')),
    ]

    for label, value in datos_acera:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # SECCIÓN 3: LUMINARIAS
    ws[f'A{row}'] = "LUMINARIAS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    # Encabezados luminarias
    headers = ["Modelo", "Óptica", "Potencia [W]", "Lúmenes [lm]", "Objetivo"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")
    row += 1

    luminarias = data.get('luminarias', [])
    if luminarias:
        for lum in luminarias:
            ws[f'A{row}'] = lum.get('modelo', '')
            ws[f'B{row}'] = lum.get('optica', '')
            ws[f'C{row}'] = lum.get('potencia', '')
            ws[f'D{row}'] = lum.get('lumenes', '')
            ws[f'E{row}'] = lum.get('objetivo', '')
            row += 1

    row += 1

    # SECCIÓN 4: ENERGÍA
    ws[f'A{row}'] = "ENERGÍA Y SOSTENIBILIDAD"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    datos_energia = [
        ("Horas funcionamiento anual [h]", data.get('horas_funcionamiento', '')),
        ("Tarifa eléctrica [€/kWh]", data.get('tarifa_electrica', '')),
        ("Factor CO₂ [kg/kWh]", data.get('factor_co2', '')),
        ("Zona ambiental", data.get('zona_ambiental', '')),
        ("ULOR máximo [%]", data.get('ulor_maximo', '')),
    ]

    for label, value in datos_energia:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # SECCIÓN 5: ENTREGABLES
    ws[f'A{row}'] = "ENTREGABLES SOLICITADOS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    row += 1

    entregables = [
        ("Generar PDF", data.get('generar_pdf', False)),
        ("Generar Excel", data.get('generar_excel', False)),
        ("Generar Isolíneas", data.get('generar_isolineas', False)),
    ]

    for label, value in entregables:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = "✓ Sí" if value else "✗ No"
        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40

    return wb

def crear_pdf(data, graficos_bytes=None):
    """Crea un PDF con los datos del estudio"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=1  # Center
    )
    story.append(Paragraph("CÁLCULO FOTOMÉTRICO SALVI", title_style))
    story.append(Spacer(1, 12))

    # Datos del proyecto
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=10,
        spaceBefore=10
    )

    story.append(Paragraph("DATOS DEL PROYECTO", heading_style))

    datos_proyecto = [
        ["Nombre del proyecto", str(data.get('nombre_proyecto', ''))],
        ["Cliente final", str(data.get('cliente_final', ''))],
        ["Localización", str(data.get('localizacion', ''))],
        ["Proyectista", str(data.get('proyectista', ''))],
        ["Fecha del estudio", str(data.get('fecha_estudio', ''))],
        ["Norma aplicable", str(data.get('norma_aplicable', ''))],
    ]

    t = Table(datos_proyecto, colWidths=[2.5*inch, 4*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    # Geometría
    story.append(Paragraph("GEOMETRÍA", heading_style))

    datos_geom = [
        ["Disposición", str(data.get('disposicion', ''))],
        ["Altura montaje", f"{data.get('altura_montaje', '')} m"],
        ["Interdistancia", f"{data.get('interdistancia', '')} m"],
        ["Ancho calzada", f"{data.get('calzada1_ancho', '')} m"],
        ["Clase de iluminación", str(data.get('calzada1_clase', ''))],
    ]

    t = Table(datos_geom, colWidths=[2.5*inch, 4*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    # Luminarias
    story.append(Paragraph("LUMINARIAS", heading_style))

    luminarias = data.get('luminarias', [])
    if luminarias:
        lum_data = [["Modelo", "Potencia [W]", "Lúmenes [lm]", "Objetivo"]]
        for lum in luminarias:
            lum_data.append([
                str(lum.get('modelo', '')),
                str(lum.get('potencia', '')),
                str(lum.get('lumenes', '')),
                str(lum.get('objetivo', ''))
            ])

        t = Table(lum_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

    # Energía
    story.append(Paragraph("ENERGÍA Y SOSTENIBILIDAD", heading_style))

    datos_energia = [
        ["Horas funcionamiento anual", f"{data.get('horas_funcionamiento', '')} h"],
        ["Tarifa eléctrica", f"{data.get('tarifa_electrica', '')} €/kWh"],
        ["Factor CO₂", f"{data.get('factor_co2', '')} kg/kWh"],
        ["Zona ambiental", str(data.get('zona_ambiental', ''))],
        ["ULOR máximo", f"{data.get('ulor_maximo', '')} %"],
    ]

    t = Table(datos_energia, colWidths=[2.5*inch, 4*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(t)

    # Agregar gráficos si existen
    if graficos_bytes and graficos_bytes.get('iluminancia'):
        story.append(PageBreak())
        story.append(Paragraph("GRÁFICOS DE ISOCURVAS", heading_style))
        story.append(Spacer(1, 12))

        try:
            img_iluminancia = Image(graficos_bytes['iluminancia'], width=6*inch, height=4*inch)
            story.append(img_iluminancia)
            story.append(Spacer(1, 12))
        except Exception as e:
            print(f"Error al agregar gráfico de iluminancia al PDF: {str(e)}")

        try:
            img_luminancia = Image(graficos_bytes['luminancia'], width=6*inch, height=4*inch)
            story.append(img_luminancia)
        except Exception as e:
            print(f"Error al agregar gráfico de luminancia al PDF: {str(e)}")

    doc.build(story)
    buffer.seek(0)
    return buffer

@app.route('/api/calcular', methods=['POST'])
def calcular():
    """Endpoint para calcular fotometría según EN 13201"""
    try:
        data = request.get_json()
        print(f"\nDEBUG: Calculando fotometría...")
        print(f"DEBUG: nombre_proyecto = {data.get('nombre_proyecto')}")

        # Validación mínima
        nombre = data.get('nombre_proyecto', '').strip()
        cliente = data.get('cliente_final', '').strip()

        if not nombre:
            return jsonify({
                'success': False,
                'error': 'Nombre del proyecto es obligatorio'
            }), 200

        if not cliente:
            return jsonify({
                'success': False,
                'error': 'Cliente final es obligatorio'
            }), 200

        # Validar luminarias
        luminarias = data.get('luminarias', [])
        if not luminarias:
            return jsonify({
                'success': False,
                'error': 'Debe definir al menos una luminaria'
            }), 200

        # Realizar cálculos fotométricos
        resultado_calc = calcular_fotometria(data)

        if resultado_calc['success']:
            print(f"✓ Cálculos realizados exitosamente")
            print(f"DEBUG: Resultados: {resultado_calc['resultados']}")

            # Generar gráficos de isocurvas
            resultado_graficos = generar_graficos_isocurvas(data, resultado_calc['resultados'])

            response = {
                'success': True,
                'resultados': resultado_calc['resultados']
            }

            # Agregar gráficos si se generaron exitosamente
            if resultado_graficos['success']:
                print(f"✓ Gráficos de isocurvas generados")
                response['graficos'] = {
                    'iluminancia_b64': resultado_graficos['iluminancia_b64'],
                    'luminancia_b64': resultado_graficos['luminancia_b64']
                }
            else:
                print(f"⚠ Advertencia al generar gráficos: {resultado_graficos['error']}")

            return jsonify(response), 200
        else:
            print(f"ERROR en cálculos: {resultado_calc['error']}")
            return jsonify({
                'success': False,
                'error': resultado_calc['error']
            }), 200

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        print(f"\n❌ ERROR en calcular: {error_msg}")
        print(f"❌ Traceback:\n{traceback_str}")
        return jsonify({
            'success': False,
            'error': f'Error del servidor: {error_msg}'
        }), 200

@app.route('/api/submit-form', methods=['POST'])
def submit_form():
    """Test de envío de formulario"""
    try:
        data = request.get_json()
        print(f"\nDEBUG: Datos recibidos: {list(data.keys())}")
        print(f"DEBUG: nombre_proyecto = {data.get('nombre_proyecto')}")
        print(f"DEBUG: cliente_final = {data.get('cliente_final')}")

        # Validación SUPER flexible
        nombre = data.get('nombre_proyecto', '').strip()
        cliente = data.get('cliente_final', '').strip()

        if not nombre:
            print("ERROR: nombre_proyecto vacío")
            return jsonify({
                'success': False,
                'error': 'Nombre del proyecto es obligatorio'
            }), 200

        if not cliente:
            print("ERROR: cliente_final vacío")
            return jsonify({
                'success': False,
                'error': 'Cliente final es obligatorio'
            }), 200

        print(f"✓ Validación pasada. Nombre: {nombre}, Cliente: {cliente}")

        # Generar Excel completo
        filename_xlsx = f"estudio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath_xlsx = os.path.join(DOWNLOAD_FOLDER, filename_xlsx)

        print(f"DEBUG: Creando Excel en: {filepath_xlsx}")
        wb = crear_excel_completo(data)
        wb.save(filepath_xlsx)

        if not os.path.exists(filepath_xlsx):
            print(f"ERROR CRÍTICO: Excel no fue creado")
            return jsonify({
                'success': False,
                'error': f'Error: Excel no fue creado'
            }), 200

        file_size = os.path.getsize(filepath_xlsx)
        print(f"✓ Excel creado exitosamente!")
        print(f"✓ Nombre: {filename_xlsx}")
        print(f"✓ Tamaño: {file_size} bytes")

        # Generar PDF si está solicitado
        generar_pdf = data.get('generar_pdf', False)
        filename_pdf = None
        filepath_pdf = None

        if generar_pdf:
            filename_pdf = f"estudio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath_pdf = os.path.join(DOWNLOAD_FOLDER, filename_pdf)
            print(f"DEBUG: Creando PDF en: {filepath_pdf}")

            try:
                # Generar gráficos para incluir en PDF
                resultado_graficos = generar_graficos_isocurvas(data, resultados_calculo)
                graficos_bytes = None
                if resultado_graficos['success']:
                    graficos_bytes = {
                        'iluminancia': resultado_graficos['iluminancia_bytes'],
                        'luminancia': resultado_graficos['luminancia_bytes']
                    }

                pdf_buffer = crear_pdf(data, graficos_bytes)
                with open(filepath_pdf, 'wb') as f:
                    f.write(pdf_buffer.getvalue())
                print(f"✓ PDF creado exitosamente!")
                print(f"✓ Nombre: {filename_pdf}")
            except Exception as e:
                print(f"❌ ERROR al crear PDF: {str(e)}")
                filename_pdf = None

        # Construir respuesta
        response = {
            'success': True,
            'message': f'✓ Excel generado correctamente ({file_size} bytes)',
            'download_url': f'/api/download/{filename_xlsx}',
            'filename': filename_xlsx
        }

        if filename_pdf:
            response['pdf_url'] = f'/api/download/{filename_pdf}'
            response['pdf_filename'] = filename_pdf

        print(f"DEBUG: Enviando respuesta: {response}\n")
        return jsonify(response), 200

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        print(f"\n❌ ERROR en submit_form: {error_msg}")
        print(f"❌ Traceback:\n{traceback_str}")
        return jsonify({
            'success': False,
            'error': f'Error del servidor: {error_msg}',
            'traceback': traceback_str
        }), 200

@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """Descarga archivo generado"""
    try:
        filepath = os.path.join(DOWNLOAD_FOLDER, filename)

        print(f"DEBUG: Descargando: {filepath}")

        if not os.path.exists(filepath):
            print(f"ERROR: Archivo no existe: {filepath}")
            print(f"DEBUG: Archivos en downloads: {os.listdir(DOWNLOAD_FOLDER)}")
            return jsonify({
                'success': False,
                'error': f'Archivo no encontrado: {filepath}'
            }), 404

        print(f"✓ Descargando archivo: {filename}")

        # Determinar MIME type por extensión
        if filename.endswith('.pdf'):
            mimetype = 'application/pdf'
        else:
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype
        )
    except Exception as e:
        print(f"ERROR en download: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/debug/files', methods=['GET'])
def debug_files():
    """Debug endpoint"""
    return jsonify({
        'status': 'ok',
        'base_dir': os.path.dirname(__file__),
        'download_folder': DOWNLOAD_FOLDER,
        'download_folder_exists': os.path.isdir(DOWNLOAD_FOLDER),
        'files_in_downloads': os.listdir(DOWNLOAD_FOLDER) if os.path.isdir(DOWNLOAD_FOLDER) else [],
        'assets_folder': ASSETS_FOLDER,
        'assets_folder_exists': os.path.isdir(ASSETS_FOLDER),
        'files_in_assets': os.listdir(ASSETS_FOLDER) if os.path.isdir(ASSETS_FOLDER) else [],
        'timestamp': datetime.now().isoformat()
    }), 200

@app.errorhandler(404)
def not_found(error):
    print(f"ERROR 404: {error}")
    return jsonify({'error': f'Endpoint no encontrado: {request.path}'}), 404

@app.errorhandler(500)
def server_error(error):
    print(f"ERROR 500: {error}")
    return jsonify({'error': 'Error interno del servidor'}), 500

if __name__ == '__main__':
    print("\n" + "="*70)
    print("SalviLux - TEST APP (Versión Completa con LDTs)")
    print("="*70)
    print(f"\n✓ Download folder: {DOWNLOAD_FOLDER}")
    print(f"✓ Assets folder: {ASSETS_FOLDER}")
    print(f"\n✓ Archivos en assets: {os.listdir(ASSETS_FOLDER) if os.path.exists(ASSETS_FOLDER) else 'CARPETA VACÍA'}")

    # Cargar LDTs al iniciar
    print("\n📚 Cargando biblioteca de luminarias...")
    cargar_ldts_cache()

    print("\n🌐 ABRE EN TU NAVEGADOR:")
    print("   → http://localhost:5000")
    print("\n📊 PRUEBA LA API:")
    print("   → http://localhost:5000/api/test")
    print("   → http://localhost:5000/api/luminarias")
    print("   → http://localhost:5000/api/debug/files")
    print("\n⏹️  PARA DETENER: Presiona Ctrl+C")
    print("\n" + "="*70 + "\n")

    try:
        app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
