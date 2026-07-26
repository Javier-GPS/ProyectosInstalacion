import re
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..database import DATA_DIR, get_db
from ..schemas.models import (
    AdvancedOptimizationRequest,
    BatchCalculationResponse,
    CalculationConfig,
    CalculationResult,
    OptimizationResponse,
)
from ..services.calculator import calculate_config, run_calculation
from ..services.catalog_service import get_eficiencia
from ..services.ifc_exporter import build_ifc
from ..services.ldt_matcher import require_ldt_for_config
from ..services.luminaire_catalog import clamp_power_to_pmax
from ..services.optimizer import (
    run_simple_optimization,
    run_flux_optimization,
    run_advanced_optimization,
    run_advanced_optimization_batch,
)
from ..services.led_calculator import LedModelError

router = APIRouter()


@router.post("/calculate", response_model=CalculationResult)
async def calculate(config: CalculationConfig, db: Session = Depends(get_db), skip_optimization: bool = False):
    """Run a full CIE 140 / EN 13201 calculation.

    By default auto-optimises flux to reach the target Lavg. When
    ``skip_optimization=true`` the supplied ``target_flux`` (or
    ``power``) is used verbatim — useful for testing specific flux
    values against EN 13201.
    """
    try:
        if skip_optimization:
            result = calculate_config(db, config)
        else:
            optimization = run_flux_optimization(db, config)
            if optimization.config is not None:
                config = optimization.config
            result = calculate_config(db, config)
        return result
    except (LedModelError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/export/ifc")
async def export_ifc(config: CalculationConfig, db: Session = Depends(get_db)):
    """Re-run the calculation and return an IFC 4.3 file (DIALux evo 7+ compatible).

    The generated IFC contains the road geometry, a carriageway slab, and
    one ``IfcLightFixture`` per pole baseline. The maintenance factor used
    for the calculation is written to ``Pset_LightFixtureTypeCommon`` /
    ``Pset_LightFixtureCommon`` so DIALux evo re-applies the same MF when
    it imports the file. The response is a one-shot download.
    """
    config = clamp_power_to_pmax(db, config)
    lente_eff, difusor_eff = get_eficiencia(db, config.lente, config.difusor)
    ldt_id, _ = require_ldt_for_config(config)
    result = run_calculation(config, ldt_id, lente_eficiencia=lente_eff, difusor_eficiencia=difusor_eff)

    out_dir = DATA_DIR / "ifc_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / f"luxstudio_{uuid4().hex}.ifc"
    build_ifc(config, result, result.luminaire, output_path)

    return FileResponse(
        str(output_path),
        media_type="application/x-step",
        filename="luxstudio_export.ifc",
        background=None,
    )


@router.post("/optimize/simple", response_model=OptimizationResponse)
async def optimize_simple(config: CalculationConfig, db: Session = Depends(get_db)):
    """Find the lowest compliant power while keeping the selected geometry and luminaire fixed."""
    return run_simple_optimization(db, config)


@router.post("/optimize/flux", response_model=OptimizationResponse)
async def optimize_flux(config: CalculationConfig, db: Session = Depends(get_db)):
    """Find the smallest ``target_flux`` that nails the class Lavg limit from below.

    Bumps the user-supplied luminous flux up or down (and lets the V2
    LED model re-pick a PCB / current) so the resulting Lavg matches
    the EN 13201 class requirement as closely as possible while
    staying above it.  Falls back to the power-based optimizer when
    the 4-tuple is incomplete or the class has no Lavg requirement.
    """
    return run_flux_optimization(db, config)


@router.post("/optimize/advanced", response_model=OptimizationResponse)
async def optimize_advanced(request: AdvancedOptimizationRequest, db: Session = Depends(get_db)):
    """Optimize selected installation variables against installed W/m."""
    return run_advanced_optimization(db, request)


@router.post("/optimize/advanced-batch", response_model=BatchCalculationResponse)
async def optimize_advanced_batch(request: AdvancedOptimizationRequest, db: Session = Depends(get_db)):
    """Optimize each selected optic and return one downloadable row per lens."""
    return run_advanced_optimization_batch(db, request)


def _float(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, str):
        value = value.strip().replace(",", ".")
        value = re.sub(r"[^0-9.+-]", "", value)
        if value in ("", "-", "+", ".", "-.", "+."):
            return default
    return float(value)


def _int(value, default=0):
    return int(round(_float(value, default)))


def _text(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def _cct(value):
    match = re.search(r"\d+", _text(value))
    return int(match.group(0)) if match else 4000


def _arrangement(value):
    raw = _text(value, "Lineal").lower()
    if "alternada" in raw or "staggered" in raw:
        return "Bilateral Alternada"
    if "bilateral" in raw:
        return "Bilateral"
    if "doble" in raw or "twin" in raw:
        return "Central Doble"
    if "isleta" in raw or "single" in raw:
        return "En Isleta"
    return "Lineal"


def _lighting_class(value):
    raw = _text(value, "M3").upper()
    match = re.search(r"\b[MP][1-6]\b", raw)
    return match.group(0) if match else "M3"


def _optic_family(value, height, road_width):
    raw = _text(value).upper()
    match = re.search(r"\bF[0-9A-Z]{2,4}\b", raw)
    if match:
        return match.group(0)
    ratio = height / road_width if road_width else 0
    if ratio <= 1:
        return "F151"
    return "F2MD"


def _config_from_row(ws, row):
    height = _float(ws.cell(row, 3).value, 9)
    road_width = _float(ws.cell(row, 6).value, 7)
    optic_family = _optic_family(ws.cell(row, 12).value, height, road_width)

    return CalculationConfig(
        road_width=road_width,
        sidewalk_left=_float(ws.cell(row, 5).value, 0),
        sidewalk_right=0,
        lanes=2,
        arrangement=_arrangement(ws.cell(row, 2).value),
        height=height,
        spacing=_float(ws.cell(row, 4).value, 30),
        arm_length=_float(ws.cell(row, 7).value, 0),
        tilt=_float(ws.cell(row, 13).value, 0),
        optic_family=optic_family,
        power=_float(ws.cell(row, 14).value, 100),
        lighting_class=_lighting_class(ws.cell(row, 8).value),
        mf=_float(ws.cell(row, 9).value, 0.85),
        pavement="R3",
        cct=_cct(ws.cell(row, 10).value),
    )


def _config_from_values(values):
    height = _float(values[2] if len(values) > 2 else None, 9)
    road_width = _float(values[5] if len(values) > 5 else None, 7)
    optic_family = _optic_family(values[11] if len(values) > 11 else None, height, road_width)

    return CalculationConfig(
        road_width=road_width,
        sidewalk_left=_float(values[4] if len(values) > 4 else None, 0),
        sidewalk_right=0,
        lanes=2,
        arrangement=_arrangement(values[1] if len(values) > 1 else None),
        height=height,
        spacing=_float(values[3] if len(values) > 3 else None, 30),
        arm_length=_float(values[6] if len(values) > 6 else None, 0),
        tilt=_float(values[12] if len(values) > 12 else None, 0),
        optic_family=optic_family,
        power=_float(values[13] if len(values) > 13 else None, 100),
        lighting_class=_lighting_class(values[7] if len(values) > 7 else None),
        mf=_float(values[8] if len(values) > 8 else None, 0.85),
        pavement="R3",
        cct=_cct(values[9] if len(values) > 9 else None),
    )


@router.post("/batch-excel", response_model=BatchCalculationResponse)
async def calculate_batch_excel(file: UploadFile = File(...)):
    """Read a DIALux-style Excel file and calculate every model row."""
    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.active

    rows = []
    for row_index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = tuple(values)
        model_id = _text(values[0] if values else None, f"Row {row_index}")
        if not model_id:
            continue
        rows.append((row_index, model_id, values))

    def calculate_row(row_data):
        row, model_id, values = row_data
        try:
            config = _config_from_values(values)
            ldt_id, _ = require_ldt_for_config(config)
            result = run_calculation(config, ldt_id)
            return BatchCalculationItem(model_id=model_id, row=row, config=config, result=result)
        except Exception as exc:
            return BatchCalculationItem(model_id=model_id, row=row, error=str(exc))

    with ThreadPoolExecutor(max_workers=8) as executor:
        items = list(executor.map(calculate_row, rows))

    return BatchCalculationResponse(filename=file.filename or "uploaded.xlsx", count=len(items), items=items)
