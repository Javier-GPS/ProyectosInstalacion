"""GIS — FastAPI router for SALVI GIS endpoints.

Maintains 1:1 API compatibility with the old ``api_server.py`` so the GIS
frontend (``SALVI GIS.html``) keeps working without changes while the data
lives in PostgreSQL alongside LuxStudio.
"""

import io
import json
import math
import os
import secrets
import time
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import (
    Project, User,
    GisZone, GisZoneConfig, GisZoneOsmData, GisZoneTrees,
    GisLuminaire, GisInventoryLuminaire, GisPhotometricResult,
    GisProjectUiConfig, ensure_gis_tables,
)
from ..routers.auth import current_user
from ..services.auth import create_token, hash_password, verify_password

router = APIRouter()

# ── Pending import cache (expires after 30 min) ────────────────────────────
_PENDING_IMPORTS: dict[str, dict] = {}

# ── Colour palette (same as api_server.py) ──────────────────────────────────
DEFAULT_COLORS = [
    '#4caf82', '#e67e22', '#3498db', '#9b59b6', '#e74c3c',
    '#1abc9c', '#f39c12', '#2980b9', '#8e44ad', '#c0392b',
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def _fval(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None

def _sval(v):
    return str(v).strip() if v is not None else None

def _parse_json_field(v, default=None):
    if v is None:
        return default if default is not None else []
    if isinstance(v, (list, dict)):
        return v
    if isinstance(v, str):
        try:
            return json.loads(v)
        except (json.JSONDecodeError, TypeError):
            return default if default is not None else []
    return v

def _zone_to_dict(z: GisZone, spacing: int = 30) -> dict:
    """Convert a GisZone ORM row to the dict the GIS frontend expects."""
    return {
        "id": z.id,
        "name": z.name,
        "type": z.type or "",
        "color": z.color or DEFAULT_COLORS[0],
        "priority": z.priority if z.priority is not None else 2,
        "center_lat": z.center_lat,
        "center_lon": z.center_lon,
        "zoom": z.zoom if z.zoom is not None else 12,
        "bbox": z.bbox or "",
        "description": z.description or "",
        "corridors": _parse_json_field(z.corridors, []),
        "bounds_polygon": _parse_json_field(z.bounds_polygon, []),
        "est": _parse_json_field(z.est, {}),
        "source": z.source or "manual",
        "project_id": None,  # we'll resolve later — see below
        "created_at": z.created_at.isoformat() if z.created_at else None,
        "spacing": spacing,
    }

def _row2dict(r) -> dict | None:
    """Convert SQLAlchemy model row to dict (column name → value)."""
    if r is None:
        return None
    return {c.name: getattr(r, c.name) for c in r.__table__.columns}

def _rows2list(rows) -> list[dict]:
    return [_row2dict(r) for r in rows]


# ── require_admin dependency (defined early so routes can reference it) ────
def require_admin(user: User = Depends(current_user)) -> User:
    if user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin role required")
    return user


# ═══════════════════════════════════════════════════════════════════════════
# AUTH — GIS-compatible login (username or email)
# ═══════════════════════════════════════════════════════════════════════════

class GisLoginBody(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    password: str


@router.post("/api/auth/login")
async def gis_login(body: GisLoginBody, db: Session = Depends(get_db)):
    """Unified login: accepts ``email`` (LuxStudio) or ``username`` (GIS)."""
    ensure_gis_tables()
    user: User | None = None
    if body.email:
        user = db.query(User).filter(User.email == body.email.lower().strip()).first()
    if not user and body.username:
        user = db.query(User).filter(
            (User.email == body.username.lower().strip()) |
            (User.name == body.username.strip())
        ).first()
    if not user or not verify_password(body.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Credenciales no validas")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Usuario inactivo")

    user.last_login_at = datetime.now(timezone.utc)
    db.commit()

    token = create_token(user)
    # Return superset format: both frontends understand it
    return {
        "token": token,                 # GIS
        "access_token": token,          # LuxStudio
        "token_type": "bearer",
        "user": {
            "id": str(user.id),         # GIS (string)
            "user_id": user.id,         # LuxStudio (int)
            "username": user.name,      # GIS
            "name": user.name,          # LuxStudio
            "email": user.email,
            "role": user.role,
            "company_name": user.company_name or "SALVI LIGHTING",
            "is_active": user.is_active,
            "must_reset_password": user.must_reset_password,
        },
    }


class GisSetupBody(BaseModel):
    username: str
    email: str = ""
    password: str

@router.post("/api/auth/setup")
async def gis_auth_setup(body: GisSetupBody, db: Session = Depends(get_db)):
    """Initial admin setup — GIS compat. Creates first admin if no users exist."""
    from ..services.auth import ensure_users_table
    ensure_users_table()
    existing = db.query(User).count()
    if existing > 0:
        raise HTTPException(status_code=400, detail="Ya hay usuarios en el sistema")
    from ..services.auth import COMPANY_NAME
    user = User(
        name=body.username.strip(),
        email=body.email.lower().strip() or f"{body.username.strip().lower()}@salvi.lighting",
        password_hash=hash_password(body.password),
        role="ADMIN",
        is_active=True,
        must_reset_password=False,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"ok": True, "user": {"id": str(user.id), "username": user.name, "email": user.email, "role": user.role}}


@router.get("/api/auth/me")
async def gis_auth_me(user: User = Depends(current_user)):
    """Unified /me — returns both GIS and LuxStudio formats."""
    return {
        "id": str(user.id),
        "user_id": user.id,
        "username": user.name,
        "name": user.name,
        "email": user.email,
        "role": user.role,
        "company_name": user.company_name or "SALVI LIGHTING",
        "is_active": user.is_active,
        "must_reset_password": user.must_reset_password,
    }


# ── Password reset (simple token-in-memory, same as api_server.py) ──────────
_RESET_TOKENS: dict[str, dict] = {}

class ResetRequest(BaseModel):
    email: str

class ResetApply(BaseModel):
    token: str
    password: str

@router.post("/api/auth/reset-request")
async def gis_reset_request(body: ResetRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == body.email.lower().strip()).first()
    if not user:
        # Don't reveal whether the email exists — same as api_server.py
        return {"ok": True, "message": "Si el email existe, recibirás un enlace"}
    token = secrets.token_urlsafe(32)
    _RESET_TOKENS[token] = {"uid": str(user.id), "email": user.email, "exp": time.time() + 3600}
    # In production, send email. For now, log the token.
    print(f"[GIS] Password-reset token for {user.email}: {token}")
    return {"ok": True, "message": "Si el email existe, recibirás un enlace"}

@router.post("/api/auth/reset-apply")
async def gis_reset_apply(body: ResetApply, db: Session = Depends(get_db)):
    entry = _RESET_TOKENS.pop(body.token, None)
    if not entry or entry["exp"] < time.time():
        raise HTTPException(status_code=400, detail="Token invalido o expirado")
    user = db.get(User, int(entry["uid"]))
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    user.password_hash = hash_password(body.password)
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════
# USERS — GIS-compatible user management
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/api/users")
async def gis_users_list(user: User = Depends(require_admin), db: Session = Depends(get_db)):
    users = db.query(User).all()
    return [{"id": str(u.id), "username": u.name, "email": u.email, "role": u.role} for u in users]

class GisCreateUserBody(BaseModel):
    username: str
    email: str = ""
    password: str
    role: str = "user"

@router.post("/api/users")
async def gis_users_create(body: GisCreateUserBody, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(User).filter(
        (User.email == body.email.lower().strip()) | (User.name == body.username.strip())
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Usuario ya existe")
    new_user = User(
        name=body.username.strip(),
        email=body.email.lower().strip() or f"{body.username.strip().lower()}@salvi.lighting",
        password_hash=hash_password(body.password),
        role="ADMIN" if body.role.lower() == "admin" else "USER",
        is_active=True,
        must_reset_password=False,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"id": str(new_user.id), "username": new_user.name, "email": new_user.email, "role": new_user.role}


class GisUpdateUserBody(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None


@router.put("/api/users/{user_id}")
async def gis_users_update(user_id: int, body: GisUpdateUserBody, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if body.username:
        target.name = body.username.strip()
    if body.email:
        target.email = body.email.lower().strip()
    if body.password:
        target.password_hash = hash_password(body.password)
    if body.role:
        target.role = "ADMIN" if body.role.lower() == "admin" else "USER"
    db.commit()
    return {"ok": True}


@router.delete("/api/users/{user_id}")
async def gis_users_delete(user_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if target.id == user.id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    db.delete(target)
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════
# NOMINATIM proxy (forward to OpenStreetMap)
# ═══════════════════════════════════════════════════════════════════════════

import urllib.request as _ur
import urllib.parse as _up
import urllib.error as _ue
import ssl

_NOM_HEADERS = {"User-Agent": "SalviGIS/1.0 (contact@salvi.es)", "Accept-Language": "es,en"}

def _nom_ssl_ctx():
    try:
        return ssl.create_default_context()
    except Exception:
        return ssl._create_unverified_context()

@router.get("/api/nominatim/search")
async def gis_nominatim_search(q: str = Query(...), featuretype: Optional[str] = None):
    try:
        params = f"q={_up.quote(q)}&format=json&addressdetails=1&polygon_geojson=1&limit=10"
        if featuretype:
            params += f"&featuretype={featuretype}"
        req = _ur.Request(f"https://nominatim.openstreetmap.org/search?{params}", headers=_NOM_HEADERS)
        with _ur.urlopen(req, timeout=12, context=_nom_ssl_ctx()) as resp:
            data = json.loads(resp.read())
        return data
    except _ue.URLError as e:
        raise HTTPException(status_code=502, detail=f"Nominatim unreachable: {e.reason}")

@router.get("/api/nominatim/reverse")
async def gis_nominatim_reverse(lat: float = Query(...), lon: float = Query(...), zoom: int = 14):
    try:
        params = f"lat={lat}&lon={lon}&format=json&polygon_geojson=1&zoom={zoom}"
        req = _ur.Request(f"https://nominatim.openstreetmap.org/reverse?{params}", headers=_NOM_HEADERS)
        with _ur.urlopen(req, timeout=12, context=_nom_ssl_ctx()) as resp:
            data = json.loads(resp.read())
        return data
    except _ue.URLError as e:
        raise HTTPException(status_code=502, detail=f"Nominatim unreachable: {e.reason}")


# ═══════════════════════════════════════════════════════════════════════════
# ZONES  (core GIS entity)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/api/zones")
async def gis_zones_list(
    project_id: Optional[int] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    ensure_gis_tables()
    query = db.query(GisZone, GisZoneConfig.spacing).outerjoin(
        GisZoneConfig, GisZone.id == GisZoneConfig.zone_id
    )
    if project_id is not None:
        query = query.filter(GisZone.project_id == project_id)
    query = query.order_by(GisZone.created_at.desc())
    results = []
    for z, sp in query.all():
        d = _zone_to_dict(z, spacing=sp or 30)
        results.append(d)
    return results


class GisCreateZoneBody(BaseModel):
    id: Optional[str] = None
    name: str = "Zona"
    type: str = ""
    color: Optional[str] = None
    priority: int = 2
    center_lat: Optional[float] = None
    center_lon: Optional[float] = None
    zoom: int = 12
    bbox: str = ""
    description: str = ""
    est: dict = {}
    corridors: list = []
    bounds_polygon: list = []
    source: str = "manual"
    project_id: Optional[int] = None
    osm_relation: Optional[int] = None
    center: list = []  # alternative to center_lat/center_lon

@router.post("/api/zones", status_code=201)
async def gis_zones_create(body: GisCreateZoneBody, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    zid = body.id or uuid.uuid4().hex[:12]
    clat = body.center_lat or (body.center[0] if len(body.center) > 0 else None)
    clon = body.center_lon or (body.center[1] if len(body.center) > 1 else None)

    zone = GisZone(
        id=zid,
        name=body.name.strip(),
        type=body.type,
        color=body.color or DEFAULT_COLORS[0],
        priority=body.priority,
        center_lat=_fval(clat),
        center_lon=_fval(clon),
        zoom=body.zoom,
        bbox=body.bbox,
        description=body.description,
        est=body.est,
        corridors=body.corridors,
        bounds_polygon=body.bounds_polygon,
        source=body.source,
        project_id=body.project_id,
        osm_relation=body.osm_relation,
    )
    db.add(zone)
    # Create default zone_config
    db.add(GisZoneConfig(zone_id=zid))
    db.commit()
    db.refresh(zone)
    sp = db.get(GisZoneConfig, zid)
    return _zone_to_dict(zone, spacing=sp.spacing if sp else 30)


@router.put("/api/zones/{zone_id}")
async def gis_zones_update(zone_id: str, body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    zone = db.get(GisZone, zone_id)
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    # Map allowed fields
    field_map = {
        "name": "name", "type": "type", "color": "color", "priority": "priority",
        "center_lat": "center_lat", "center_lon": "center_lon", "zoom": "zoom",
        "bbox": "bbox", "description": "description", "source": "source",
        "project_id": "project_id",
    }
    for gis_key, orm_key in field_map.items():
        if gis_key in body:
            setattr(zone, orm_key, body[gis_key])
    # JSON fields
    if "corridors" in body and isinstance(body["corridors"], (list, dict)):
        zone.corridors = body["corridors"]
    if "bounds_polygon" in body and isinstance(body["bounds_polygon"], (list, dict)):
        zone.bounds_polygon = body["bounds_polygon"]
    if "est" in body and isinstance(body["est"], dict):
        zone.est = body["est"]
    db.commit()
    return {"ok": True}


@router.delete("/api/zones/{zone_id}")
async def gis_zones_delete(zone_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    zone = db.get(GisZone, zone_id)
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    db.delete(zone)  # CASCADE deletes related config, osm_data, luminaires, etc.
    db.commit()
    return {"ok": True}


# ── Zone OSM ────────────────────────────────────────────────────────────────

@router.get("/api/zones/osm/all")
async def gis_zones_osm_all(
    project_id: Optional[int] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    ensure_gis_tables()
    query = db.query(GisZoneOsmData).join(GisZone)
    if project_id is not None:
        query = query.filter(GisZone.project_id == project_id)
    result = {}
    for row in query.all():
        result[row.zone_id] = {
            "zone_id": row.zone_id,
            "km_by_type": row.km_by_type or {},
            "ways": row.ways or [],
            "source": row.source or "",
            "loaded_at": row.loaded_at.isoformat() if row.loaded_at else None,
        }
    return result


@router.get("/api/zones/{zone_id}/osm")
async def gis_zone_osm(zone_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    row = db.get(GisZoneOsmData, zone_id)
    if not row:
        return {}
    return {
        "zone_id": row.zone_id,
        "km_by_type": row.km_by_type or {},
        "ways": row.ways or [],
        "source": row.source or "",
        "loaded_at": row.loaded_at.isoformat() if row.loaded_at else None,
    }


@router.put("/api/zones/{zone_id}/osm")
async def gis_zone_osm_save(zone_id: str, body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    existing = db.get(GisZoneOsmData, zone_id)
    data = {
        "km_by_type": body.get("kmByType", body.get("km_by_type", {})),
        "ways": body.get("ways", []),
        "source": body.get("source", "osm"),
    }
    if existing:
        existing.km_by_type = data["km_by_type"]
        existing.ways = data["ways"]
        existing.source = data["source"]
        existing.loaded_at = datetime.now(timezone.utc)
    else:
        db.add(GisZoneOsmData(zone_id=zone_id, **data))
    db.commit()
    return {"ok": True}

# ── Zone Config ─────────────────────────────────────────────────────────────

@router.put("/api/zones/{zone_id}/config")
async def gis_zone_config(zone_id: str, body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    cfg = db.get(GisZoneConfig, zone_id)
    if not cfg:
        cfg = GisZoneConfig(zone_id=zone_id)
        db.add(cfg)
    for key in ("spacing", "watt_hps", "watt_led", "efficacy", "hours_night"):
        if key in body:
            setattr(cfg, key, body[key])
    cfg.updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"ok": True}

# ── Zone Trees ──────────────────────────────────────────────────────────────

@router.get("/api/zones/{zone_id}/trees")
async def gis_zone_trees_get(zone_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    row = db.get(GisZoneTrees, zone_id)
    return {"trees": row.trees if row else [], "loaded_at": row.loaded_at.isoformat() if row and row.loaded_at else None}

@router.put("/api/zones/{zone_id}/trees")
async def gis_zone_trees_put(zone_id: str, body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    existing = db.get(GisZoneTrees, zone_id)
    trees = body.get("trees", [])
    if existing:
        existing.trees = trees
        existing.loaded_at = datetime.now(timezone.utc)
    else:
        db.add(GisZoneTrees(zone_id=zone_id, trees=trees))
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════
# LUMINAIRES  (designed / placed)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/api/luminaires")
async def gis_luminaires_list(
    zone_id: Optional[str] = None,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    ensure_gis_tables()
    query = db.query(GisLuminaire)
    if zone_id:
        query = query.filter(GisLuminaire.zone_id == zone_id)
    return _rows2list(query.order_by(GisLuminaire.id).all())


class GisBulkLuminaire(BaseModel):
    project_id: Optional[int] = None
    zone_id: str
    road_type: Optional[str] = None
    lighting_class: Optional[str] = None
    street_name: Optional[str] = None
    lat: float
    lon: float
    watts: Optional[float] = None
    spacing: Optional[float] = None
    tilt: Optional[float] = None
    height_m: Optional[float] = None
    arm_len: Optional[float] = None
    distribution: Optional[str] = None

@router.post("/api/luminaires/bulk")
async def gis_luminaires_bulk(body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Bulk upsert luminaires for a zone. Accepts ``{"luminaires": [...], "zone_id": "..."}``."""
    ensure_gis_tables()
    zone_id = body.get("zone_id") or (body.get("luminaires", [{}])[0].get("zone_id") if body.get("luminaires") else None)
    if not zone_id:
        raise HTTPException(status_code=400, detail="zone_id required")

    luminaires = body.get("luminaires", [])
    if not luminaires:
        return {"ok": True, "count": 0}

    # Delete existing for this zone, then bulk insert
    db.query(GisLuminaire).filter(GisLuminaire.zone_id == zone_id).delete()

    for item in luminaires:
        db.add(GisLuminaire(
            project_id=item.get("project_id"),
            zone_id=zone_id,
            road_type=item.get("road_type"),
            lighting_class=item.get("lighting_class"),
            street_name=item.get("street_name"),
            lat=item["lat"],
            lon=item["lon"],
            watts=_fval(item.get("watts")),
            spacing=_fval(item.get("spacing")),
            tilt=_fval(item.get("tilt")),
            height_m=_fval(item.get("height_m") or item.get("height")),
            arm_len=_fval(item.get("arm_len") or item.get("arm_length")),
            distribution=item.get("distribution"),
        ))
    db.commit()
    return {"ok": True, "count": len(luminaires)}


# ── Luminaires export ─────────────────────────────────────────────────────

@router.get("/api/luminaires/export")
async def gis_luminaires_export(
    zone_id: str = Query(...),
    format: str = Query("xlsx"),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Export luminaires as XLSX. Same as old api_server.py ``h_luminaires_export``."""
    import openpyxl
    ensure_gis_tables()
    luminaires = db.query(GisLuminaire).filter(GisLuminaire.zone_id == zone_id).order_by(GisLuminaire.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Luminarias"
    ws.append(["ID", "Zona", "Tipo via", "Clase ilum.", "Calle", "Lat", "Lon", "W", "Espaciado", "Colocada"])
    for r in luminaires:
        ws.append([r.id, r.zone_id, r.road_type, r.lighting_class, r.street_name,
                   r.lat, r.lon, r.watts, r.spacing, r.placed_at.isoformat() if r.placed_at else ""])

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="luminaires_{zone_id}.xlsx"'},
    )


# ── Inventory ──────────────────────────────────────────────────────────────

@router.get("/api/zones/{zone_id}/inventory")
async def gis_zone_inventory(zone_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    rows = db.query(GisInventoryLuminaire).filter(
        GisInventoryLuminaire.zone_id == zone_id
    ).order_by(GisInventoryLuminaire.id).all()
    return _rows2list(rows)


@router.post("/api/parse/inventory_excel")
async def gis_parse_inventory(request: Request):
    """Parse uploaded inventory Excel. Returns column detection + caches rows."""
    import openpyxl
    raw = await request.body()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise HTTPException(status_code=400, detail="Excel vacio")

    headers = [str(h).lower().strip() if h else "" for h in rows_iter[0]]

    FIELD_KEYS = {
        'lat': [' lat', 'latitud', 'latitude', 'coord_y', 'y_coord', 'coordenada y'],
        'lon': [' lon', ' lng', 'longitud', 'longitude', 'coord_x', 'x_coord', 'coordenada x'],
        'power_w': [' w ', 'potencia', 'power', 'watt', 'kw'],
        'height_m': ['altura', 'height', 'alto', 'h_m', 'h (m)'],
        'brand': ['marca', 'brand', 'fabricante'],
        'model': ['modelo', 'model', 'tipo'],
        'lamp_type': ['lampara', 'lamp', 'tecnologia', 'tech'],
        'point_id': ['id punto', 'punto_id', 'point_id', 'cod', 'codigo', 'num'],
        'circuit_id': ['cuadro', 'panel', 'circuit', 'armario'],
        'line_id': ['linea', 'ramal', 'line', 'circuito'],
        'support_type': ['soporte', 'support', 'bac', 'columna', 'poste'],
    }
    col_map = {}
    for i, h in enumerate(headers):
        hl = f" {h.lower().strip()}"
        for field, aliases in FIELD_KEYS.items():
            if field not in col_map and any(k in hl for k in aliases):
                col_map[field] = i
                break

    # Cache parsed rows for subsequent import call
    temp_id = uuid.uuid4().hex[:12]
    _PENDING_IMPORTS[temp_id] = {
        "rows": [list(r) for r in rows_iter[1:]],
        "headers": headers,
        "ts": time.time(),
    }

    # Sample data
    sample_rows = []
    for row in rows_iter[1:11]:
        if any(row):
            sample_rows.append([str(v) if v is not None else "" for v in row])

    return {
        "temp_id": temp_id,
        "filename": "upload.xlsx",
        "total_rows": len(rows_iter) - 1,
        "detected_columns": col_map,
        "headers": headers,
        "sample_rows": sample_rows,
        "field_order": list(FIELD_KEYS.keys()),
    }


class GisImportInventoryBody(BaseModel):
    temp_id: str
    mapping: dict
    zone_name: str = "Nueva zona"
    project_id: Optional[int] = None
    color: str = "#4caf82"


@router.post("/api/import/inventory")
async def gis_import_inventory(body: GisImportInventoryBody, db: Session = Depends(get_db)):
    """Import parsed inventory rows, create zone."""
    ensure_gis_tables()

    pending = _PENDING_IMPORTS.pop(body.temp_id, None)
    if not pending:
        raise HTTPException(status_code=400, detail="temp_id no valido o expirado")
    if time.time() - pending["ts"] > 1800:
        raise HTTPException(status_code=400, detail="Datos expirados, sube el archivo otra vez")

    rows_raw = pending["rows"]
    mapping = body.mapping

    def _val(row, field):
        idx = mapping.get(field)
        if idx is not None and idx < len(row):
            v = row[idx]
            return str(v).strip() if v is not None else None
        return None

    def _num(row, field):
        v = _val(row, field)
        if v is None or v == "":
            return None
        try:
            return float(v.replace(",", ".").replace(" ", ""))
        except ValueError:
            return None

    # Create zone from imported data
    zid = uuid.uuid4().hex[:12]
    zone = GisZone(id=zid, name=body.zone_name.strip(), color=body.color or "#4caf82",
                   source="inventory", project_id=body.project_id)
    db.add(zone)
    db.add(GisZoneConfig(zone_id=zid))
    db.flush()

    count = 0
    for row in rows_raw:
        if not any(row):
            continue
        lat = _num(row, "lat")
        lon = _num(row, "lon")
        if lat is None or lon is None:
            continue
        db.add(GisInventoryLuminaire(
            zone_id=zid, lat=lat, lon=lon,
            point_id=_val(row, "point_id"),
            power_w=_num(row, "power_w"),
            height_m=_num(row, "height_m"),
            brand=_val(row, "brand"), model=_val(row, "model"),
            lamp_type=_val(row, "lamp_type"),
            support_type=_val(row, "support_type"),
            circuit_id=_val(row, "circuit_id"),
            line_id=_val(row, "line_id"),
        ))
        count += 1

    db.commit()
    db.refresh(zone)
    return {"zone": _zone_to_dict(zone, spacing=30), "count": count}


# (luminaire list is at GET /api/luminaires above)


@router.delete("/api/luminaires/{zone_id}/{lum_id}")
async def gis_luminaires_delete(zone_id: str, lum_id: int, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    lum = db.get(GisLuminaire, lum_id)
    if not lum or lum.zone_id != zone_id:
        raise HTTPException(status_code=404, detail="Luminaire not found")
    db.delete(lum)
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════
# PHOTOMETRIC results
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/api/zones/{zone_id}/photometric")
async def gis_zone_photometric(zone_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    rows = db.query(GisPhotometricResult).filter(
        GisPhotometricResult.zone_id == zone_id
    ).order_by(GisPhotometricResult.id).all()
    return [{
        "id": r.id,
        "zone_id": r.zone_id,
        "segment_name": r.segment_name,
        "match_key": r.match_key,
        "road_width": r.road_width,
        "spacing": r.spacing,
        "lighting_class": r.lighting_class,
        "power_w": r.power_w,
        "lm_em": r.lm_em,
        "uo": r.uo,
        "ui": r.ui,
        "ti": r.ti,
        "sr": r.sr,
        "model": r.model,
        "lente": r.lente,
        "tilt": r.tilt,
        "phi_lm": r.phi_lm,
        "cumple": r.cumple,
        "notes": r.notes,
        "imported_at": r.imported_at.isoformat() if r.imported_at else None,
    } for r in rows]


@router.post("/api/import/photometric")
async def gis_import_photometric(request: Request, db: Session = Depends(get_db)):
    """Import photometric results XLSX from LuxStudio. Expects raw XLSX bytes."""
    import openpyxl
    ensure_gis_tables()

    raw = await request.body()
    # Try to get zone_id from query
    import urllib.parse
    # We get it from the URL manually
    zone_id = request.query_params.get("zone_id", "")

    if not zone_id:
        raise HTTPException(status_code=400, detail="zone_id query parameter required")

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return {"imported": 0, "rows": []}

    headers = [str(h).lower().strip() if h else "" for h in rows_raw[0]]
    # Column detection
    def _col(names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None
    ci = {
        "segment_name":  _col(["segmento", "segment", "name", "tramo"]),
        "match_key":     _col(["match_key", "key", "clave"]),
        "road_width":    _col(["ancho", "road_width", "w_m", "width"]),
        "spacing":       _col(["espaciado", "spacing", "sp"]),
        "lighting_class":_col(["clase", "class", "lc", "lighting_class"]),
        "power_w":       _col(["w", "potencia", "power", "watts"]),
        "lm_em":         _col(["lm/m2", "lm_em", "em", "e_m"]),
        "uo":            _col(["uo", "u0"]),
        "ui":            _col(["ui"]),
        "ti":            _col(["ti"]),
        "sr":            _col(["sr"]),
        "model":         _col(["modelo", "model"]),
        "lente":         _col(["lente", "lens"]),
        "tilt":          _col(["tilt", "inclinacion"]),
        "phi_lm":        _col(["phi", "flujo", "phi_lm", "lm"]),
        "cumple":        _col(["cumple", "ok", "pass", "resultado"]),
        "notes":         _col(["notas", "notes", "observaciones"]),
    }

    imported = 0
    rows_out = []
    for row in rows_raw[1:]:
        if not any(row):
            continue
        def g(k):
            idx = ci.get(k)
            return row[idx] if idx is not None and idx < len(row) else None
        road_w  = _fval(g("road_width"))
        sp_val  = _fval(g("spacing"))
        lc      = _sval(g("lighting_class"))
        mk      = _sval(g("match_key")) or (f"{road_w}|{sp_val}|{lc}" if road_w and sp_val and lc else None)
        if not mk:
            continue

        existing = db.query(GisPhotometricResult).filter(
            GisPhotometricResult.zone_id == zone_id,
            GisPhotometricResult.match_key == mk
        ).first()
        data = {
            "segment_name": _sval(g("segment_name")),
            "match_key": mk,
            "road_width": road_w,
            "spacing": sp_val,
            "lighting_class": lc,
            "power_w": _fval(g("power_w")),
            "lm_em": _fval(g("lm_em")),
            "uo": _fval(g("uo")),
            "ui": _fval(g("ui")),
            "ti": _fval(g("ti")),
            "sr": _fval(g("sr")),
            "model": _sval(g("model")),
            "lente": _sval(g("lente")),
            "tilt": _fval(g("tilt")),
            "phi_lm": _fval(g("phi_lm")),
            "cumple": _sval(g("cumple")),
            "notes": _sval(g("notes")),
        }
        if existing:
            for k, v in data.items():
                setattr(existing, k, v)
        else:
            db.add(GisPhotometricResult(zone_id=zone_id, **data))
        rows_out.append(data)
        imported += 1
    db.commit()
    return {"imported": imported, "rows": rows_out}


# ═══════════════════════════════════════════════════════════════════════════
# UI CONFIG  (per-project key-value)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/api/projects/{project_id}/ui-config")
async def gis_ui_config_get(project_id: int, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    rows = db.query(GisProjectUiConfig).filter(
        GisProjectUiConfig.project_id == project_id
    ).all()
    result = {}
    for r in rows:
        result[r.config_key] = r.config_value
    return result


@router.put("/api/projects/{project_id}/ui-config")
async def gis_ui_config_put(project_id: int, body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    ensure_gis_tables()
    for key, value in body.items():
        existing = db.query(GisProjectUiConfig).filter(
            GisProjectUiConfig.project_id == project_id,
            GisProjectUiConfig.config_key == key
        ).first()
        if existing:
            existing.config_value = value if isinstance(value, dict) else {"value": value}
            existing.updated_at = datetime.now(timezone.utc)
        else:
            db.add(GisProjectUiConfig(
                project_id=project_id,
                config_key=key,
                config_value=value if isinstance(value, dict) else {"value": value},
            ))
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════
# EXPORTS
# ═══════════════════════════════════════════════════════════════════════════

# ── DXF export ────────────────────────────────────────────────────────────

_DXF_CLR = {'motorway':1,'motorway_link':1,'trunk':14,'trunk_link':14,
             'primary':30,'primary_link':30,'secondary':2,'secondary_link':2,
             'tertiary':3,'tertiary_link':3,'residential':4,'unclassified':9,
             'living_street':8,'pedestrian':6,'service':8,'tunnel':5,'trees':82}
_DXF_LW  = {'motorway':50,'trunk':40,'primary':35,'secondary':30,
             'tertiary':25,'residential':18,'tunnel':30}

def _dxf_ldef(name, color, lw=18, lt='CONTINUOUS'):
    return ["0","LAYER","2",name,"70","0","62",str(color),"6",lt,"370",str(lw)]

def _perp_off(lon1, lat1, lon2, lat2, half_m):
    mid = math.radians((lat1 + lat2) / 2)
    cos_mid = math.cos(mid) or 0.001
    dlat_m = (lat2 - lat1) * 111320
    dlon_m = (lon2 - lon1) * 111320 * cos_mid
    dist = math.sqrt(dlat_m**2 + dlon_m**2)
    if dist < 0.01:
        return 0.0, 0.0
    return (-dlon_m/dist * half_m / 111320,
             dlat_m/dist * half_m / (111320 * cos_mid))

@router.get("/api/export/dxf")
async def gis_export_dxf(
    zone_id: str = Query(...),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Export zone as DXF with optional layers."""
    ensure_gis_tables()
    zone = db.get(GisZone, zone_id)
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")

    osm_row = db.get(GisZoneOsmData, zone_id)
    ways = _parse_json_field(osm_row.ways if osm_row else None, [])
    lum_rows = db.query(GisLuminaire).filter(GisLuminaire.zone_id == zone_id).all()
    inv_rows = db.query(GisInventoryLuminaire).filter(GisInventoryLuminaire.zone_id == zone_id).all()
    trees_row = db.get(GisZoneTrees, zone_id)
    tree_data = _parse_json_field(trees_row.trees if trees_row else None, [])
    boundary = _parse_json_field(zone.bounds_polygon, [])

    rtypes = sorted({w.get('type', 'road') for w in ways})

    layers = [("0", 7, 18, "CONTINUOUS")]
    for rt in rtypes:
        layers.append((f"STREETS_{rt.upper()}", _DXF_CLR.get(rt, 7), _DXF_LW.get(rt, 18), "CONTINUOUS"))
    for rt in rtypes:
        layers.append((f"WIDTH_{rt.upper()}", _DXF_CLR.get(rt, 7), 9, "DASHED"))

    seen_names = set()
    for w in ways:
        nm = w.get('name')
        if nm and nm not in seen_names:
            seen_names.add(nm)
            layers.append(("STREET_LABELS", 7, 13, "CONTINUOUS"))
            break

    lum_types = sorted({r.road_type or 'GEN' for r in lum_rows})
    for rt in lum_types:
        layers.append((f"LUM_{rt.upper()}", 50, 18, "CONTINUOUS"))
    if inv_rows:
        layers.append(("INVENTORY", 140, 18, "CONTINUOUS"))
    if boundary:
        layers.append(("ZONE_BOUNDARY", 7, 25, "CONTINUOUS"))
    if tree_data:
        layers.append(("TREES", 82, 18, "CONTINUOUS"))

    L = []
    L += ["0","SECTION","2","HEADER","0","ENDSEC"]
    L += ["0","SECTION","2","TABLES"]
    L += ["0","TABLE","2","LTYPE","70","2"]
    L += ["0","LTYPE","2","CONTINUOUS","70","0","3","Solid","72","65","73","0","40","0.0"]
    L += ["0","LTYPE","2","DASHED","70","0","3","__ __","72","65","73","2","40","0.75",
          "49","0.5","74","0","49","-0.25","74","0"]
    L += ["0","ENDTAB"]
    L += ["0","TABLE","2","LAYER","70",str(len(layers))]
    for nm, clr, lw, lt in layers:
        L += _dxf_ldef(nm, clr, lw, lt)
    L += ["0","ENDTAB","0","ENDSEC"]
    L += ["0","SECTION","2","ENTITIES"]

    # Centerlines
    for w in ways:
        geom = w.get("geom", [])
        if len(geom) < 2:
            continue
        rt = w.get('type', 'road')
        lnm = f"STREETS_{rt.upper()}"
        clr = _DXF_CLR.get(rt, 7)
        for i in range(len(geom) - 1):
            p0, p1 = geom[i], geom[i + 1]
            L += ["0","LINE","8",lnm,"62",str(clr),
                  "10",f"{p0['lon']:.6f}","20",f"{p0['lat']:.6f}","30","0.0",
                  "11",f"{p1['lon']:.6f}","21",f"{p1['lat']:.6f}","31","0.0"]

    # Width polygons
    for w in ways:
        geom = w.get("geom", [])
        if len(geom) < 2:
            continue
        rt = w.get('type', 'road')
        lnm = f"WIDTH_{rt.upper()}"
        clr = _DXF_CLR.get(rt, 7)
        half = (w.get('estWidth') or 6.0) / 2.0
        for i in range(len(geom) - 1):
            p0, p1 = geom[i], geom[i + 1]
            dlat, dlon = _perp_off(p0['lon'], p0['lat'], p1['lon'], p1['lat'], half)
            if not dlat and not dlon:
                continue
            for s in (1, -1):
                L += ["0","LINE","8",lnm,"62",str(clr),"370","9",
                      "10",f"{p0['lon']+s*dlon:.6f}","20",f"{p0['lat']+s*dlat:.6f}","30","0.0",
                      "11",f"{p1['lon']+s*dlon:.6f}","21",f"{p1['lat']+s*dlat:.6f}","31","0.0"]

    # Street labels
    seen = set()
    for w in ways:
        nm = w.get('name')
        if not nm or nm in seen:
            continue
        seen.add(nm)
        geom = w.get("geom", [])
        if not geom:
            continue
        mid = geom[len(geom) // 2]
        L += ["0","TEXT","8","STREET_LABELS","62","7",
              "10",f"{mid['lon']:.6f}","20",f"{mid['lat']:.6f}","30","0.0",
              "40","0.000045","1",nm[:63]]

    # Luminaires
    for r in lum_rows:
        rt = r.road_type or 'GEN'
        L += ["0","POINT","8",f"LUM_{rt.upper()}","62","50",
              "10",f"{r.lon:.6f}","20",f"{r.lat:.6f}","30","0.0"]

    # Inventory
    for r in inv_rows:
        L += ["0","POINT","8","INVENTORY","62","140",
              "10",f"{r.lon:.6f}","20",f"{r.lat:.6f}","30","0.0"]

    # Boundary
    n = len(boundary)
    for i in range(n):
        p0 = boundary[i]
        p1 = boundary[(i + 1) % n]
        lat0, lon0 = (p0[0], p0[1]) if isinstance(p0, list) else (p0.get('lat'), p0.get('lon'))
        lat1, lon1 = (p1[0], p1[1]) if isinstance(p1, list) else (p1.get('lat'), p1.get('lon'))
        L += ["0","LINE","8","ZONE_BOUNDARY","62","7","370","25",
              "10",f"{lon0:.6f}","20",f"{lat0:.6f}","30","0.0",
              "11",f"{lon1:.6f}","21",f"{lat1:.6f}","31","0.0"]

    # Trees
    for t in tree_data:
        L += ["0","POINT","8","TREES","62","82",
              "10",f"{t.get('lon',0):.6f}","20",f"{t.get('lat',0):.6f}","30","0.0"]

    L += ["0","ENDSEC","0","EOF"]
    dxf_content = "\n".join(L).encode("utf-8")

    return Response(
        content=dxf_content,
        media_type="application/dxf",
        headers={"Content-Disposition": f'attachment; filename="zone_{zone_id}.dxf"'},
    )


# ── Plantilla luminotécnica export ────────────────────────────────────────

class GisPlantillaRow(BaseModel):
    name: str = ""
    description: str = ""
    road_width: float = 7.0
    sidewalk_left: float = 0.0
    sidewalk_right: float = 0.0
    lanes: int = 2
    median_width: float = 0.0
    arrangement: str = "Unilateral"
    height: float = 9.0
    spacing: float = 30.0
    arm_length: float = 1.0
    pole_offset: float = 1.0
    pole_side: str = "right"
    tilt: float = 5.0
    manufacturer: str = "Salvi"
    gama: str = "Clap M"
    difusor: str = "Vidrio ultrawhite transp plano"
    lente: str = "F151"
    led_type: str = "Luxeon HOP 5050"
    power: Optional[float] = None
    cct: int = 3000
    cri: int = 70
    lighting_class: str = "M4"
    mf: float = 0.80
    pavement: str = "R3"

class GisPlantillaRequest(BaseModel):
    zone_id: str
    rows: list[GisPlantillaRow]


@router.post("/api/export/plantilla_luminotecnica")
async def gis_export_plantilla(body: GisPlantillaRequest, user: User = Depends(current_user)):
    """Export plantilla Excel for LuxStudio calculation."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla luminotecnica"

    HEADERS = [
        "Zona", "Tipo via", "Longitud (km)", "N luminarias", "Disposicion",
        "Clase ilum.", "Ancho calzada (m)", "Espaciado (m)",
        "Fabricante", "Gama", "Potencia (W)", "Difusor", "Lente", "Tipo LED",
        "CCT (K)", "CRI", "Brazo (m)", "Inclinacion", "Pavimento", "MF",
    ]
    gold = "C8A96E"
    dark = "1C1C1A"
    hfill = PatternFill("solid", fgColor=gold)
    hfont = Font(bold=True, color=dark, size=10)
    thin = Side(style="thin", color="2A3A4A")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bord

    for r in body.rows:
        ws.append([
            body.zone_id, r.name, r.description, "",
            r.arrangement, r.lighting_class, r.road_width, r.spacing,
            r.manufacturer, r.gama, r.power or "", r.difusor, r.lente, r.led_type,
            r.cct, r.cri, r.arm_length, r.tilt, r.pavement, r.mf,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plantilla_{body.zone_id}.xlsx"'},
    )


# ═══════════════════════════════════════════════════════════════════════════
# AI — Anthropic proxy
# ═══════════════════════════════════════════════════════════════════════════

_DB_SCHEMA_SUMMARY = """
ESQUEMA DE LA BASE DE DATOS (PostgreSQL):
- projects(id, project_name, client, location, ...)
- gis_zones(id, name, type, color, priority, center_lat, center_lon, ...)
- gis_zone_config(zone_id, spacing[m], watt_hps, watt_led, efficacy, hours_night)
- gis_zone_osm_data(zone_id, km_by_type[JSON:{primary,secondary,tertiary,...}], ways[JSON], source)
- gis_luminaires(id, zone_id, road_type, lighting_class, street_name, lat, lon, watts, spacing, ...)
- gis_inventory_luminaires(id, zone_id, lat, lon, power_w, height_m, brand, model, lamp_type, ...)
- gis_photometric_results(id, zone_id, segment_name, match_key, road_width, spacing, ...)
- gis_project_ui_config(project_id, config_key, config_value[JSON])
"""

@router.post("/api/ai/ask")
async def gis_ai_ask(body: dict, user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Ask Anthropic Claude a question about the project data."""
    _AI_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
    if not _AI_KEY:
        raise HTTPException(status_code=501, detail="ANTHROPIC_API_KEY not configured")

    _AI_MODEL = os.environ.get("SALVI_AI_MODEL", "claude-haiku-4-5-20251001")
    project_id = body.get("project_id")
    question = body.get("question", "")

    # Build project context
    context_lines = [_DB_SCHEMA_SUMMARY.strip(), ""]
    if project_id:
        proj = db.get(Project, project_id)
        if proj:
            context_lines.append(f"PROYECTO: {proj.project_name} (id={proj.id})")
            zones = db.query(GisZone).filter(GisZone.project_id == proj.id).all()
            for z in zones:
                osm = db.get(GisZoneOsmData, z.id)
                km_by_type = _parse_json_field(osm.km_by_type if osm else None, {})
                lum_count = db.query(GisLuminaire).filter(GisLuminaire.zone_id == z.id).count()
                total_km = sum(float(v) for v in km_by_type.values() if v)
                cfg = db.get(GisZoneConfig, z.id)
                spacing = cfg.spacing if cfg else 30
                watt_led = cfg.watt_led if cfg else 60
                hours = cfg.hours_night if cfg else 11.5
                n_est = int(total_km * 1000 / spacing) if spacing > 0 else 0
                kw_total = n_est * watt_led / 1000
                mwh_yr = kw_total * hours * 365 / 1000
                context_lines.append(f"\nZONA: {z.name} (id={z.id})")
                if km_by_type:
                    for t, km in sorted(km_by_type.items(), key=lambda x: -float(x[1] or 0)):
                        if km:
                            context_lines.append(f"  - {t}: {float(km):.2f} km")
                    context_lines.append(f"  Total red: {total_km:.2f} km")
                context_lines.append(f"  Config: espaciado={spacing}m, LED={watt_led}W, {hours}h/noche")
                context_lines.append(f"  Lums estimadas: {n_est} | {kw_total:.1f} kW | {mwh_yr:.1f} MWh/año")
                if lum_count:
                    context_lines.append(f"  Lums diseñadas: {lum_count}")

    context = "\n".join(context_lines)
    prompt = f"""{context}

PREGUNTA DEL USUARIO: {question}

Responde de forma clara y concisa, usando datos del proyecto cuando sea relevante.
Si no tienes suficiente información, dilo directamente.
"""
    # Call Anthropic API
    import urllib.request as _ur
    import urllib.error as _ue
    body_json = json.dumps({
        "model": _AI_MODEL,
        "max_tokens": 2000,
        "system": "Eres un asistente experto en alumbrado público que analiza proyectos de iluminación urbana.",
        "messages": [{"role": "user", "content": prompt}],
    }).encode()
    req = _ur.Request(
        "https://api.anthropic.com/v1/messages", data=body_json,
        headers={"Content-Type": "application/json", "x-api-key": _AI_KEY,
                 "anthropic-version": "2023-06-01"}
    )
    try:
        with _ur.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())
            answer = data["content"][0]["text"]
            usage = data.get("usage", {})
            return {"answer": answer, "usage": usage, "context": context}
    except _ue.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"API error: {e.code}: {e.read().decode('utf-8','replace')}")


# ═══════════════════════════════════════════════════════════════════════════
# DB query (safe read-only SQL)
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/api/db/query")
async def gis_db_query(body: dict, user: User = Depends(current_user)):
    """Execute a safe read-only SQL query. Same as old /api/db/query."""
    from ..database import engine
    sql = body.get("query", "").strip()
    if not sql:
        raise HTTPException(status_code=400, detail="query required")
    upper = sql.upper()
    if not (upper.startswith("SELECT") or upper.startswith("WITH")):
        raise HTTPException(status_code=400, detail="Solo SELECT o WITH")
    if sql.count(";") > 1 or (sql.endswith(";") and sql[:-1].count(";") > 0):
        raise HTTPException(status_code=400, detail="Multiples sentencias no permitidas")
    with engine.connect() as conn:
        result = conn.exec_driver_sql(sql)
        cols = list(result.keys()) if result.keys() else []
        rows_data = [list(r) for r in result.fetchmany(500)]
    return {"columns": cols, "rows": rows_data}


# require_admin is defined near the top of this file
