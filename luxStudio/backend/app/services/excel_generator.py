from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..schemas.models import CalculationResult
from .i18n import translator


def _headers(t):
    return [
        t("excel.model_id"),
        t("excel.arrangement"),
        t("excel.height"),
        t("excel.spacing"),
        t("excel.sidewalk_width"),
        t("excel.road_width"),
        t("excel.arm_length"),
        t("excel.lighting_class"),
        t("excel.maintenance"),
        t("excel.cct"),
        "CRI",
        t("excel.proposed_luminaire"),
        t("excel.optic"),
        t("excel.tilt"),
        t("excel.proposed_power"),
        "Lm (cd/m2)\nEm (lux)",
        "UNIFORMIDAD\nUo\nEmin (Lux)",
        "UI",
        "TI",
        "SR",
        t("excel.project_efficiency"),
        "int/h",
        "h/a",
        None,
        "P_calc\n(W)",
        "Phi_calc\n(lm)",
        "CRI_calc",
        "LDT base\n/interp",
        "Lm/Em\ncalc",
        "Uo\ncalc",
        "UI/Ul\ncalc",
        "TI\ncalc",
        "SR\ncalc",
        t("excel.complies"),
        t("excel.notes"),
    ]


def _metric(result: CalculationResult, key: str):
    value = getattr(result, key, None)
    return round(value, 2) if value is not None else None


def _notes(result: CalculationResult) -> str:
    failed = [c.name for c in result.criteria if not c.passed]
    return ", ".join(failed) if failed else "OK"


def _spacing_note(lighting_class: str, spacing: float, road_width: float, t) -> str:
    if road_width <= 0:
        return ""
    ratio = spacing / road_width
    if lighting_class in ("M1", "M2"):
        return "OK m1-m2" if ratio < 4 else t("excel.excessive_spacing")
    if lighting_class in ("M3", "M4", "M5"):
        return "OK m3-m4-m5" if ratio < 4.5 else t("excel.excessive_spacing")
    return ""


def _optic_note(height: float, road_width: float, t) -> str:
    if road_width <= 0:
        return ""
    ratio = height / road_width
    if ratio <= 1:
        return "f151"
    if ratio > 1.5:
        return t("excel.watch")
    return "f2md"


def _sidewalk_value(left: float, right: float):
    return left if abs(left - right) < 0.001 else f"L {left:.1f} / R {right:.1f}"


def _row_values(result: CalculationResult, t, tramo_name: str = "MODELO 1"):
    cfg = result.config
    lum = result.luminaire
    main_metric = result.Lavg if result.mode == "ME" else result.Eavg
    uniformity = result.Uo if result.mode == "ME" else result.Emin
    mf_origen = float(getattr(lum, "mf_origen", 1.0) or 1.0)
    mf_efectivo = float(cfg.mf) / mf_origen if mf_origen > 0 else float(cfg.mf)

    return [
        tramo_name,
        cfg.arrangement,
        cfg.height,
        cfg.spacing,
        _sidewalk_value(cfg.sidewalk_left, cfg.sidewalk_right),
        cfg.road_width,
        cfg.arm_length,
        cfg.lighting_class,
        mf_efectivo,
        f"{cfg.cct}K",
        cfg.cri,
        lum.luminaire_name,
        lum.optic_family,
        cfg.tilt,
        cfg.power,
        _metric(result, "Lavg") if result.mode == "ME" else _metric(result, "Eavg"),
        _metric(result, "Uo") if result.mode == "ME" else _metric(result, "Emin"),
        _metric(result, "Ul"),
        _metric(result, "TI"),
        _metric(result, "SR"),
        None,
        _spacing_note(cfg.lighting_class, cfg.spacing, cfg.road_width, t),
        _optic_note(cfg.height, cfg.road_width, t),
        None,
        lum.power,
        round(lum.flux, 0),
        lum.cri,
        lum.filename,
        round(main_metric, 2) if main_metric is not None else None,
        round(uniformity, 2) if uniformity is not None else None,
        _metric(result, "Ul"),
        _metric(result, "TI"),
        _metric(result, "SR"),
        t("status.pass_short") if result.compliant else t("status.fail_short"),
        _notes(result),
    ]


def generate_excel(result: CalculationResult, project: dict | None = None, tramo_name: str = "MODELO 1") -> bytes:
    t = translator(result.config.language)
    headers = _headers(t)
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    if project:
        ws.append(["1. Proyecto / Cliente", None, None, None, None, None])
        ws.append(["Nombre del proyecto", project.get("project_name") or "", "Cliente final", project.get("client") or "", "Localizacion", project.get("location") or ""])
        ws.append(["Proyectista", project.get("designer") or "", "Fecha del estudio", project.get("study_date") or "", "No referencia / plano", project.get("reference") or ""])
        ws.append(["Tipo de calculo / aplicacion", project.get("calculation_type") or "", "Norma aplicable", project.get("standard") or "", "Notas / observaciones", project.get("notes") or ""])
        ws.append([None] * 6)

    header_row = ws.max_row + 1
    data_row = header_row + 1
    ws.append(headers)
    ws.append(_row_values(result, t, tramo_name))
    ws[f"U{data_row}"] = f'=IFERROR(D{data_row}*F{data_row}*P{data_row}*15/O{data_row},"")'

    ws.freeze_panes = f"A{data_row}"
    ws.sheet_view.showGridLines = False

    fills = {
        "input": PatternFill("solid", fgColor="C6E0B4"),
        "selection": PatternFill("solid", fgColor="FFD966"),
        "spacer": PatternFill("solid", fgColor="FFFFFF"),
        "calc": PatternFill("solid", fgColor="2F5597"),
    }
    thin = Side(style="thin", color="A6A6A6")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    if project:
        title_fill = PatternFill("solid", fgColor="D9E2F3")
        input_fill = PatternFill("solid", fgColor="FFFFFF")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.cell(1, 1).fill = title_fill
        ws.cell(1, 1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
        for row in range(2, 5):
            for col in range(1, 7):
                cell = ws.cell(row, col)
                cell.fill = input_fill
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.font = Font(name="Calibri", size=9, bold=col in (1, 3, 5))

    for col in range(1, len(headers) + 1):
        cell = ws.cell(header_row, col)
        if col <= 8:
            cell.fill = fills["input"]
            font_color = "000000"
        elif col <= 23:
            cell.fill = fills["selection"]
            font_color = "000000"
        elif col == 24:
            cell.fill = fills["spacer"]
            font_color = "000000"
        else:
            cell.fill = fills["calc"]
            font_color = "FFFFFF"
        cell.font = Font(name="Calibri", size=8, bold=True, color=font_color)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

        value_cell = ws.cell(data_row, col)
        value_cell.font = Font(name="Calibri", size=9, bold=col in (34, 35))
        value_cell.alignment = Alignment(wrap_text=True, vertical="center")
        value_cell.border = border
        if col == 34:
            value_cell.fill = PatternFill("solid", fgColor="E2F0D9" if result.compliant else "F4CCCC")

    number_formats = {
        "C": "0.00",
        "D": "0.0",
        "E": "0.0",
        "F": "0.0",
        "G": "0.0",
        "I": "0.00",
        "K": "0",
        "N": "0",
        "O": "0",
        "P": "0.000",
        "Q": "0.000",
        "R": "0.000",
        "S": "0.0",
        "T": "0.000",
        "U": "0.00",
        "Y": "0.0",
        "Z": "#,##0",
        "AA": "0",
        "AC": "0.000",
        "AD": "0.000",
        "AE": "0.000",
        "AF": "0.0",
        "AG": "0.000",
    }
    for col, number_format in number_formats.items():
        ws[f"{col}{data_row}"].number_format = number_format

    widths = {
        "A": 16, "B": 14, "C": 10, "D": 15, "E": 15, "F": 15, "G": 13,
        "H": 14, "I": 18, "J": 18, "K": 8, "L": 28, "M": 14, "N": 16,
        "O": 16, "P": 14, "Q": 16, "R": 10, "S": 8, "T": 8, "U": 16,
        "V": 22, "W": 12, "X": 5, "Y": 11, "Z": 13, "AA": 10, "AB": 30,
        "AC": 13, "AD": 12, "AE": 12, "AF": 10, "AG": 10, "AH": 10, "AI": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if project:
        ws.row_dimensions[1].height = 22
        for row in range(2, 5):
            ws.row_dimensions[row].height = 28
    ws.row_dimensions[header_row].height = 48
    ws.row_dimensions[data_row].height = 38
    ws.auto_filter.ref = f"A{header_row}:AI{data_row}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_multi_excel(results: list[CalculationResult], project: dict | None = None, tramo_names: list[str] | None = None) -> bytes:
    """Generate one Excel workbook with one row per calculation result."""
    if not results:
        raise ValueError("At least one result is required")

    t = translator(results[0].config.language)
    headers = _headers(t)
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    if project:
        ws.append(["1. Proyecto / Cliente", None, None, None, None, None])
        ws.append(["Nombre del proyecto", project.get("project_name") or "", "Cliente final", project.get("client") or "", "Localizacion", project.get("location") or ""])
        ws.append(["Proyectista", project.get("designer") or "", "Fecha del estudio", project.get("study_date") or "", "No referencia / plano", project.get("reference") or ""])
        ws.append(["Tipo de calculo / aplicacion", project.get("calculation_type") or "", "Norma aplicable", project.get("standard") or "", "Notas / observaciones", project.get("notes") or ""])
        ws.append([None] * 6)

    header_row = ws.max_row + 1
    first_data_row = header_row + 1
    ws.append(headers)

    for i, result in enumerate(results):
        tramo_name = tramo_names[i] if tramo_names and i < len(tramo_names) else f"Tramo {i + 1}"
        ws.append(_row_values(result, t, tramo_name))
        row_num = first_data_row + i
        ws[f"U{row_num}"] = f'=IFERROR(D{row_num}*F{row_num}*P{row_num}*15/O{row_num},"")'

    last_data_row = ws.max_row

    ws.freeze_panes = f"A{first_data_row}"
    ws.sheet_view.showGridLines = False

    fills = {
        "input": PatternFill("solid", fgColor="C6E0B4"),
        "selection": PatternFill("solid", fgColor="FFD966"),
        "spacer": PatternFill("solid", fgColor="FFFFFF"),
        "calc": PatternFill("solid", fgColor="2F5597"),
    }
    thin = Side(style="thin", color="A6A6A6")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    if project:
        title_fill = PatternFill("solid", fgColor="D9E2F3")
        input_fill = PatternFill("solid", fgColor="FFFFFF")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.cell(1, 1).fill = title_fill
        ws.cell(1, 1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
        for row in range(2, 5):
            for col in range(1, 7):
                cell = ws.cell(row, col)
                cell.fill = input_fill
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.font = Font(name="Calibri", size=9, bold=col in (1, 3, 5))

    # Header row styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(header_row, col)
        if col <= 8:
            cell.fill = fills["input"]
            font_color = "000000"
        elif col <= 23:
            cell.fill = fills["selection"]
            font_color = "000000"
        elif col == 24:
            cell.fill = fills["spacer"]
            font_color = "000000"
        else:
            cell.fill = fills["calc"]
            font_color = "FFFFFF"
        cell.font = Font(name="Calibri", size=8, bold=True, color=font_color)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

    # Data rows styling
    number_formats = {
        "C": "0.00", "D": "0.0", "E": "0.0", "F": "0.0", "G": "0.0",
        "I": "0.00", "K": "0", "N": "0", "O": "0",
        "P": "0.000", "Q": "0.000", "R": "0.000", "S": "0.0", "T": "0.000",
        "U": "0.00", "Y": "0.0", "Z": "#,##0", "AA": "0",
        "AC": "0.000", "AD": "0.000", "AE": "0.000", "AF": "0.0", "AG": "0.000",
    }
    for row_num in range(first_data_row, last_data_row + 1):
        result = results[row_num - first_data_row]
        for col in range(1, len(headers) + 1):
            value_cell = ws.cell(row_num, col)
            value_cell.font = Font(name="Calibri", size=9, bold=col in (34, 35))
            value_cell.alignment = Alignment(wrap_text=True, vertical="center")
            value_cell.border = border
            if col == 34:
                value_cell.fill = PatternFill("solid", fgColor="E2F0D9" if result.compliant else "F4CCCC")
            col_letter = ws.cell(row=row_num, column=col).column_letter
            if col_letter in number_formats:
                ws[f"{col_letter}{row_num}"].number_format = number_formats[col_letter]

    widths = {
        "A": 16, "B": 14, "C": 10, "D": 15, "E": 15, "F": 15, "G": 13,
        "H": 14, "I": 18, "J": 18, "K": 8, "L": 28, "M": 14, "N": 16,
        "O": 16, "P": 14, "Q": 16, "R": 10, "S": 8, "T": 8, "U": 16,
        "V": 22, "W": 12, "X": 5, "Y": 11, "Z": 13, "AA": 10, "AB": 30,
        "AC": 13, "AD": 12, "AE": 12, "AF": 10, "AG": 10, "AH": 10, "AI": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    if project:
        ws.row_dimensions[1].height = 22
        for row in range(2, 5):
            ws.row_dimensions[row].height = 28
    ws.row_dimensions[header_row].height = 48
    for row_num in range(first_data_row, last_data_row + 1):
        ws.row_dimensions[row_num].height = 38

    ws.auto_filter.ref = f"A{header_row}:AI{last_data_row}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
