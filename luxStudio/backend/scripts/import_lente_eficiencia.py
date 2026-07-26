"""Import lens efficiency (Eficiencia Lente) from the Lentes sheet of
``Referencias_productos_pcb_go.xlsx`` into the ``lentes`` table.

Usage:
    python scripts/import_lente_eficiencia.py
    python scripts/import_lente_eficiencia.py --xlsx path/to/file.xlsx
    python scripts/import_lente_eficiencia.py --dry-run
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from app.core.text_utils import norm  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.models import Lente  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s :: %(message)s")
log = logging.getLogger("import_lente_eficiencia")


def _as_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().replace(",", ".").replace(" ", "")
    try:
        v = float(value)
        return v if not (v != v) else None  # nan → None
    except (TypeError, ValueError):
        return None


def read_eficiencia(xlsx_path: Path) -> dict[str, float | None]:
    """Return {lens_name: eficiencia} deduplicated (first non-null wins)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "Lentes" not in wb.sheetnames:
        log.warning("Sheet 'Lentes' not found; no data imported.")
        return {}
    ws = wb["Lentes"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col = {str(v).strip(): i for i, v in enumerate(header)}
    desc_idx = col.get("Descripcion")
    eff_idx = col.get("Eficiencia Lente")
    if desc_idx is None or eff_idx is None:
        log.warning("Required columns not found; got: %s", list(col))
        return {}

    result: dict[str, float | None] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = norm(row[desc_idx]) if desc_idx < len(row) else ""
        if not name:
            continue
        eff = _as_float(row[eff_idx]) if eff_idx < len(row) else None
        if name not in result:
            result[name] = eff
        elif eff is not None and result[name] is None:
            result[name] = eff
    log.info("Unique lenses with efficiency data: %d", len(result))
    return result


def import_eficiencia(xlsx_path: Path, dry_run: bool = False) -> dict:
    eficiencia_map = read_eficiencia(xlsx_path)
    if dry_run:
        return {"lenses_read": len(eficiencia_map), "updated": 0, "created": 0}

    db = SessionLocal()
    stats = {"lenses_read": len(eficiencia_map), "updated": 0, "created": 0}
    try:
        for name, eff in eficiencia_map.items():
            lente = db.query(Lente).filter(Lente.name == name).first()
            if lente is None:
                lente = Lente(name=name, eficiencia=eff)
                db.add(lente)
                stats["created"] += 1
            else:
                lente.eficiencia = eff
                stats["updated"] += 1
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
    return stats


def main() -> None:
    default_xlsx = ROOT.parent / "docs" / "Referencias_productos_pcb_go.xlsx"
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=default_xlsx)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if not args.xlsx.exists():
        raise SystemExit(f"xlsx not found: {args.xlsx}")

    stats = import_eficiencia(args.xlsx, dry_run=args.dry_run)
    print("\n=== Resumen ===")
    if args.dry_run:
        print(f"  Lenses leídas del xlsx:  {stats['lenses_read']}")
    else:
        print(f"  Lenses leídas:  {stats['lenses_read']}")
        print(f"  Creadas:        {stats['created']}")
        print(f"  Actualizadas:   {stats['updated']}")


if __name__ == "__main__":
    main()
