"""Import the LED / PCB / driver catalogs and 4-tuple -> LED binding from
``Referencias_productos_pcb_go.xlsx``.

This script is the bridge between the (temporary) Excel data source and
the database: once it has been run, the application no longer touches
the xlsx. Re-running it is idempotent (it UPSERTs by ``LED_REF``,
``DR_REF``, ``PCB_REF`` and 4-tuple).

What the script reads
---------------------
The xlsx has a sheet ``SALVI Lighting$Param_ Configura`` with a single
wide table; the rows of interest are those where a particular column
is non-empty (different rows populate different column groups):

- **LED catalog** (rows where ``LED_REF`` is set):
  ``LED_REF``, ``LED_Desc Corta``, ``LED_TIPO``,
  ``LED_ pot Max lum``, ``LED_I Max led``, ``LED_Pot Max Ajustada``.

- **Driver catalog** (rows where ``DR_REF`` is set):
  ``DR_REF``, ``DR_ Pot Max driver``.

- **PCB catalog** (rows where ``PCB_REF`` is set):
  ``PCB_REF``, ``PCB_No Drivers``, ``PCB_V Nominal``,
  ``PCB_No LED``, ``PCB_No Circuitos``, ``PCB_Imax LED``.

The 4-tuple ``(gama, difusor, lente, led_type)`` comes from the sheet
``Variantes SALVI``:

- col ``G_Gama``      → gama
- col ``D_Descr corta`` → difusor (falls back to ``D_Ref`` only if the
  description is empty)
- col ``Lente_Ref``   → lente
- col ``LED_REF``     → led_type
- col ``Temp_Color``  → CCT (informational only; ``LED_REF`` is the
  identifier the configurator uses to look up the cap).
- col ``Lista materiales`` (col 20) holds the driver and PCB refs as a
  concatenated string — useful for diagnostics, not used for the cap.

Integrity check
---------------
A 4-tuple may map to several ``LED_REF``s in the xlsx because the same
configurable luminaire can be built with more or fewer LEDs/PCBs. For
each 4-tuple the script keeps the LED with the **highest**
``LED_Pot Max Ajustada``: that is the real ceiling the luminaire can
support. It emits a ``warnings.warn`` + ``logging.warning`` listing
every alternative it dropped, so an operator can audit the choice. The
same 4-tuple is **also** a key in ``valid_combinations``;
the script logs any 4-tuple in ``Variantes SALVI`` that has no matching
``valid_combinations`` row (an audit hint for the operator).

Usage
-----
    python scripts/import_salvi_leds.py
    python scripts/import_salvi_leds.py --xlsx path/to/file.xlsx
    python scripts/import_salvi_leds.py --dry-run
"""
from __future__ import annotations

import argparse
import logging
import sys
import warnings
from collections import defaultdict
from pathlib import Path

# Make ``app`` importable when the script is run from any cwd.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from app.core.text_utils import norm  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.models import (  # noqa: E402
    Difusor,
    Driver,
    Gama,
    LED,
    Lente,
    LedType,
    LuminaireLED,
    PCB,
    ValidCombination,
)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s :: %(message)s",
)
log = logging.getLogger("import_salvi_leds")


CONFIG_SHEET = "SALVI Lighting$Param_ Configura"
VARIANTES_SHEET = "Variantes SALVI"
LENTES_SHEET = "Lentes"

# How we recognise a "real" ref vs header / footer noise in the column.
# Catalog refs are short tokens (M18, 16W, 1ME2432, ...). Anything longer
# than 24 chars in the LED_REF/DR_REF/PCB_REF column is a description
# that ended up in the wrong column.
MAX_REF_LEN = 24


_WARNING_COUNTER: list[int] = [0]


def _as_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        # The Spanish locale uses ',' as decimal separator.
        value = value.strip().replace(" ", "").replace(",", ".")
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_int(value) -> int | None:
    f = _as_float(value)
    return int(round(f)) if f is not None else None


def _is_real_ref(value) -> bool:
    """Reject the header literal and any non-ref noise."""
    if value is None:
        return False
    text = norm(value)
    if not text or text == value and not text:  # empty
        return False
    if text in {"LED_REF", "DR_REF", "PCB_REF"}:
        return False
    if len(text) > MAX_REF_LEN:
        return False
    return True


def _warn(message: str, *args) -> None:
    """Emit a warning both to ``logging`` and via ``warnings.warn``.

    Going through both channels makes the message visible whether the
    script is run from a terminal, a notebook, or CI.
    """
    log.warning(message, *args)
    warnings.warn(message % args if args else message, stacklevel=2)
    _WARNING_COUNTER[0] += 1


# ---------------------------------------------------------------------------
# Param_Configura reader (single wide table, filtered by column)
# ---------------------------------------------------------------------------


def _column_index(ws) -> dict[str, int]:
    """Return ``{column_name: 0-based index}`` for the header row."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(name).strip(): idx for idx, name in enumerate(header)}


def _read_param_configura(xlsx_path: Path) -> tuple[dict, dict, dict]:
    log.info("Loading %s :: sheet %r", xlsx_path, CONFIG_SHEET)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[CONFIG_SHEET]
    col = _column_index(ws)

    required = {
        "LED_REF": col.get("LED_REF"),
        "LED_Desc Corta": col.get("LED_Desc Corta"),
        "LED_TIPO": col.get("LED_TIPO"),
        "LED_ pot Max lum": col.get("LED_ pot Max lum"),
        "LED_I Max led": col.get("LED_I Max led"),
        "LED_Pot Max Ajustada": col.get("LED_Pot Max Ajustada"),
        "DR_REF": col.get("DR_REF"),
        "DR_ Pot Max driver": col.get("DR_ Pot Max driver"),
        "PCB_REF": col.get("PCB_REF"),
        "PCB_No Drivers": col.get("PCB_No Drivers"),
        "PCB_V Nominal": col.get("PCB_V Nominal"),
        "PCB_No LED": col.get("PCB_No LED"),
        "PCB_No Circuitos": col.get("PCB_No Circuitos"),
        "PCB_Imax LED": col.get("PCB_Imax LED"),
    }
    missing = [name for name, idx in required.items() if idx is None]
    if missing:
        raise SystemExit(
            f"Param_Configura is missing expected columns: {missing}"
        )

    leds: dict[str, dict] = {}
    drivers: dict[str, dict] = {}
    pcbs: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # LED block: a row is an LED catalog row if LED_REF is set.
        led_ref_raw = row[required["LED_REF"]]
        if _is_real_ref(led_ref_raw):
            ref = norm(led_ref_raw)
            pmax_raw = _as_float(row[required["LED_Pot Max Ajustada"]])
            if ref in leds:
                existing = leds[ref]
                existing_pmax = existing.get("pmax_ajustada")
                if pmax_raw is not None and (existing_pmax is None or pmax_raw > existing_pmax):
                    existing["pmax_ajustada"] = pmax_raw
            else:
                leds[ref] = {
                    "ref": ref,
                    "desc_corta": norm(row[required["LED_Desc Corta"]]) or None,
                    "tipo": norm(row[required["LED_TIPO"]]) or None,
                    "pmax_lum": _as_float(row[required["LED_ pot Max lum"]]),
                    "i_max_led": _as_float(row[required["LED_I Max led"]]),
                    "pmax_ajustada": pmax_raw,
                }

        # Driver block: a row is a driver catalog row if DR_REF is set.
        dr_ref_raw = row[required["DR_REF"]]
        if _is_real_ref(dr_ref_raw):
            ref = norm(dr_ref_raw)
            drivers[ref] = {
                "ref": ref,
                "pmax": _as_float(row[required["DR_ Pot Max driver"]]),
            }

        # PCB block: a row is a PCB catalog row if PCB_REF is set.
        pcb_ref_raw = row[required["PCB_REF"]]
        if _is_real_ref(pcb_ref_raw):
            ref = norm(pcb_ref_raw)
            pcbs[ref] = {
                "ref": ref,
                "no_drivers": _as_int(row[required["PCB_No Drivers"]]),
                "v_nominal": _as_float(row[required["PCB_V Nominal"]]),
                "no_led": _as_int(row[required["PCB_No LED"]]),
                "no_circuitos": _as_int(row[required["PCB_No Circuitos"]]),
                "imax_led": _as_float(row[required["PCB_Imax LED"]]),
            }

    log.info("LED catalog: %d rows", len(leds))
    log.info("Driver catalog: %d rows", len(drivers))
    log.info("PCB catalog: %d rows", len(pcbs))
    return leds, drivers, pcbs


# ---------------------------------------------------------------------------
# Variantes SALVI reader
# ---------------------------------------------------------------------------


def _read_variantes(xlsx_path: Path) -> list[dict]:
    """Read the 4-tuple -> LED binding sheet.

    Returns a list of dicts:
        ``{gama, difusor, lente, led_type, led_ref, cct}``
    plus optional ``pcb_ref`` and ``dr_ref`` extracted from the
    ``Lista materiales`` column (concatenated string), used only for
    audit / diagnostics.
    """
    log.info("Loading %s :: sheet %r", xlsx_path, VARIANTES_SHEET)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[VARIANTES_SHEET]
    col = _column_index(ws)

    required = ("G_Gama", "D_Ref", "Lente_Ref", "LED_REF", "Temp_Color")
    for name in required:
        if col.get(name) is None:
            raise SystemExit(
                f"Variantes SALVI is missing required column {name!r}"
            )

    cct_idx = col.get("Temp_Color")
    led_ref_idx = col.get("LED_REF")
    lista_idx = col.get("Lista materiales")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        gama = norm(row[col["G_Gama"]])
        difusor = norm(row[col["D_Descr corta"]]) if col.get("D_Descr corta") is not None else ""
        difusor = difusor or norm(row[col["D_Ref"]])
        lente = norm(row[col["Lente_Ref"]])
        led_type = norm(row[led_ref_idx])
        if not (gama and difusor and lente):
            # Incomplete 4-tuple: skip (these will be flagged by the
            # ``valid_combinations`` audit below if applicable).
            continue

        # ``Temp_Color`` is a number like 4000. ``Lista materiales`` is
        # a concatenation of refs; we keep the first two non-empty
        # tokens as PCB / Driver for diagnostics only.
        pcb_ref, dr_ref = _split_lista_materiales(
            row[lista_idx] if lista_idx is not None and lista_idx < len(row) else None
        )

        rows.append({
            "gama": gama,
            "difusor": difusor,
            "lente": lente,
            "led_type": led_type or None,
            "led_ref": led_type or None,
            "cct": _as_int(row[cct_idx]) if cct_idx < len(row) else None,
            "pcb_ref": pcb_ref,
            "dr_ref": dr_ref,
        })
    log.info("Variantes SALVI: %d data rows", len(rows))
    return rows


def _split_lista_materiales(value) -> tuple[str | None, str | None]:
    """Best-effort split of ``Lista materiales`` into (pcb_ref, dr_ref).

    The column is a concatenation of refs without a strict separator;
    the driver ref always starts with ``MS`` or ``ST`` in the data we
    have, so we look for the first token that matches that pattern.
    """
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    tokens = [t.strip() for t in text.replace(";", " ").split() if t.strip()]
    pcb_ref = tokens[0] if tokens else None
    dr_ref = next(
        (t for t in tokens[1:] if t.upper().startswith(("MS", "ST"))),
        None,
    )
    return norm(pcb_ref) or None, norm(dr_ref) or None


def _read_lentes_mapping(xlsx_path: Path) -> dict[str, str]:
    """Return ``{bbdd_lente: variantes_lente}`` for the Lentes sheet.

    The sheet has columns ``Descripcion`` (e.g. ``F151`` — the BBDD
    code) and `` ref lente `` (e.g. ``M3`` — the Variantes SALVI
    code).  We index the BBDD name → Variantes name.  Either side may
    be empty in some rows; those are skipped.
    """
    log.info("Loading %s :: sheet %r", xlsx_path, LENTES_SHEET)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if LENTES_SHEET not in wb.sheetnames:
        log.warning("Lentes sheet not present; lens mapping will be empty.")
        return {}
    ws = wb[LENTES_SHEET]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col = {str(name).strip(): idx for idx, name in enumerate(header)}
    desc_idx = col.get("Descripcion")
    ref_idx = col.get("ref lente")
    if desc_idx is None or ref_idx is None:
        log.warning(
            "Lentes sheet is missing expected columns (Descripcion / ' ref lente ')."
        )
        return {}
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        bbdd = norm(row[desc_idx]) if desc_idx < len(row) else ""
        variantes = norm(row[ref_idx]) if ref_idx < len(row) else ""
        if bbdd and variantes:
            mapping[bbdd] = variantes
    log.info("Lens mapping rows: %d", len(mapping))
    return mapping


# ---------------------------------------------------------------------------
# Upserts
# ---------------------------------------------------------------------------


def _upsert_led(db, led_data: dict) -> int:
    ref = led_data["ref"]
    instance = db.query(LED).filter(LED.led_ref == ref).first()
    if instance is None:
        instance = LED(led_ref=ref)
        db.add(instance)
    instance.led_desc_corta = led_data.get("desc_corta")
    instance.led_tipo = led_data.get("tipo")
    instance.pmax_lum = led_data.get("pmax_lum")
    instance.i_max_led = led_data.get("i_max_led")
    instance.pmax_ajustada = led_data.get("pmax_ajustada")
    db.flush()
    return instance.id


def _upsert_driver(db, dr_data: dict) -> int:
    ref = dr_data["ref"]
    instance = db.query(Driver).filter(Driver.dr_ref == ref).first()
    if instance is None:
        instance = Driver(dr_ref=ref)
        db.add(instance)
    instance.dr_pot_max_driver = dr_data.get("pmax")
    db.flush()
    return instance.id


def _upsert_pcb(db, pcb_data: dict) -> int:
    ref = pcb_data["ref"]
    instance = db.query(PCB).filter(PCB.pcb_ref == ref).first()
    if instance is None:
        instance = PCB(pcb_ref=ref)
        db.add(instance)
    instance.pcb_no_drivers = pcb_data.get("no_drivers")
    instance.pcb_v_nominal = pcb_data.get("v_nominal")
    instance.pcb_no_led = pcb_data.get("no_led")
    instance.pcb_no_circuitos = pcb_data.get("no_circuitos")
    instance.pcb_imax_led = pcb_data.get("imax_led")
    db.flush()
    return instance.id


def _get_or_create_dim(db, model, name: str) -> int:
    instance = db.query(model).filter(model.name == name).first()
    if instance is None:
        instance = model(name=name)
        db.add(instance)
        db.flush()
    return instance.id


def _dim_id(db, model, name: str) -> int | None:
    instance = db.query(model).filter(model.name == name).first()
    return instance.id if instance else None


# ---------------------------------------------------------------------------
# Main import routine
# ---------------------------------------------------------------------------


def import_xlsx(xlsx_path: Path, dry_run: bool = False) -> dict:
    leds_raw, drivers_raw, pcbs_raw = _read_param_configura(xlsx_path)
    variantes = _read_variantes(xlsx_path)
    lente_mapping = _read_lentes_mapping(xlsx_path)
    # Variantes SALVI stores the lens as the short code (M3, 2D, ...);
    # the rest of the application uses the BBDD code (F151, F2MD, ...).
    # Build the inverse mapping so the cap lookup joins cleanly on
    # the BBDD lens table.
    variantes_to_bbdd: dict[str, str] = {v: k for k, v in lente_mapping.items()}

    if dry_run:
        return {
            "leds": len(leds_raw),
            "drivers": len(drivers_raw),
            "pcbs": len(pcbs_raw),
            "variantes": len(variantes),
            "lens_mapping": len(lente_mapping),
            "warnings": 0,
        }

    db = SessionLocal()
    stats = {
        "leds_upserted": 0,
        "drivers_upserted": 0,
        "pcbs_upserted": 0,
        "luminaire_leds_upserted": 0,
        "unknown_led_refs": 0,
        "missing_4tuple_in_catalog": 0,
        "lens_translation_warnings": 0,
    }
    try:
        # ``luminaire_leds`` is a derived table: each row is the result
        # of a deterministic 4-tuple -> LED binding using the LED
        # catalog's ``led_tipo`` as the FK.  The previous version of
        # this script keyed the FK on the short ``LED_REF`` from
        # Variantes SALVI, which doesn't match the LDT catalog
        # (descriptive name).  Wipe and rebuild so the FKs are
        # consistent with the LDT catalog after the change.
        db.query(LuminaireLED).delete()
        db.flush()
        for led_data in leds_raw.values():
            _upsert_led(db, led_data)
            stats["leds_upserted"] += 1
        for dr_data in drivers_raw.values():
            _upsert_driver(db, dr_data)
            stats["drivers_upserted"] += 1
        for pcb_data in pcbs_raw.values():
            _upsert_pcb(db, pcb_data)
            stats["pcbs_upserted"] += 1
        db.flush()
        # Group variantes by 4-tuple.  When the same 4-tuple maps to
        # several LED_REFs, keep the highest supported adjusted power.
        grouped: dict[tuple, list[dict]] = defaultdict(list)
        for v in variantes:
            key = (v["gama"], v["difusor"], v["lente"], v["led_type"])
            grouped[key].append(v)

        # Cache all LEDs so the resolution below is in-memory only.
        led_cache: dict[str, "LED"] = {led.led_ref: led for led in db.query(LED).all()}

        # First pass: for every (gama, difusor, lente, short_code) group,
        # pick the LED with the highest adjusted maximum power. This is the per-short-ref
        # reduction; a second pass below collapses the result into the
        # descriptive (led_tipo) 4-tuple so multiple short refs sharing
        # the same series (e.g. M18 + M08 → LUXEON HO 5050) collapse to
        # the highest available cap in the family.
        per_short: dict[tuple, dict] = {}
        for key, group in grouped.items():
            gama_name, difusor_name, lente_name, led_type_name = key

            # Translate the Variantes SALVI lens code to the BBDD
            # code.  When the lens is unknown we drop the entry
            # (already-flagged below) and warn.
            lente_bbdd = variantes_to_bbdd.get(lente_name, lente_name)
            if lente_bbdd != lente_name:
                lente_name = lente_bbdd
                key = (gama_name, difusor_name, lente_name, led_type_name)

            # Filter to only variants that have a known LED_REF.
            known = [v for v in group if v["led_ref"] in led_cache]
            unknown = [v for v in group if v["led_ref"] not in led_cache]
            for v in unknown:
                stats["unknown_led_refs"] += 1
                _warn(
                    "Variantes SALVI refers to LED_REF %r (4-tuple=%r) "
                    "which is not in the LED catalog; skipping.",
                    v["led_ref"], key,
                )
            if not known:
                continue

            if lente_name not in lente_mapping.values() and lente_name not in lente_mapping:
                # The Variantes SALVI lente code didn't translate to a
                # BBDD code AND isn't itself a known BBDD code; flag it
                # so the operator can extend the Lentes sheet.
                stats["lens_translation_warnings"] += 1
                _warn(
                    "Variantes SALVI lente code %r (4-tuple=%r) is not present "
                    "in the Lentes mapping; cap will not be enforced unless the "
                    "Lentes sheet is updated to add this code.",
                    lente_name, key,
                )

            # Pick the LED with the highest pmax_ajustada.  If a LED has
            # pmax_ajustada=None we treat it as -inf and skip it in
            # favour of any LED with a numeric value.
            def _pmax(v: dict) -> float:
                led = led_cache.get(v["led_ref"])
                p = led.pmax_ajustada if led else None
                return p if p is not None else float("-inf")

            chosen = max(known, key=_pmax)
            chosen_led = led_cache.get(chosen["led_ref"])
            if chosen_led is None or chosen_led.pmax_ajustada is None:
                # No LED with a numeric pmax_ajustada in the group: we
                # still record the binding to the first non-null one
                # so the configurator can show "no pmax known" if
                # needed; otherwise the 4-tuple would be silently
                # missing from the cap lookup.
                fallback = known[0]
                chosen_led = led_cache.get(fallback["led_ref"])
                _warn(
                    "No LED with numeric pmax_ajustada in 4-tuple %r; "
                    "using %r (pmax_ajustada=None) as fallback.",
                    key, fallback["led_ref"],
                )

            if len(known) > 1:
                ccts = sorted({v["cct"] for v in known if v["cct"] is not None})
                alts = sorted({v["led_ref"] for v in known})
                _warn(
                    "4-tuple %r maps to multiple LED_REFs %r; keeping the highest "
                    "cap (%r, pmax_ajustada=%s). Affected CCTs: %s",
                    key, alts, chosen_led.led_ref, chosen_led.pmax_ajustada, ccts,
                )

            per_short[key] = {
                "gama_name": gama_name,
                "difusor_name": difusor_name,
                "lente_name": lente_name,
                "led_type_name": led_type_name,
                "led": chosen_led,
            }

        # Second pass: collapse per-short entries that share the same
        # descriptive (gama, difusor, lente, led_tipo).  The LDT catalog
        # keys ``led_type`` by the descriptive name (LUXEON HO 5050),
        # not the short code, so we need the join to match.  When
        # multiple short refs collapse to the same descriptive name
        # (e.g. M18, M08, 12C all → LUXEON HO 5050), we keep the LED
        # with the highest pmax_ajustada.
        per_descriptive: dict[tuple, dict] = {}
        for key, entry in per_short.items():
            gama_name, difusor_name, lente_name, _ = key
            chosen_led = entry["led"]
            led_tipo_name = (chosen_led.led_tipo or "").strip().upper() or None
            if led_tipo_name is None and chosen_led.led_ref:
                _warn(
                    "LED %r has no led_tipo (descriptive name); using LED_REF %r "
                    "as the FK for the cap lookup of 4-tuple %r.",
                    chosen_led.led_ref, chosen_led.led_ref, key,
                )
                led_tipo_name = chosen_led.led_ref
            desc_key = (gama_name, difusor_name, lente_name, led_tipo_name)
            current = per_descriptive.get(desc_key)
            if current is None:
                per_descriptive[desc_key] = {
                    "gama_name": gama_name,
                    "difusor_name": difusor_name,
                    "lente_name": lente_name,
                    "led_type_name": led_tipo_name,
                    "led": chosen_led,
                }
                continue
            cur_led = current["led"]
            cur_p = cur_led.pmax_ajustada if cur_led else None
            cur_p = cur_p if cur_p is not None else float("-inf")
            new_p = chosen_led.pmax_ajustada if chosen_led else None
            new_p = new_p if new_p is not None else float("-inf")
            if new_p > cur_p:
                alts = sorted({cur_led.led_ref, chosen_led.led_ref})
                _warn(
                    "Descriptive 4-tuple %r collapses multiple short LED_REFs %r; "
                    "keeping the highest cap (%r, pmax_ajustada=%s).",
                    desc_key, alts, chosen_led.led_ref, chosen_led.pmax_ajustada,
                )
                per_descriptive[desc_key]["led"] = chosen_led
            else:
                alts = sorted({cur_led.led_ref, chosen_led.led_ref})
                _warn(
                    "Descriptive 4-tuple %r collapses multiple short LED_REFs %r; "
                    "keeping the highest cap (%r, pmax_ajustada=%s).",
                    desc_key, alts, cur_led.led_ref, cur_led.pmax_ajustada,
                )

        # Third pass: insert one row per descriptive 4-tuple.
        for desc_key, entry in per_descriptive.items():
            gama_name = entry["gama_name"]
            difusor_name = entry["difusor_name"]
            lente_name = entry["lente_name"]
            led_type_name = entry["led_type_name"]
            chosen_led = entry["led"]
            gama_id = _get_or_create_dim(db, Gama, gama_name)
            difusor_id = _get_or_create_dim(db, Difusor, difusor_name)
            lente_id = _get_or_create_dim(db, Lente, lente_name)
            led_type_id = (
                _get_or_create_dim(db, LedType, led_type_name)
                if led_type_name else None
            )

            existing = (
                db.query(LuminaireLED)
                .filter(
                    LuminaireLED.gama_id == gama_id,
                    LuminaireLED.difusor_id == difusor_id,
                    LuminaireLED.lente_id == lente_id,
                    LuminaireLED.led_type_id == led_type_id,
                )
                .first()
            )
            if existing is None:
                db.add(LuminaireLED(
                    gama_id=gama_id,
                    difusor_id=difusor_id,
                    lente_id=lente_id,
                    led_type_id=led_type_id,
                    led_id=chosen_led.id,
                ))
                stats["luminaire_leds_upserted"] += 1
            else:
                existing.led_id = chosen_led.id

        # Audit: variants referring to 4-tuples that have no
        # valid_combinations row in the DB.  The cap cannot be
        # enforced for those 4-tuples.  Iterate over the descriptive
        # (post-collapse) 4-tuples so the comparison is apples-to-
        # apples with the BBDD ``valid_combinations`` table.
        vc_keys = {
            (vc.gama_id, vc.difusor_id, vc.lente_id, vc.led_type_id)
            for vc in db.query(ValidCombination).all()
        }
        for desc_key, entry in per_descriptive.items():
            gama_name, difusor_name, lente_name, led_type_name = desc_key
            if not (gama_name and difusor_name and lente_name):
                continue
            gama_id = _dim_id(db, Gama, gama_name)
            difusor_id = _dim_id(db, Difusor, difusor_name)
            lente_id = _dim_id(db, Lente, lente_name)
            led_type_id = _dim_id(db, LedType, led_type_name) if led_type_name else None
            if None in (gama_id, difusor_id, lente_id):
                stats["missing_4tuple_in_catalog"] += 1
                _warn(
                    "4-tuple %r in Variantes SALVI is missing one of "
                    "gama/difusor/lente in the dimension catalog; cap will "
                    "not be enforced for this 4-tuple.",
                    desc_key,
                )
                continue
            if (gama_id, difusor_id, lente_id, led_type_id) not in vc_keys:
                stats["missing_4tuple_in_catalog"] += 1
                _warn(
                    "4-tuple %r in Variantes SALVI is not present in "
                    "valid_combinations (the seed of BBDD_Fotometrias did not "
                    "list it). The cap will not be enforced for this 4-tuple.",
                    desc_key,
                )

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    return stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=ROOT.parent / "Referencias_productos_pcb_go.xlsx",
        help="Path to Referencias_productos_pcb_go.xlsx (default: project root).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read the xlsx but do not touch the database.",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"xlsx not found: {args.xlsx}")

    _WARNING_COUNTER[0] = 0
    stats = import_xlsx(args.xlsx, dry_run=args.dry_run)

    print("\n=== Resumen de import ===")
    if args.dry_run:
        print(f"  (dry-run, no DB writes)")
        print(f"  LEDs leídos:        {stats['leds']}")
        print(f"  Drivers leídos:     {stats['drivers']}")
        print(f"  PCBs leídos:        {stats['pcbs']}")
        print(f"  Variantes leídas:   {stats['variantes']}")
        print(f"  Lentes mapping:     {stats['lens_mapping']}")
    else:
        print(f"  LEDs UPSERT:        {stats['leds_upserted']}")
        print(f"  Drivers UPSERT:     {stats['drivers_upserted']}")
        print(f"  PCBs UPSERT:        {stats['pcbs_upserted']}")
        print(f"  4-tuplas -> LED:    {stats['luminaire_leds_upserted']}")
        print(f"  Warnings emitidos:  {_WARNING_COUNTER[0]}")
        if stats["unknown_led_refs"]:
            print(f"    • LED_REFs no catalogados:   {stats['unknown_led_refs']}")
        if stats["missing_4tuple_in_catalog"]:
            print(f"    • 4-tuplas sin valid_combinations: {stats['missing_4tuple_in_catalog']}")
        if stats["lens_translation_warnings"]:
            print(f"    • Lentes sin mapping:        {stats['lens_translation_warnings']}")


if __name__ == "__main__":
    main()
