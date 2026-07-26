"""Update the difusores table with the canonical long descriptions and
efficiency values from ``DifusoresConfigurador 1.xlsx``.

Usage:
    python scripts/seed_difusores.py
    python scripts/seed_difusores.py --dry-run
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.models import Difusor, Fotometria, ValidCombination, LuminaireLED  # noqa: E402


XLSX = ROOT.parent / "docs" / "DifusoresConfigurador 1.xlsx"
SHEET_GAMA = "gama difusor"
SHEET_EFI = "eficienciaDifusor"


def _norm(s):
    return str(s).strip().upper().replace(" ", "").replace("_", "") if s else ""


def _load_efficiency_map(ws) -> dict[str, float]:
    """Return {upper_name: eficiencia} from the efficiency sheet."""
    efis: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, eff = (row + (None, None))[:2]
        n = _norm(name)
        if n and eff is not None:
            try:
                efis[n] = float(str(eff).replace(",", "."))
            except (ValueError, TypeError):
                pass
    return efis


def _load_gama_mapping(ws) -> dict[str, str]:
    """Return two maps:
    - short_to_long: short desc (col E) → long desc (col C)
    - ref_to_long: ref code (col B) → long desc (col C)

    Handles multiple short codes pointing to the same long desc by
    keeping whatever was found first (they're equivalent).
    """
    short_to_long: dict[str, str] = {}
    ref_to_long: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gama, ref, long_desc, _price, short_desc = (row + (None,) * 5)[:5]
        long_desc = str(long_desc).strip() if long_desc else ""
        short_desc = str(short_desc).strip() if short_desc else ""
        ref = str(ref).strip() if ref is not None else ""
        if not long_desc:
            continue
        if short_desc and _norm(short_desc) not in {_norm(k) for k in short_to_long}:
            short_to_long[short_desc] = long_desc
        if ref and ref not in ref_to_long:
            ref_to_long[ref] = long_desc
    return short_to_long, ref_to_long


def _load_gama_difusor_bindings(ws) -> dict[str, set[str]]:
    """Return {gama_name: {long_diffuser_name, ...}} from the sheet."""
    bindings: dict[str, set[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gama, _, long_desc = (row + (None,) * 3)[:3]
        gama_s = str(gama).strip() if gama else ""
        long_desc_s = str(long_desc).strip() if long_desc else ""
        if gama_s and long_desc_s:
            bindings.setdefault(gama_s.upper(), set()).add(long_desc_s)
    return bindings


def _resolve_name(
    current_name: str,
    short_to_long: dict[str, str],
    ref_to_long: dict[str, str],
    efficiency: dict[str, float],
) -> tuple[str | None, float | None]:
    """Try to find the long description for a current difusor name.

    Returns (new_name, eficiencia) or (None, None) if no mapping found.
    """
    n = _norm(current_name)
    # Try exact match in efficiency (already a long name)
    if n in efficiency:
        return current_name, efficiency[n]
    # Try short desc → long desc
    if n in {_norm(k) for k in short_to_long}:
        for short, long in short_to_long.items():
            if _norm(short) == n:
                eff = efficiency.get(_norm(long))
                return long, eff
    # Try ref code → long desc
    if n in {_norm(k) for k in ref_to_long}:
        for ref, long in ref_to_long.items():
            if _norm(ref) == n:
                eff = efficiency.get(_norm(long))
                return long, eff
    # Maybe current name IS the long description but with different case
    for eff_name in efficiency:
        if _norm(eff_name) == n:
            return eff_name, efficiency[eff_name]
    return None, None


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Show what would change without applying.")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    efis = _load_efficiency_map(wb[SHEET_EFI])
    short_to_long, ref_to_long = _load_gama_mapping(wb[SHEET_GAMA])
    _load_gama_difusor_bindings(wb[SHEET_GAMA])

    db = SessionLocal()
    try:
        all_difusores = db.query(Difusor).order_by(Difusor.id).all()
        print(f"Current difusores in DB: {len(all_difusores)}")
        print(f"Efficiencies from Excel: {len(efis)}")
        print(f"Short->long mappings: {len(short_to_long)}")
        print(f"Ref->long mappings: {len(ref_to_long)}")

        renamed = 0
        eff_added = 0
        deleted = 0

        for d in all_difusores:
            new_name, eff = _resolve_name(d.name, short_to_long, ref_to_long, efis)
            if new_name is None:
                print(f"  ??? {d.id:3d} {d.name!r:30s} — SKIP (no mapping)")
                continue

            # Check if another difusor already has this name (merge case)
            existing = db.query(Difusor).filter(Difusor.name == new_name, Difusor.id != d.id).first()
            if existing is not None:
                # Merge: update all FK references from d to existing, then delete d
                print(f"  MERGE {d.id:3d} {d.name!r:30s} -> {new_name!r:30s} (into id={existing.id})")
                if not args.dry_run:
                    for tbl, fk_col in [
                        (Fotometria, "difusor_id"),
                        (ValidCombination, "difusor_id"),
                        (LuminaireLED, "difusor_id"),
                    ]:
                        (db.query(tbl).filter(getattr(tbl, fk_col) == d.id).update(
                            {fk_col: existing.id}, synchronize_session=False
                        ))
                    if eff is not None:
                        existing.eficiencia = eff
                    db.flush()
                    db.delete(d)
                    deleted += 1
                continue

            changes = []
            if d.name != new_name:
                changes.append(f"name: {d.name!r} -> {new_name!r}")
            if d.eficiencia != eff:
                changes.append(f"eficiencia: {d.eficiencia} -> {eff}")

            if changes:
                print(f"  UPDATE {d.id:3d} {d.name!r:30s} — {'; '.join(changes)}")
                if not args.dry_run:
                    d.name = new_name
                    d.eficiencia = eff
                    db.flush()
                    renamed += 1
                    if eff is not None:
                        eff_added += 1

        # Add new difusores from gama difusor sheet that don't exist in DB
        existing_names = {_norm(d.name) for d in db.query(Difusor).all()}
        all_long_descs = set(short_to_long.values()) | set(ref_to_long.values())
        for long_desc in sorted(all_long_descs):
            if _norm(long_desc) not in existing_names:
                eff = efis.get(_norm(long_desc))
                print(f"  INSERT {long_desc!r:30s} eficiencia={eff}")
                if not args.dry_run:
                    db.add(Difusor(name=long_desc, eficiencia=eff))

        if not args.dry_run:
            db.commit()
            print(f"\nApplied: {renamed} renamed, {eff_added} with eficiencia, {deleted} merged/removed.")
        else:
            print(f"\nDry-run: {renamed} to rename, {eff_added} to get eficiencia, {deleted} to merge/remove.")
    except Exception:
        if not args.dry_run:
            db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
