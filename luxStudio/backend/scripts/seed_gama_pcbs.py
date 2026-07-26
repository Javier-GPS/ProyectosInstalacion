"""Seed gama_pcbs from motor_configurador_v7 Excel (PCB dropdown per gama)."""
from __future__ import annotations
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.core.text_utils import norm
from app.database import SessionLocal
from app.models import Gama, GamaPCB, PCB
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent.parent / "docs" / "motor_configurador_v7 2.xlsx"


def main():
    wb = load_workbook(str(XLSX), data_only=True)
    ws = wb["Configurador"]

    db = SessionLocal()

    gama_cols = {}
    for c in range(7, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            gama_cols[str(v).strip()] = c

    all_pcbs = {norm(p.pcb_descripcion): p for p in db.query(PCB).all() if p.pcb_descripcion}
    gama_map = {norm(g.name): g for g in db.query(Gama).all()}
    added = 0
    skipped = 0
    missing_pcb = 0
    missing_gama = 0

    for gama_name, col in gama_cols.items():
        gama = gama_map.get(norm(gama_name))
        if not gama:
            print(f"  SKIP gama={gama_name!r} (not in DB)")
            missing_gama += 1
            continue

        for r in range(10, 17):
            desc = ws.cell(r, col).value
            if not desc:
                continue
            desc = str(desc).strip()
            if not desc:
                continue

            pcb = all_pcbs.get(norm(desc))
            if not pcb:
                print(f"  SKIP {gama_name} PCB desc={desc!r} (not in pcbs table)")
                missing_pcb += 1
                continue

            exists = (
                db.query(GamaPCB)
                .filter(GamaPCB.gama_id == gama.id, GamaPCB.pcb_id == pcb.id)
                .first()
            )
            if exists:
                skipped += 1
                continue

            db.add(GamaPCB(gama_id=gama.id, pcb_id=pcb.id))
            added += 1

    db.commit()
    db.close()
    print(f"Added {added} gama_pcbs, skipped {skipped} (dups), missing_gama={missing_gama}, missing_pcb_desc={missing_pcb}")


if __name__ == "__main__":
    main()
