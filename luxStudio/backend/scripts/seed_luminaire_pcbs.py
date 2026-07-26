"""Seed n_pcbs / n_leds_per_pcb / pcb_id into luminaire_leds from Excel (bulk)."""
from __future__ import annotations
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.core.text_utils import norm
from app.database import SessionLocal
from app.models import Gama, Difusor, Lente, LedType, LuminaireLED, PCB
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent.parent / "docs" / "Referencias_productos_pcb_go.xlsx"


def main():
    wb = load_workbook(str(XLSX), read_only=True, data_only=True)

    ws_gf = wb["GO que faltan"]
    go_data = {}
    for row in ws_gf.iter_rows(min_row=2, values_only=True):
        go = row[0]
        if not go: continue
        go_data[norm(go)] = {
            "n_pcbs": int(row[5]) if row[5] is not None and str(row[5]).isdigit() else None,
            "n_leds_per_pcb": int(row[6]) if row[6] is not None and str(row[6]).isdigit() else None,
            "pcb_desc": str(row[7]).strip().lower() if row[7] else "",
        }

    ws_pc = wb["SALVI Lighting$Param_ Configura"]
    pcb_desc_ref = {}
    for row in ws_pc.iter_rows(min_row=2, values_only=True):
        if row[144] and row[141]:
            ref = str(row[144]).strip()
            desc = str(row[141]).strip().lower()
            pcb_desc_ref[desc] = ref

    ws_v = wb["Variantes SALVI"]
    variants = []
    for row in ws_v.iter_rows(min_row=2, values_only=True):
        gama_v = norm(row[2])
        dif_v = norm(row[6] or row[7])
        lente_v = norm(row[14])
        led_type_v = norm(row[11])
        go_v = norm(row[17])
        if gama_v and dif_v and lente_v and go_v:
            variants.append((gama_v, dif_v, lente_v, led_type_v, go_v))

    wb.close()
    print(f"Total variants: {len(variants)}")

    db = SessionLocal()
    try:
        gama_map = {norm(g.name): g.id for g in db.query(Gama).all()}
        dif_map = {norm(d.name): d.id for d in db.query(Difusor).all()}
        lent_map = {norm(l.name): l.id for l in db.query(Lente).all()}
        lt_map = {norm(t.name): t.id for t in db.query(LedType).all()}
        pcb_by_ref = {norm(p.pcb_ref): p.id for p in db.query(PCB).all()}

        # Build all LuminaireLED by combo key
        leds_by_combo = {}
        for ll in db.query(LuminaireLED).all():
            key = (ll.gama_id, ll.difusor_id, ll.lente_id, ll.led_type_id)
            leds_by_combo.setdefault(key, []).append(ll)

        updated = 0
        not_found_go = 0
        matched_pcb = 0

        for gv, dv, lv, ltv, go_v in variants:
            gid = gama_map.get(gv)
            did = dif_map.get(dv)
            lid = lent_map.get(lv)
            ltid = lt_map.get(ltv) if ltv else None
            if not (gid and did and lid):
                continue

            key = (gid, did, lid, ltid)
            lls = leds_by_combo.get(key, [])
            if not lls:
                continue

            gd = go_data.get(go_v)
            if not gd:
                not_found_go += 1
                continue

            for ll in lls:
                ll.n_pcbs = gd["n_pcbs"]
                ll.n_leds_per_pcb = gd["n_leds_per_pcb"]
                pcb_desc = gd["pcb_desc"]
                if pcb_desc and pcb_desc in pcb_desc_ref:
                    ref = pcb_desc_ref[pcb_desc]
                    pcb_id = pcb_by_ref.get(norm(ref))
                    if pcb_id:
                        ll.pcb_id = pcb_id
                        matched_pcb += 1
                updated += 1

        db.commit()
        print(f"Updated: {updated} luminaire_leds rows")
        print(f"GO not found: {not_found_go}")
        print(f"PCB matched: {matched_pcb}")
        print(f"GO entries: {len(go_data)}, PCB descs: {len(pcb_desc_ref)}")

        # Verify
        with_p = db.query(LuminaireLED).filter(LuminaireLED.pcb_id.isnot(None)).count()
        with_n = db.query(LuminaireLED).filter(LuminaireLED.n_pcbs.isnot(None)).count()
        print(f"Verify - n_pcbs set: {with_n}, pcb_id set: {with_p}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
